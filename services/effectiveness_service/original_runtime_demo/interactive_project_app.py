# -*- coding: utf-8 -*-
"""Workbook-driven interactive A/B parameter comparison application.

Run from the workspace root:
    python .\compare\demo\interactive_project_app.py

Open:
    http://127.0.0.1:8776
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import random
import shutil
import statistics
import sys
import zipfile
from dataclasses import asdict
from datetime import datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import urlparse

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from scipy.stats import qmc


DEMO_DIR = Path(__file__).resolve().parent
if str(DEMO_DIR) not in sys.path:
    sys.path.insert(0, str(DEMO_DIR))

from project_excel import (  # noqa: E402
    AttributeSpec,
    ProjectDataError,
    ProjectDataset,
    load_project_workbook,
    write_generation_ranges,
)
from coupling_model import CouplingSystem  # noqa: E402
from feasibility_model import ExpertFeasibilityModel  # noqa: E402
from requirement_model import RequirementEvaluator  # noqa: E402
from preference_models import (  # noqa: E402
    DEFAULT_UTA_SEGMENTS,
    MAX_UTA_SEGMENTS,
    MIN_UTA_SEGMENTS,
    LPUTAModel,
    OnlineBTModel,
    validate_segment_count,
)


DEFAULT_WORKBOOK = DEMO_DIR / "data" / "aircraft_door_lock_demo.xlsx"
DEFAULT_STATE_DIR = DEMO_DIR / "interactive_project"
DEFAULT_PORT = 8776
PROFILE_VERSION = 10
LEGACY_REQUIREMENT_PROFILE_ID = "legacy_default"

STATUS_LABELS = {
    "likely_feasible_learned": "较可能实现",
    "likely_feasible_extrapolation": "较可能实现，但属于经验外推",
    "uncertain_feasibility": "还需要专家确认",
    "likely_infeasible_learned": "可能无法实现",
    "learned_expert_boundary": "超出专家已确认的可行边界",
    "likely_feasible_by_coupling": "参数搭配较合理",
    "coupling_boundary": "接近当前可行边界",
    "coupling_outside_experience": "参数搭配需要确认",
    "outside_sample_experience": "超出已有样本范围，等待确认",
    "infeasible_by_range": "有参数超出允许范围",
}
FEASIBILITY_LABELS = {"feasible": "可以实现", "infeasible": "不能实现", "uncertain": "暂时看不准"}
PREFERENCE_LABELS = {"A": "方案 A 更好", "B": "方案 B 更好", "tie": "两个方案差不多", "unknown": "暂时无法比较"}
SOURCE_LABELS = {
    "existing_sample": "Excel 已有样本",
    "coupling_consistent": "单向耦合带内生成",
    "coupling_boundary_probe": "单向耦合边界试探",
    "global_space_fill": "全局空间填充探索",
    "maximin_exploration": "远离已探索区域",
    "multiscale_local": "优胜方案多尺度搜索",
    "directional_coupling_probe": "方向耦合控制变量试探",
}


def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def clamp(value: float, lo: float, hi: float) -> float:
    return max(lo, min(hi, value))


def json_response(handler: BaseHTTPRequestHandler, payload: Any, status: int = 200) -> None:
    data = json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "application/json; charset=utf-8")
    handler.send_header("Content-Length", str(len(data)))
    handler.end_headers()
    handler.wfile.write(data)


def html_response(handler: BaseHTTPRequestHandler, html: str) -> None:
    data = html.encode("utf-8")
    handler.send_response(200)
    handler.send_header("Content-Type", "text/html; charset=utf-8")
    handler.send_header("Content-Length", str(len(data)))
    handler.end_headers()
    handler.wfile.write(data)


def binary_response(
    handler: BaseHTTPRequestHandler,
    data: bytes,
    content_type: str,
    filename: str,
) -> None:
    handler.send_response(200)
    handler.send_header("Content-Type", content_type)
    handler.send_header("Content-Disposition", f'attachment; filename="{filename}"')
    handler.send_header("Content-Length", str(len(data)))
    handler.end_headers()
    handler.wfile.write(data)


def read_body(handler: BaseHTTPRequestHandler) -> Dict[str, Any]:
    length = int(handler.headers.get("Content-Length", "0"))
    if length <= 0:
        return {}
    return json.loads(handler.rfile.read(length).decode("utf-8"))


class ProjectApp:
    def __init__(self, workbook_path: Path | str, state_dir: Path | str = DEFAULT_STATE_DIR, seed: int = 20260710):
        self.seed = int(seed)
        self.state_root = Path(state_dir).expanduser().resolve()
        self.configure_project(load_project_workbook(workbook_path))

    def configure_project(self, project: ProjectDataset) -> None:
        self.project = project
        self.state_path = self.state_root / f"state_{self.project.learning_fingerprint}.json"
        self.attributes = self.project.attributes
        self.attribute_by_key = self.project.attribute_by_key()
        self.coupling_system = CouplingSystem.fit(self.project)
        self.prior_feasibility_model = ExpertFeasibilityModel(self.project, self.coupling_system)
        self.reason_options = self._build_reason_options()

    def _build_reason_options(self) -> Dict[str, Dict[str, str]]:
        options = {
            "single_range": {
                "title": "某个参数超出允许范围",
                "description": "某个参数低于或高于当前允许范围。",
            }
        }
        seen_targets: set[str] = set()
        for edge in self.project.couplings:
            if edge.target_key in seen_targets:
                continue
            seen_targets.add(edge.target_key)
            options[f"coupling_{edge.target_key}"] = {
                "title": f"{edge.target_label}与其他参数搭配不合理",
                "description": f"当前{edge.target_label}与其他参数组合可能无法同时实现。",
            }
        options["other"] = {"title": "其他原因", "description": "请在下方补充具体问题。"}
        return options

    def param_profile(self) -> Dict[str, Dict[str, Any]]:
        return {
            item.key: {
                "key": item.key,
                "label": item.label,
                "unit": item.unit,
                "data_type": item.data_type,
                "design_stage": item.design_stage,
                "precision": item.precision,
                "min": item.generation_min,
                "max": item.generation_max,
                "feasible_min": item.feasible_min,
                "feasible_max": item.feasible_max,
                "preference_direction": item.preference_direction,
                "marginal_trend": item.marginal_trend,
                "participates_utility": item.participates_utility,
                "participates_generation": item.participates_generation,
                "description": item.description,
                "inferred_fields": item.inferred_fields,
                "scale_profile": self.adaptive_scale_profile(item),
                "range_semantics": "生成范围是允许探索的外层边界；经验范围来自已有样本最小值和最大值",
            }
            for item in self.attributes
        }

    def existing_schemes(self) -> List[Dict[str, Any]]:
        return [
            {
                "id": item.id,
                "source": "existing_sample",
                "source_detail": item.source,
                "known_feasibility": item.known_feasibility,
                "params": dict(item.params),
            }
            for item in self.project.schemes
        ]

    @staticmethod
    def automatic_scale_mode(spec: AttributeSpec) -> str:
        """Choose a unit-free coordinate system without asking the expert."""
        if spec.data_type == "integer":
            return "discrete"
        if not spec.is_numeric:
            return "categorical"
        assert spec.generation_min is not None and spec.generation_max is not None
        if spec.generation_min > 0 and spec.generation_max / max(spec.generation_min, 1e-12) >= 4.0:
            return "log"
        return "linear"

    def normalize_coordinate(self, value: Any, spec: AttributeSpec) -> float:
        assert spec.generation_min is not None and spec.generation_max is not None
        numeric = float(value)
        mode = self.automatic_scale_mode(spec)
        if mode == "log":
            lo = math.log(max(float(spec.generation_min), 1e-12))
            hi = math.log(max(float(spec.generation_max), 1e-12))
            return clamp((math.log(max(numeric, 1e-12)) - lo) / max(hi - lo, 1e-12), 0.0, 1.0)
        return clamp(
            (numeric - float(spec.generation_min))
            / max(float(spec.generation_max) - float(spec.generation_min), 1e-12),
            0.0,
            1.0,
        )

    def denormalize_coordinate(self, coordinate: float, spec: AttributeSpec) -> Any:
        assert spec.generation_min is not None and spec.generation_max is not None
        z = clamp(float(coordinate), 0.0, 1.0)
        if self.automatic_scale_mode(spec) == "log":
            lo = math.log(max(float(spec.generation_min), 1e-12))
            hi = math.log(max(float(spec.generation_max), 1e-12))
            value = math.exp(lo + z * (hi - lo))
        else:
            value = float(spec.generation_min) + z * (
                float(spec.generation_max) - float(spec.generation_min)
            )
        return int(round(value)) if spec.data_type == "integer" else round(value, spec.precision)

    def adaptive_scale_profile(self, spec: AttributeSpec) -> Dict[str, Any]:
        if not spec.is_numeric:
            return {"mode": "categorical", "local_radius": None, "radii": []}
        coordinates = sorted(
            {round(self.normalize_coordinate(item.params[spec.key], spec), 10) for item in self.project.schemes}
        )
        gaps = [right - left for left, right in zip(coordinates, coordinates[1:]) if right - left > 1e-8]
        typical_gap = statistics.median(gaps) if gaps else 0.10
        local_radius = clamp(0.75 * typical_gap, 0.035, 0.12)
        radii = [
            clamp(0.55 * local_radius, 0.025, 0.08),
            clamp(1.35 * local_radius, 0.06, 0.18),
            clamp(2.75 * local_radius, 0.14, 0.35),
        ]
        return {
            "mode": self.automatic_scale_mode(spec),
            "local_radius": round(local_radius, 4),
            "radii": [round(value, 4) for value in radii],
            "sample_unique_values": len(coordinates),
        }

    def normalized_distance(self, a: Dict[str, Any], b: Dict[str, Any]) -> float:
        values: List[float] = []
        for spec in self.attributes:
            if spec.is_numeric:
                values.append(self.normalize_coordinate(a[spec.key], spec) - self.normalize_coordinate(b[spec.key], spec))
            else:
                values.append(0.0 if a[spec.key] == b[spec.key] else 1.0)
        return math.sqrt(sum(value * value for value in values) / max(len(values), 1))

    def global_novelty(self, params: Dict[str, Any], state: Dict[str, Any]) -> float:
        distances = [
            self.normalized_distance(params, item["params"])
            for item in state.get("known_schemes", [])
            if item.get("params")
        ]
        return min(distances, default=1.0)

    def coverage_scarcity(self, params: Dict[str, Any], state: Dict[str, Any], bins: int = 6) -> float:
        scores: List[float] = []
        known = state.get("known_schemes", [])
        for spec in self.attributes:
            if not spec.is_numeric or not spec.participates_generation:
                continue
            target_bin = min(bins - 1, int(self.normalize_coordinate(params[spec.key], spec) * bins))
            occupied = sum(
                1
                for item in known
                if min(bins - 1, int(self.normalize_coordinate(item["params"][spec.key], spec) * bins)) == target_bin
            )
            scores.append(1.0 / (1.0 + occupied))
        return sum(scores) / max(len(scores), 1)

    def query_signature(self, base: Dict[str, Any], candidate: Dict[str, Any]) -> str:
        """Describe the engineering question independently of units and scheme ids."""
        target_keys = self.coupling_system.target_keys()
        tokens: List[str] = []
        for spec in self.attributes:
            if not spec.is_numeric or spec.key in target_keys or not spec.participates_generation:
                continue
            before = self.normalize_coordinate(base["params"][spec.key], spec)
            after = self.normalize_coordinate(candidate["params"][spec.key], spec)
            delta = after - before
            if abs(delta) < 0.018:
                continue
            direction = "+" if delta > 0.0 else "-"
            position_bin = min(7, max(0, int(after * 8.0)))
            tokens.append(f"{spec.key}:{direction}:{position_bin}")
        explanation = candidate.get("generation_explanation") or {}
        probe_target = explanation.get("probe_target")
        probe_side = explanation.get("probe_side")
        if probe_target:
            tokens.append(f"probe:{probe_target}:{probe_side or 'direction'}")
        return "|".join(tokens) if tokens else "no_upstream_change"

    @staticmethod
    def seen_query_signatures(state: Dict[str, Any]) -> Dict[str, int]:
        counts: Dict[str, int] = {}
        for scheme in state.get("known_schemes", []):
            signature = (scheme.get("generation_explanation") or {}).get("query_signature")
            if signature:
                counts[signature] = counts.get(signature, 0) + 1
        return counts

    @staticmethod
    def recent_effective_yield(state: Dict[str, Any], window: int = 20) -> Optional[float]:
        recent = state.get("interactions", [])[-window:]
        if not recent:
            return None
        effective = sum(
            1
            for item in recent
            if item.get("feasibility_a") == "feasible"
            and item.get("feasibility_b") == "feasible"
            and item.get("preference") in {"A", "B", "tie"}
        )
        return effective / len(recent)

    def normalized_benefit(self, params: Dict[str, Any], spec: AttributeSpec) -> Optional[float]:
        if not spec.is_numeric or not spec.participates_utility:
            return None
        assert spec.generation_min is not None and spec.generation_max is not None
        value = float(params[spec.key])
        position = clamp((value - spec.generation_min) / max(spec.generation_max - spec.generation_min, 1e-9), 0.0, 1.0)
        if spec.preference_direction == "higher_better":
            return position
        if spec.preference_direction == "lower_better":
            return 1.0 - position
        if spec.preference_direction == "neutral":
            return None
        return 0.5

    def active_requirement_profile(self, state: Optional[Dict[str, Any]]) -> Any:
        profile_id = (state or {}).get("active_protocol_profile_id") or (state or {}).get("active_requirement_profile_id")
        if profile_id and profile_id != LEGACY_REQUIREMENT_PROFILE_ID:
            return self.project.requirement_profile_by_id().get(profile_id)
        if profile_id == LEGACY_REQUIREMENT_PROFILE_ID:
            return None
        return self.project.default_requirement_profile()

    def active_requirement_profile_id(self, state: Optional[Dict[str, Any]]) -> str:
        profile = self.active_requirement_profile(state)
        return profile.id if profile is not None else LEGACY_REQUIREMENT_PROFILE_ID

    def preference_evidence_for_context(self, state: Dict[str, Any]) -> List[Dict[str, Any]]:
        # Expert A/B preferences learn a protocol-independent product model.
        return OnlineBTModel.active_evidence(state.get("preference_evidence", []))

    def feasibility_model_from_state(self, state: Optional[Dict[str, Any]]) -> ExpertFeasibilityModel:
        weights = (state or {}).get("feasibility_weights") or self.prior_feasibility_model.prior_weights()
        active_evidence = [
            item
            for item in (state or {}).get("feasibility_evidence", [])
            if item.get("status", "active") == "active"
        ]
        frontier = [
            item
            for item in active_evidence
            if item.get("coupling_feedback")
        ]
        return ExpertFeasibilityModel(
            self.project,
            self.coupling_system,
            weights=weights,
            frontier_evidence=frontier,
            expert_evidence=active_evidence,
        )

    def bt_model_from_state(self, state: Optional[Dict[str, Any]]) -> OnlineBTModel:
        return OnlineBTModel(
            self.project,
            weights=(state or {}).get("bt_weights"),
            requirement_profile=None,
        )

    def uta_model_from_state(self, state: Optional[Dict[str, Any]]) -> LPUTAModel:
        state = state or {}
        segments = validate_segment_count(state.get("uta_segments", DEFAULT_UTA_SEGMENTS))
        return LPUTAModel(
            self.project,
            segments=segments,
            increments=state.get("uta_increments"),
            requirement_profile=None,
        )

    def retrain_preference_models(self, state: Dict[str, Any]) -> Tuple[OnlineBTModel, LPUTAModel]:
        evidence = self.preference_evidence_for_context(state)
        bt_model = OnlineBTModel(self.project, requirement_profile=None)
        bt_summary = bt_model.fit(evidence)
        uta_model = LPUTAModel(
            self.project,
            segments=state.get("uta_segments", DEFAULT_UTA_SEGMENTS),
            requirement_profile=None,
        )
        uta_summary = uta_model.fit(evidence)
        for summary in (bt_summary, uta_summary):
            summary["learning_context"] = "generic_product_effectiveness"
            summary["generic_preference_pairs"] = len(evidence)
            summary["protocol_used_for_training"] = False
        state["bt_weights"] = list(bt_summary["weights"])
        state["bt_model"] = bt_summary
        state["uta_increments"] = list(uta_summary["increments"])
        state["uta_model"] = uta_summary
        used_ids = {item.get("id") for item in evidence}
        for item in state.get("preference_evidence", []):
            if item.get("id") in used_ids and item.get("status") == "pending_bt_uta":
                item["status"] = "used_bt_uta"
        self.sync_preference_reviews(state)
        return bt_model, uta_model

    @staticmethod
    def classify_feasibility_status(
        probability: float,
        *,
        hard_violation: bool = False,
        mature_boundary_violation: bool = False,
        outside_sample_experience: bool = False,
        coupling_outside: bool = False,
        coupling_severity: float = 0.0,
    ) -> str:
        """Keep feasibility probability separate from evidence-coverage warnings."""
        if hard_violation:
            return "infeasible_by_range"
        if mature_boundary_violation or probability < 0.35:
            return "likely_infeasible_learned"
        if probability < 0.65:
            return "uncertain_feasibility"
        if coupling_outside and (probability < 0.75 or coupling_severity >= 0.65):
            return "uncertain_feasibility"
        if outside_sample_experience or coupling_outside:
            return "likely_feasible_extrapolation"
        return "likely_feasible_learned"

    def evaluate(self, params: Dict[str, Any], state: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        state = state or {}
        uta_summary = state.get("uta_model", {})
        bt_summary = state.get("bt_model", {})
        if uta_summary.get("evidence_pairs") and uta_summary.get("attribute_weights"):
            protocol_weights = uta_summary.get("attribute_weights")
            protocol_weight_source = "通用LP-UTA学习权重"
        elif bt_summary.get("training_pairs") and bt_summary.get("attribute_weights"):
            protocol_weights = bt_summary.get("attribute_weights")
            protocol_weight_source = "通用BT学习权重"
        else:
            protocol_weights = None
            protocol_weight_source = "设计批次50/30/20先验"
        requirement_assessment = RequirementEvaluator(
            self.project,
            self.active_requirement_profile(state),
            attribute_weights=protocol_weights,
            weight_source=protocol_weight_source,
        ).evaluate(params)
        violations: List[Dict[str, Any]] = []
        edge_distances: List[float] = []
        experience_extrapolations: List[Dict[str, Any]] = []
        contributions: List[Dict[str, Any]] = []
        benefits: List[float] = []
        for spec in self.attributes:
            value = params[spec.key]
            if spec.is_numeric:
                numeric = float(value)
                assert spec.generation_min is not None and spec.generation_max is not None
                assert spec.feasible_min is not None and spec.feasible_max is not None
                generation_span = max(spec.generation_max - spec.generation_min, 1e-9)
                if numeric < spec.generation_min:
                    severity = clamp((spec.generation_min - numeric) / generation_span, 0.05, 1.0)
                    violations.append(
                        {
                            "code": f"range_low_{spec.key}",
                            "title": f"{spec.label}低于生成下限",
                            "attribute": spec.key,
                            "message": f"当前值 {numeric:g}，生成下限 {spec.generation_min:g} {spec.unit}。",
                            "severity": round(severity, 3),
                        }
                    )
                elif numeric > spec.generation_max:
                    severity = clamp((numeric - spec.generation_max) / generation_span, 0.05, 1.0)
                    violations.append(
                        {
                            "code": f"range_high_{spec.key}",
                            "title": f"{spec.label}高于生成上限",
                            "attribute": spec.key,
                            "message": f"当前值 {numeric:g}，生成上限 {spec.generation_max:g} {spec.unit}。",
                            "severity": round(severity, 3),
                        }
                    )
                if numeric < spec.feasible_min:
                    experience_extrapolations.append(
                        {
                            "attribute": spec.key,
                            "label": spec.label,
                            "side": "below",
                            "message": (
                                f"{spec.label}低于已有样本最小值 {spec.feasible_min:g} {spec.unit}，"
                                "属于有意探索区域，不等于物理不可行。"
                            ),
                        }
                    )
                elif numeric > spec.feasible_max:
                    experience_extrapolations.append(
                        {
                            "attribute": spec.key,
                            "label": spec.label,
                            "side": "above",
                            "message": (
                                f"{spec.label}高于已有样本最大值 {spec.feasible_max:g} {spec.unit}，"
                                "属于有意探索区域，不等于物理不可行。"
                            ),
                        }
                    )
                inside_distance = min(numeric - spec.generation_min, spec.generation_max - numeric) / generation_span
                edge_distances.append(clamp(inside_distance, 0.0, 0.5) * 2.0)
            benefit = self.normalized_benefit(params, spec)
            if benefit is not None:
                benefits.append(benefit)
                contributions.append(
                    {
                        "feature": spec.key,
                        "label": spec.label,
                        "value": round(benefit, 3),
                        "contribution": round(benefit / max(sum(1 for item in self.attributes if item.participates_utility), 1), 3),
                    }
                )

        baseline_score = (
            requirement_assessment.coverage_percent
            if requirement_assessment.profile_id is not None
            else 100.0 * sum(benefits) / max(len(benefits), 1)
        )
        coupling_assessments = self.coupling_system.assess(params)
        outside = [item for item in coupling_assessments if item.status in {"below_band", "above_band"}]
        near_boundary = [item for item in coupling_assessments if item.status == "near_boundary"]
        if violations:
            physics_status = "infeasible_by_range"
            physics_risk = max(float(item["severity"]) for item in violations)
        elif outside:
            physics_status = "coupling_outside_experience"
            physics_risk = clamp(0.42 + 0.34 * max(item.severity for item in outside), 0.42, 0.78)
        elif near_boundary:
            physics_status = "coupling_boundary"
            physics_risk = 0.32
        elif experience_extrapolations:
            physics_status = "outside_sample_experience"
            physics_risk = 0.24
        else:
            physics_status = "likely_feasible_by_coupling"
            physics_risk = 0.10
        contributions.sort(key=lambda item: item["contribution"], reverse=True)
        if outside:
            coupling_status = f"{len(outside)} 个下游属性超出已有样本条件经验带，需要专家确认。"
        elif near_boundary:
            coupling_status = f"{len(near_boundary)} 个下游属性接近条件经验边界。"
        elif self.coupling_system.direction_only:
            coupling_status = (
                f"{len(self.coupling_system.direction_only)} 个耦合目标目前只有方向先验，"
                "将通过专家可行性判断逐步学习强度。"
            )
        else:
            coupling_status = f"{len(coupling_assessments)} 个耦合目标均位于已有样本条件经验带内。"
        feasibility_model = self.feasibility_model_from_state(state)
        feasibility_probability = feasibility_model.probability(params)
        learned_boundary_violations = feasibility_model.boundary_violations(params)
        mature_boundary_violations = [item for item in learned_boundary_violations if item["mature"]]
        if mature_boundary_violations and not violations:
            physics_status = "learned_expert_boundary"
            physics_risk = max(
                physics_risk,
                clamp(0.58 + max(item["severity"] for item in mature_boundary_violations) * 0.22, 0.58, 0.92),
            )
        bt_model = self.bt_model_from_state(state)
        uta_model = self.uta_model_from_state(state)
        bt_pairs = int(bt_summary.get("training_pairs", 0))
        uta_status = uta_summary.get("status", "no_data")
        bt_score = bt_model.score(params)
        uta_score = uta_model.score(params)
        validation_scores = []
        for increments in uta_summary.get("validation_increments", []):
            validation_model = LPUTAModel(
                self.project,
                segments=state.get("uta_segments", DEFAULT_UTA_SEGMENTS),
                increments=increments,
                requirement_profile=None,
            )
            validation_scores.append(validation_model.score(params))
        if len(validation_scores) >= 2:
            ordered_scores = sorted(validation_scores)
            low_index = max(0, int(math.floor(0.10 * (len(ordered_scores) - 1))))
            high_index = min(len(ordered_scores) - 1, int(math.ceil(0.90 * (len(ordered_scores) - 1))))
            uta_score_interval = [round(ordered_scores[low_index], 2), round(ordered_scores[high_index], 2)]
        else:
            uta_score_interval = None
        if requirement_assessment.profile_id is not None:
            effectiveness_score = requirement_assessment.coverage_percent
            effectiveness_source = f"新技术协议100分相对评分（{protocol_weight_source}）"
            effectiveness_status = "protocol_relative"
        elif uta_status in {"preliminary", "validated", "needs_review"}:
            effectiveness_score = uta_score
            effectiveness_source = "LP-UTA"
            effectiveness_status = uta_status
        elif bt_pairs:
            effectiveness_score = bt_score
            effectiveness_source = "在线 BT 趋势"
            effectiveness_status = "insufficient_data"
        else:
            effectiveness_score = baseline_score
            effectiveness_source = "属性方向范围基线"
            effectiveness_status = "no_preference_data"
        coupling_severity = max([float(item.severity) for item in outside] or [0.0])
        status = self.classify_feasibility_status(
            feasibility_probability,
            hard_violation=bool(violations),
            mature_boundary_violation=bool(mature_boundary_violations),
            outside_sample_experience=bool(experience_extrapolations),
            coupling_outside=bool(outside),
            coupling_severity=coupling_severity,
        )
        risk = max(physics_risk, 1.0 - feasibility_probability)
        feasibility_summary = (state or {}).get("feasibility_model", {})
        expert_feasibility_samples = int(feasibility_summary.get("expert_samples", 0))
        if not expert_feasibility_samples:
            feasibility_confidence = "低：主要来自已有样本弱先验和耦合经验带"
        elif expert_feasibility_samples < 8:
            feasibility_confidence = "初步：专家证据较少"
        elif feasibility_summary.get("positive_samples", 0) and feasibility_summary.get("negative_samples", 0):
            feasibility_confidence = "中等：已包含正负专家证据"
        else:
            feasibility_confidence = "初步：专家证据类别仍不完整"
        if status == "likely_feasible_extrapolation":
            feasibility_confidence += "；预测概率较高，但当前组合超出部分样本或耦合经验覆盖"
        generalization_status = uta_summary.get("generalization_status", "insufficient_evidence")
        effectiveness_confidence = {
            "whole_scheme_supported": "较高：整方案留出与交叉验证提供支持",
            "pair_validation_only": "中等：只有比较对层面的验证",
            "unstable_on_holdout": "较低：留出验证结果不稳定",
            "insufficient_evidence": "低：有效偏好证据不足",
        }.get(generalization_status, "低：尚未形成独立验证结论")
        assessment_basis = [
            f"{len(self.project.schemes)} 个 Excel 已有方案形成经验分布",
            f"{len(self.project.couplings)} 条有向耦合关系和 {len(self.coupling_system.models)} 个条件代理",
            f"{expert_feasibility_samples} 条专家可行性证据",
            f"{bt_pairs} 条有效偏好用于 BT，{uta_summary.get('training_pairs', 0)} 条用于 UTA 训练",
        ]
        return {
            "status": status,
            "risk": round(risk, 3),
            "physics_status": physics_status,
            "physics_risk": round(physics_risk, 3),
            "learned_feasibility_probability": round(feasibility_probability, 3),
            "feasibility_risk_contributors": feasibility_model.risk_contributors(params),
            "learned_boundary_violations": learned_boundary_violations,
            "baseline_score": round(baseline_score, 2),
            "bt_score": round(bt_score, 2),
            "bt_training_pairs": bt_pairs,
            "uta_score": round(uta_score, 2) if uta_summary.get("evidence_pairs", 0) else None,
            "uta_score_interval": uta_score_interval,
            "uta_contributions": (
                uta_model.attribute_contributions(params) if uta_summary.get("evidence_pairs", 0) else []
            ),
            "uta_status": uta_status,
            "generalization_status": uta_summary.get("generalization_status", "insufficient_evidence"),
            "feasibility_confidence": feasibility_confidence,
            "effectiveness_confidence": effectiveness_confidence,
            "assessment_basis": assessment_basis,
            "effectiveness_score": round(effectiveness_score, 2),
            "effectiveness_source": effectiveness_source,
            "effectiveness_status": effectiveness_status,
            "score_label": f"{effectiveness_source}（{effectiveness_status}）",
            "hard_violations": violations,
            "experience_extrapolations": experience_extrapolations,
            "top_contributors": contributions[:5],
            "coupling_status": coupling_status,
            "coupling_assessments": [asdict(item) for item in coupling_assessments],
            "requirement_assessment": requirement_assessment.to_dict(),
            "reuse_recommendation": (
                "物理可行性尚未通过，暂不建议复用"
                if status in {"infeasible_by_range", "likely_infeasible_learned"}
                else requirement_assessment.decision_label
            ),
        }

    def feasibility_training_samples(self, state: Dict[str, Any]) -> List[Dict[str, Any]]:
        samples = [
            {
                "params": dict(item["params"]),
                "label": 1.0,
                "weight": 0.12,
                "source": "existing_weak_positive",
                "reason_codes": [],
            }
            for item in state.get("existing_samples", [])
        ]
        for observation in state.get("feasibility_evidence", []):
            if observation.get("status", "active") != "active":
                continue
            if observation.get("label") not in {"feasible", "infeasible"}:
                continue
            reason_bonus = 0.18 if observation.get("reason_codes") else 0.0
            guided_codes = list(observation.get("reason_codes") or [])
            attribute_feedback = observation.get("attribute_feedback") or {}
            if attribute_feedback.get("attribute_key") and attribute_feedback.get("side") in {"low", "high"}:
                guided_codes.append(
                    f"attribute_{attribute_feedback['side']}_{attribute_feedback['attribute_key']}"
                )
            samples.append(
                {
                    "params": dict(observation["params"]),
                    "label": 1.0 if observation["label"] == "feasible" else 0.0,
                    "weight": (1.0 if observation["label"] == "feasible" else 1.40) + reason_bonus,
                    "source": "expert",
                    "reason_codes": guided_codes,
                }
            )
        return samples

    def retrain_feasibility_model(self, state: Dict[str, Any]) -> ExpertFeasibilityModel:
        active_evidence = [
            item
            for item in state.get("feasibility_evidence", [])
            if item.get("status", "active") == "active"
        ]
        frontier = [
            item
            for item in active_evidence
            if item.get("coupling_feedback")
        ]
        model = ExpertFeasibilityModel(
            self.project,
            self.coupling_system,
            frontier_evidence=frontier,
            expert_evidence=active_evidence,
        )
        summary = model.fit(self.feasibility_training_samples(state))
        summary["frontier_evidence"] = len(frontier)
        state["feasibility_weights"] = dict(model.weights)
        state["feasibility_model"] = summary
        return model

    def parse_params(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        params: Dict[str, Any] = {}
        for spec in self.attributes:
            if spec.key not in payload:
                raise ProjectDataError(f"缺少属性“{spec.label}”。")
            raw = payload[spec.key]
            if spec.data_type == "continuous":
                try:
                    value: Any = float(raw)
                except (TypeError, ValueError) as exc:
                    raise ProjectDataError(f"属性“{spec.label}”必须是数值。") from exc
                if not math.isfinite(value):
                    raise ProjectDataError(f"属性“{spec.label}”必须是有限数值。")
            elif spec.data_type == "integer":
                try:
                    value = int(raw)
                except (TypeError, ValueError) as exc:
                    raise ProjectDataError(f"属性“{spec.label}”必须是整数。") from exc
            else:
                value = str(raw).strip()
                if not value:
                    raise ProjectDataError(f"属性“{spec.label}”不能为空。")
            params[spec.key] = value
        return params

    def round_and_clip(self, params: Dict[str, Any]) -> Dict[str, Any]:
        output: Dict[str, Any] = {}
        for spec in self.attributes:
            value = params[spec.key]
            if spec.is_numeric:
                assert spec.generation_min is not None and spec.generation_max is not None
                numeric = clamp(float(value), spec.generation_min, spec.generation_max)
                output[spec.key] = int(round(numeric)) if spec.data_type == "integer" else round(numeric, spec.precision)
            else:
                output[spec.key] = value
        return output

    @staticmethod
    def clip_feasible(value: float, spec: AttributeSpec) -> float:
        assert spec.feasible_min is not None and spec.feasible_max is not None
        return clamp(value, spec.feasible_min, spec.feasible_max)

    def generate_candidate(
        self,
        state: Dict[str, Any],
        base: Dict[str, Any],
        mode: Optional[str] = None,
        seed_offset: int = 0,
        commit: bool = True,
        normalized_vector: Optional[Dict[str, float]] = None,
    ) -> Dict[str, Any]:
        index = int(state.get("generated_index", 1))
        rng = random.Random(
            self.seed + index * 1009 + len(state.get("interactions", [])) * 37 + int(seed_offset) * 7919
        )
        params = dict(base["params"])
        target_keys = self.coupling_system.target_keys()
        design_trace: Dict[int, List[str]] = {1: [], 2: [], 3: []}
        ordered_attributes = sorted(enumerate(self.attributes), key=lambda item: (item[1].design_stage, item[0]))
        selected_mode = mode or "multiscale_local"
        if selected_mode == "coupling_boundary_probe" and not self.coupling_system.models:
            selected_mode = "directional_coupling_probe"
        probe_target: Optional[str] = None
        probe_side: Optional[str] = None
        probe_source: Optional[str] = None

        if selected_mode in {"global_space_fill", "maximin_exploration"}:
            explorable = [
                spec
                for _, spec in ordered_attributes
                if spec.participates_generation and spec.is_numeric and spec.key not in target_keys
            ]
            coordinates = {
                spec.key: (
                    normalized_vector.get(spec.key, rng.random())
                    if normalized_vector is not None
                    else rng.random()
                )
                for spec in explorable
            }
            ranked_specs = sorted(
                explorable,
                key=lambda spec: abs(
                    coordinates[spec.key] - self.normalize_coordinate(base["params"][spec.key], spec)
                ),
                reverse=True,
            )
            change_count = min(len(ranked_specs), 2 + ((index + seed_offset) % 3 == 0))
            for spec in ranked_specs[:change_count]:
                params[spec.key] = self.denormalize_coordinate(coordinates[spec.key], spec)
                design_trace[spec.design_stage].append(spec.label)
        elif selected_mode == "directional_coupling_probe" and target_keys:
            target_key = sorted(target_keys)[(index + seed_offset) % len(target_keys)]
            probe_target = target_key
            definition = self.coupling_system.probe_definition(target_key)
            sources = definition["sources"]
            source = sources[(index + seed_offset) % len(sources)]
            probe_source = source["key"]
            source_spec = self.attribute_by_key[source["key"]]
            target_spec = self.attribute_by_key[target_key]
            source_z = self.normalize_coordinate(params[source_spec.key], source_spec)
            target_z = self.normalize_coordinate(params[target_key], target_spec)
            direction = 1.0 if source["direction"] == "positive" else -1.0
            source_step = (0.12, 0.20, 0.30)[seed_offset % 3]
            compensation = (0.0, 0.08, 0.20)[(seed_offset // 3) % 3]
            params[source_spec.key] = self.denormalize_coordinate(source_z + source_step, source_spec)
            params[target_key] = self.denormalize_coordinate(target_z + direction * compensation, target_spec)
            design_trace[source_spec.design_stage].append(source_spec.label)
            design_trace[target_spec.design_stage].append(target_spec.label)
        else:
            radius_index = seed_offset % 3
            candidates = [
                item
                for item in self.attributes
                if item.is_numeric and item.participates_generation and item.key not in target_keys
            ]
            change_count = min(len(candidates), 1 + ((index + seed_offset) % 3))
            selected_specs = rng.sample(candidates, change_count) if change_count else []
            for spec in selected_specs:
                profile = self.adaptive_scale_profile(spec)
                radius = profile["radii"][radius_index]
                current = self.normalize_coordinate(params[spec.key], spec)
                delta = rng.gauss(0.0, radius)
                if abs(delta) < 0.35 * radius:
                    delta = radius if rng.random() >= 0.5 else -radius
                params[spec.key] = self.denormalize_coordinate(current + delta, spec)
                design_trace[spec.design_stage].append(spec.label)

        target_details: Dict[str, Dict[str, Any]] = {}
        for target_key in sorted(target_keys):
            if selected_mode == "directional_coupling_probe" and target_key == probe_target:
                continue
            model = self.coupling_system.models.get(target_key)
            if model is not None:
                spec = self.attribute_by_key[target_key]
                band = model.band(params)
                assert spec.generation_min is not None and spec.generation_max is not None
                lower = max(band["lower"], spec.generation_min)
                upper = min(band["upper"], spec.generation_max)
                if lower > upper:
                    chosen = clamp(band["predicted"], spec.generation_min, spec.generation_max)
                else:
                    width = max(upper - lower, 1e-9)
                    if spec.preference_direction == "lower_better":
                        chosen = lower + rng.uniform(0.22, 0.54) * width
                    elif spec.preference_direction == "higher_better":
                        chosen = lower + rng.uniform(0.46, 0.78) * width
                    else:
                        chosen = lower + rng.uniform(0.30, 0.70) * width
                params[target_key] = chosen
                design_trace[spec.design_stage].append(spec.label)
                target_details[target_key] = {
                    "target_key": target_key,
                    "target_label": model.target_label,
                    "predicted": round(band["predicted"], spec.precision),
                    "lower": round(band["lower"], spec.precision),
                    "upper": round(band["upper"], spec.precision),
                    "chosen": round(chosen, spec.precision),
                    "unit": spec.unit,
                    "mode": "inside",
                    "model_status": "fitted_monotonic",
                }
                continue

            spec = self.attribute_by_key[target_key]
            base_coordinate = self.normalize_coordinate(base["params"][target_key], spec)
            adjustment = 0.0
            source_changes: List[Dict[str, Any]] = []
            for edge in self.coupling_system.edges_by_target.get(target_key, []):
                source_spec = self.attribute_by_key[edge.source_key]
                before = self.normalize_coordinate(base["params"][edge.source_key], source_spec)
                after = self.normalize_coordinate(params[edge.source_key], source_spec)
                source_delta = after - before
                signed_delta = source_delta if edge.direction == "positive" else -source_delta
                contribution = 0.22 * signed_delta
                adjustment += contribution
                if abs(source_delta) > 1e-8:
                    source_changes.append(
                        {
                            "source_key": edge.source_key,
                            "source_label": source_spec.label,
                            "direction": edge.direction,
                            "normalized_change": round(source_delta, 4),
                            "target_adjustment": round(contribution, 4),
                        }
                    )
            chosen_coordinate = clamp(base_coordinate + adjustment, 0.0, 1.0)
            chosen = self.denormalize_coordinate(chosen_coordinate, spec)
            params[target_key] = chosen
            if source_changes:
                design_trace[spec.design_stage].append(spec.label)
            target_details[target_key] = {
                "target_key": target_key,
                "target_label": spec.label,
                "chosen": round(float(chosen), spec.precision),
                "unit": spec.unit,
                "mode": "directional_completion",
                "model_status": "direction_only",
                "source_changes": source_changes,
                "normalized_adjustment": round(adjustment, 4),
            }

        if selected_mode == "coupling_boundary_probe" and self.coupling_system.models:
            model_keys = sorted(self.coupling_system.models)
            probe_target = model_keys[(index // 3) % len(model_keys)]
            model = self.coupling_system.models[probe_target]
            spec = self.attribute_by_key[probe_target]
            band = model.band(params)
            assert spec.generation_min is not None and spec.generation_max is not None
            span = float(spec.generation_max) - float(spec.generation_min)
            low_space = band["lower"] - spec.generation_min
            high_space = spec.generation_max - band["upper"]
            preferred_side = "below" if index % 2 else "above"
            if preferred_side == "below" and low_space > 0.015 * span:
                probe_side = "below"
            elif high_space > 0.015 * span:
                probe_side = "above"
            elif low_space > 0.0:
                probe_side = "below"
            else:
                probe_side = "inside_edge"
            if probe_side == "below":
                band_width = max(band["upper"] - band["lower"], 1e-9)
                distance = min(0.025 * span, max(0.008 * span, 0.12 * band_width))
                params[probe_target] = max(spec.generation_min, band["lower"] - distance)
            elif probe_side == "above":
                band_width = max(band["upper"] - band["lower"], 1e-9)
                distance = min(0.025 * span, max(0.008 * span, 0.12 * band_width))
                params[probe_target] = min(spec.generation_max, band["upper"] + distance)
            else:
                params[probe_target] = band["lower"] + 0.04 * max(band["upper"] - band["lower"], 1e-9)
            target_details[probe_target]["chosen"] = round(float(params[probe_target]), spec.precision)
            target_details[probe_target]["mode"] = probe_side

        rounded = self.round_and_clip(params)
        predicted_feasibility = self.feasibility_model_from_state(state).probability(rounded)
        changed_sources = []
        for spec in self.attributes:
            if spec.key in target_keys or not spec.is_numeric:
                continue
            delta = float(rounded[spec.key]) - float(base["params"][spec.key])
            if abs(delta) > 10 ** (-(spec.precision + 1)):
                changed_sources.append(f"{spec.label}{delta:+.{spec.precision}f}{spec.unit}")
        if selected_mode == "directional_coupling_probe" and target_keys:
            probe_target = sorted(target_keys)[(index + seed_offset) % len(target_keys)]
            definition = self.coupling_system.probe_definition(probe_target)
            probe_source = definition["sources"][(index + seed_offset) % len(definition["sources"])]["key"]
            target_label = self.attribute_by_key[probe_target].label
            source_label = self.attribute_by_key[probe_source].label
            purpose = f"用控制变量确认{source_label}对{target_label}的影响"
            detail = (
                f"本轮只重点改变{source_label}及{target_label}的补偿量；专家只需判断能否实现，"
                "若不能实现可直接确认系统建议的问题项。"
            )
        elif selected_mode == "coupling_boundary_probe" and probe_target:
            probe = target_details[probe_target]
            purpose = f"确认{probe['target_label']}条件经验边界"
            detail = (
                f"先扰动上游属性，再根据单调代理得到 {probe['target_label']} 经验带 "
                f"[{probe['lower']}, {probe['upper']}] {probe['unit']}；本轮取 {probe['chosen']}，"
                f"只重点试探这一条耦合边界。"
            )
        elif selected_mode in {"global_space_fill", "maximin_exploration"}:
            purpose = "用控制变量方式探索历史方案未覆盖的区域"
            detail = (
                "先用单位无关采样定位远端目标，本轮只改变 2 到 3 个上游属性，"
                "再按已知因果方向补全下游属性，便于专家判断。"
            )
        elif selected_mode == "multiscale_local":
            purpose = "在当前优胜方案附近同时进行小、中、大尺度搜索"
            detail = "各属性先自动归一化，再按样本疏密自适应选择扰动半径。"
        else:
            purpose = "生成耦合物理一致方案"
            summaries = "；".join(
                (
                    f"{item['target_label']}={item['chosen']}，经验带[{item['lower']},{item['upper']}]"
                    if "lower" in item
                    else f"{item['target_label']}={item['chosen']}，按方向先验补全"
                )
                for item in target_details.values()
            )
            detail = f"先改变上游属性，再把下游属性放入条件经验带。{summaries}。"
        candidate = {
            "id": f"G-{index:03d}",
            "source": selected_mode,
            "source_detail": (
                f"基于 {base['id']} 生成"
                if selected_mode not in {"global_space_fill", "maximin_exploration"}
                else "基于 Excel 生成范围进行全域探索"
            ),
            "params": rounded,
            "generation_explanation": {
                "purpose": purpose,
                "detail": detail,
                "base_scheme_id": base["id"],
                "mode": selected_mode,
                "probe_target": probe_target,
                "probe_source": probe_source,
                "probe_side": probe_side,
                "changed_sources": changed_sources,
                "target_bands": target_details,
                "design_sequence": [
                    {"stage": stage, "attributes": design_trace[stage]}
                    for stage in (1, 2, 3)
                    if design_trace[stage]
                ],
                "predicted_feasibility": round(predicted_feasibility, 3),
            },
        }
        if commit:
            state["generated_index"] = index + 1
        return candidate

    def generation_mode(self, state: Dict[str, Any]) -> str:
        index = int(state.get("generated_index", 1))
        expert_count = sum(
            1
            for item in state.get("feasibility_evidence", [])
            if item.get("source") == "expert" and item.get("status") == "active"
        )
        probe = (
            "coupling_boundary_probe"
            if self.coupling_system.models
            else "directional_coupling_probe"
            if self.coupling_system.target_keys()
            else "global_space_fill"
        )
        recent_yield = self.recent_effective_yield(state)
        if expert_count >= 20 and recent_yield is not None and recent_yield < 0.45:
            schedule = [
                "multiscale_local",
                "coupling_consistent",
                "multiscale_local",
                probe,
                "global_space_fill",
                "multiscale_local",
                "coupling_consistent",
                "maximin_exploration",
            ]
            return schedule[(index - 1) % len(schedule)]
        if expert_count < 10:
            schedule = [
                "global_space_fill",
                "maximin_exploration",
                probe,
                "global_space_fill",
                "maximin_exploration",
                "multiscale_local",
                "global_space_fill",
                probe,
                "maximin_exploration",
                "multiscale_local",
            ]
        elif expert_count < 30:
            schedule = [
                "global_space_fill",
                "multiscale_local",
                "maximin_exploration",
                probe,
                "multiscale_local",
                "global_space_fill",
                "multiscale_local",
                "maximin_exploration",
            ]
        else:
            schedule = [
                "multiscale_local",
                "global_space_fill",
                "multiscale_local",
                probe,
                "multiscale_local",
                "maximin_exploration",
                "multiscale_local",
                "global_space_fill",
            ]
        return schedule[(index - 1) % len(schedule)]

    def select_candidate(self, state: Dict[str, Any], base: Dict[str, Any], pool_size: int = 192) -> Dict[str, Any]:
        """Choose a candidate from a mode-specific exploration or exploitation pool."""
        index = int(state.get("generated_index", 1))
        mode = self.generation_mode(state)
        bt_model = self.bt_model_from_state(state)
        ranked: List[Tuple[float, Dict[str, Any], Dict[str, float]]] = []
        seen_signatures = self.seen_query_signatures(state)
        count = max(24, int(pool_size))
        numeric_specs = [item for item in self.attributes if item.is_numeric and item.participates_generation]
        vectors: List[Optional[Dict[str, float]]] = [None] * count
        if mode in {"global_space_fill", "maximin_exploration"} and numeric_specs:
            sampler = qmc.LatinHypercube(
                d=len(numeric_specs),
                seed=self.seed + index * 101 + len(state.get("interactions", [])) * 17,
            )
            matrix = sampler.random(n=count)
            vectors = [
                {spec.key: float(value) for spec, value in zip(numeric_specs, row)}
                for row in matrix
            ]
        for offset, vector in enumerate(vectors):
            candidate = self.generate_candidate(
                state,
                base,
                mode=mode,
                seed_offset=offset,
                commit=False,
                normalized_vector=vector,
            )
            evaluation = self.evaluate(candidate["params"], state)
            probability_b = 1.0 - bt_model.probability(base["params"], candidate["params"])
            uncertainty = bt_model.uncertainty(base["params"], candidate["params"])
            novelty = self.global_novelty(candidate["params"], state)
            base_distance = self.normalized_distance(candidate["params"], base["params"])
            coverage = self.coverage_scarcity(candidate["params"], state)
            feasibility = float(evaluation["learned_feasibility_probability"])
            boundary_violations = [
                item for item in evaluation.get("learned_boundary_violations", []) if item.get("mature")
            ]
            quality = float(evaluation["bt_score"]) / 100.0
            query_signature = self.query_signature(base, candidate)
            repeat_count = seen_signatures.get(query_signature, 0)
            if evaluation["hard_violations"]:
                objective = -10.0
            elif mode == "maximin_exploration":
                objective = 0.45 * novelty + 0.17 * coverage + 0.08 * uncertainty + 0.30 * feasibility
            elif mode == "global_space_fill":
                objective = 0.34 * novelty + 0.22 * coverage + 0.12 * uncertainty + 0.32 * feasibility
            elif mode in {"coupling_boundary_probe", "directional_coupling_probe"}:
                objective = 0.42 * uncertainty + 0.25 * novelty + 0.20 * feasibility + 0.13 * quality
            else:
                objective = (
                    0.46 * quality
                    + 0.20 * probability_b
                    + 0.16 * feasibility
                    + 0.10 * novelty
                    + 0.08 * uncertainty
                )
            if boundary_violations:
                boundary_severity = max(float(item.get("severity", 0.0)) for item in boundary_violations)
                if mode in {"coupling_boundary_probe", "directional_coupling_probe"}:
                    objective -= 0.80 * boundary_severity
                else:
                    objective = min(objective, -1.50 - 3.00 * boundary_severity)
            objective -= min(1.20, 0.38 * repeat_count)
            candidate["generation_explanation"]["query_signature"] = query_signature
            ranked.append(
                (
                    objective,
                    candidate,
                    {
                        "bt_score": float(evaluation["bt_score"]),
                        "probability_b_better": probability_b,
                        "uncertainty": uncertainty,
                        "feasibility": feasibility,
                        "novelty": novelty,
                        "base_distance": base_distance,
                        "coverage": coverage,
                        "repeat_count": float(repeat_count),
                    },
                )
            )
        objective, selected, metrics = max(ranked, key=lambda item: item[0])
        selected["generation_explanation"].update(
            {
                "candidate_pool_size": len(ranked),
                "selection_objective": round(objective, 4),
                "predicted_bt_score": round(metrics["bt_score"], 2),
                "probability_b_better": round(metrics["probability_b_better"], 3),
                "bt_uncertainty": round(metrics["uncertainty"], 3),
                "normalized_novelty": round(metrics["novelty"], 3),
                "distance_to_base": round(metrics["base_distance"], 3),
                "coverage_scarcity": round(metrics["coverage"], 3),
                "repeated_question_count": int(metrics["repeat_count"]),
                "recent_effective_yield": (
                    None
                    if self.recent_effective_yield(state) is None
                    else round(float(self.recent_effective_yield(state)), 3)
                ),
                "selection_strategy": (
                    "优先选择距离全部已见方案最远的候选"
                    if mode == "maximin_exploration"
                    else "优先填补生成空间中覆盖不足的区域"
                    if mode == "global_space_fill"
                    else "优先选择能缩小耦合可行前沿不确定性的候选"
                    if mode in {"coupling_boundary_probe", "directional_coupling_probe"}
                    else "在自动多尺度邻域中优先寻找可能更好的候选"
                ),
            }
        )
        state["generated_index"] = index + 1
        return selected

    def _initial_state(self) -> Dict[str, Any]:
        existing = self.existing_schemes()
        feasible = [item for item in existing if not self.evaluate(item["params"])["hard_violations"]]
        generic_reference = OnlineBTModel(self.project, requirement_profile=None)
        current = max(feasible or existing, key=lambda item: generic_reference.score(item["params"]))
        state: Dict[str, Any] = {
            "version": 1,
            "profile_version": PROFILE_VERSION,
            "project_name": self.project.project_name,
            "workbook_path": self.project.workbook_path,
            "workbook_fingerprint": self.project.workbook_fingerprint,
            "learning_fingerprint": self.project.learning_fingerprint,
            "created_at": now_iso(),
            "updated_at": now_iso(),
            "data_mode": "live_expert",
            "demo_profile": None,
            "active_protocol_profile_id": (
                self.project.default_requirement_profile().id
                if self.project.default_requirement_profile() is not None
                else LEGACY_REQUIREMENT_PROFILE_ID
            ),
            "generated_index": 1,
            "existing_samples": existing,
            "known_schemes": list(existing),
            "current_best_id": current["id"],
            "active_pair": None,
            "interactions": [],
            "feasibility_evidence": [],
            "preference_evidence": [],
            "review_evidence": [],
            "review_history": [],
            "feasibility_weights": {},
            "feasibility_model": {},
            "uta_segments": DEFAULT_UTA_SEGMENTS,
            "bt_weights": [],
            "bt_model": {},
            "uta_increments": [],
            "uta_model": {},
            "unknown_evaluations": [],
            "range_update_history": [],
            "learning_events": [
                {
                    "id": "E-001",
                    "created_at": now_iso(),
                    "type": "project_loaded",
                    "title": "Excel 项目已载入",
                    "detail": (
                        f"识别 {len(self.attributes)} 个属性、{len(existing)} 个已有方案和 {len(self.project.couplings)} 条耦合边；"
                        f"已拟合 {len(self.coupling_system.models)} 个单调条件代理，"
                        f"{len(self.coupling_system.direction_only)} 个目标以方向先验模式启动。"
                    ),
                }
            ],
        }
        self.retrain_feasibility_model(state)
        self.retrain_preference_models(state)
        candidate = self.select_candidate(state, current)
        state["known_schemes"].append(candidate)
        state["active_pair"] = {"a": current["id"], "b": candidate["id"]}
        self.save_state(state)
        return state

    def prepare_demo_state(self, preference_count: int = 60) -> Dict[str, Any]:
        """Build a deterministic, clearly labeled teacher-demo state."""
        state = self._initial_state()
        state["data_mode"] = "fixed_demo_simulation"
        state["demo_profile"] = {
            "name": "阶段 8 固定展示数据",
            "source": "系统模拟专家，仅用于演示，不得视为真实专家结论",
            "prepared_at": now_iso(),
            "seed": self.seed,
        }

        demo_interactions: List[Dict[str, Any]] = []
        for index, scheme in enumerate(state["existing_samples"][:6], start=1):
            interaction_id = f"D-F-{index:03d}"
            evidence = self.record_feasibility_evidence(
                state, interaction_id, scheme, "A", "feasible", [], "固定展示数据中的模拟可行判断"
            )
            if evidence:
                evidence["source"] = "fixed_demo_simulation"
                evidence["expert_id"] = "demo_simulated_expert"

        risk_schemes: List[Dict[str, Any]] = []
        generic_reference = OnlineBTModel(self.project, requirement_profile=None)
        base = max(state["existing_samples"], key=lambda item: generic_reference.score(item["params"]))
        for index, target_key in enumerate(sorted(self.coupling_system.target_keys())[:2], start=1):
            spec = self.attribute_by_key[target_key]
            params = dict(base["params"])
            assert spec.feasible_min is not None
            params[target_key] = int(round(spec.feasible_min)) if spec.data_type == "integer" else spec.feasible_min
            probe = {
                "id": f"DEMO-RISK-{index:02d}",
                "source": "coupling_boundary_probe",
                "source_detail": "固定展示数据中的耦合不匹配样本",
                "params": self.round_and_clip(params),
            }
            state["known_schemes"].append(probe)
            risk_schemes.append(probe)
            interaction_id = f"D-F-{index + 6:03d}"
            evidence = self.record_feasibility_evidence(
                state,
                interaction_id,
                probe,
                "B",
                "infeasible",
                [f"coupling_{target_key}"],
                f"固定展示：{spec.label}与上游属性组合不匹配",
            )
            if evidence:
                evidence["source"] = "fixed_demo_simulation"
                evidence["expert_id"] = "demo_simulated_expert"
        self.retrain_feasibility_model(state)

        reference_model = LPUTAModel(
            self.project,
            segments=DEFAULT_UTA_SEGMENTS,
            requirement_profile=None,
        )
        pair_candidates: List[Tuple[Dict[str, Any], Dict[str, Any], str, float]] = []
        schemes = list(state["existing_samples"])
        for left_index, scheme_a in enumerate(schemes):
            for scheme_b in schemes[left_index + 1 :]:
                difference = reference_model.utility(scheme_a["params"]) - reference_model.utility(scheme_b["params"])
                if abs(difference) < 0.012:
                    continue
                pair_candidates.append(
                    (scheme_a, scheme_b, "A" if difference > 0 else "B", abs(difference))
                )
        rng = random.Random(self.seed + 808)
        rng.shuffle(pair_candidates)
        selected_pairs = pair_candidates[: min(max(12, int(preference_count)), len(pair_candidates))]
        for index, (scheme_a, scheme_b, relation, _) in enumerate(selected_pairs, start=1):
            interaction_id = f"D-P-{index:03d}"
            evidence = self.record_preference_evidence(
                state, interaction_id, scheme_a, scheme_b, "feasible", "feasible", relation
            )
            if evidence:
                evidence["source"] = "fixed_demo_simulation"
                evidence["expert_id"] = "demo_simulated_expert"
                evidence["confidence"] = 0.92
            demo_interactions.append(
                {
                    "id": interaction_id,
                    "created_at": now_iso(),
                    "scheme_a": scheme_a["id"],
                    "scheme_b": scheme_b["id"],
                    "feasibility_a": "feasible",
                    "feasibility_b": "feasible",
                    "preference": relation,
                    "source": "fixed_demo_simulation",
                }
            )
        state["interactions"] = demo_interactions
        self.retrain_preference_models(state)

        if selected_pairs:
            scheme_a, scheme_b, relation, _ = selected_pairs[0]
            inverse = "B" if relation == "A" else "A"
            conflict = self.record_preference_evidence(
                state, "D-P-CONFLICT", scheme_a, scheme_b, "feasible", "feasible", inverse
            )
            if conflict:
                conflict["source"] = "fixed_demo_simulation"
                conflict["expert_id"] = "demo_simulated_expert"
                conflict["confidence"] = 0.92
            self.retrain_preference_models(state)
            self.save_state(state)
            review = next(
                (
                    item
                    for item in state["review_evidence"]
                    if item.get("type") == "preference_pair_conflict" and item.get("status") == "open"
                ),
                None,
            )
            if review:
                self.review_action({"reviewId": review["id"], "action": "withdraw"})
                state = self.load_state()
                for remaining in [
                    item["id"]
                    for item in state.get("review_evidence", [])
                    if item.get("status") == "open"
                ]:
                    self.review_action({"reviewId": remaining, "action": "keep"})
                state = self.load_state()

        active_evidence = OnlineBTModel.active_evidence(state["preference_evidence"])
        current = max(state["existing_samples"], key=lambda item: reference_model.utility(item["params"]))
        state["current_best_id"] = current["id"]
        candidate = self.select_candidate(state, current)
        state["known_schemes"].append(candidate)
        state["active_pair"] = {"a": current["id"], "b": candidate["id"]}
        state["unknown_evaluations"] = []
        for index, scheme in enumerate([current, *risk_schemes], start=1):
            state["unknown_evaluations"].append(
                {
                    "id": f"DEMO-U-{index:02d}",
                    "created_at": now_iso(),
                    "params": dict(scheme["params"]),
                    "result": self.evaluate(scheme["params"], state),
                    "source": "fixed_demo_simulation",
                }
            )
        state["demo_profile"].update(
            {
                "active_preference_pairs": len(active_evidence),
                "feasibility_evidence": len(state["feasibility_evidence"]),
                "resolved_reviews": len(state["review_history"]),
            }
        )
        state["learning_events"].append(
            {
                "id": f"E-{len(state['learning_events']) + 1:03d}",
                "created_at": now_iso(),
                "type": "fixed_demo_loaded",
                "title": "固定展示数据已载入",
                "detail": (
                    f"已载入 {len(active_evidence)} 条模拟有效偏好、{len(state['feasibility_evidence'])} 条模拟可行性证据，"
                    "并保留一次已处理的冲突复核历史。所有证据均标记为模拟来源。"
                ),
            }
        )
        self.save_state(state)
        return self.summarize(state)

    def backup_incompatible_state(
        self,
        state: Dict[str, Any],
        source_path: Optional[Path] = None,
    ) -> Optional[Path]:
        source = source_path or self.state_path
        if not source.exists():
            return None
        version = str(state.get("profile_version", "unknown"))
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup = source.with_name(
            f"{source.stem}.backup_profile_{version}_{timestamp}{source.suffix}"
        )
        shutil.copy2(source, backup)
        return backup

    def legacy_state_matches_project(self, state: Dict[str, Any]) -> bool:
        """Conservatively recognize pre-v10 state after a protocol-only workbook edit."""
        stored = state.get("existing_samples") or []
        if len(stored) != len(self.project.schemes):
            return False
        stored_by_id = {str(item.get("id")): item.get("params") for item in stored}
        return all(
            stored_by_id.get(item.id) == item.params
            for item in self.project.schemes
        )

    def state_source_path(self) -> Optional[Path]:
        if self.state_path.exists():
            return self.state_path
        exact_legacy = self.state_root / f"state_{self.project.workbook_fingerprint}.json"
        if exact_legacy.exists():
            return exact_legacy
        if not self.state_root.exists():
            return None
        candidates = sorted(
            (
                path
                for path in self.state_root.glob("state_*.json")
                if ".backup_" not in path.name
            ),
            key=lambda path: path.stat().st_mtime,
            reverse=True,
        )
        for path in candidates:
            try:
                state = json.loads(path.read_text(encoding="utf-8"))
            except (OSError, json.JSONDecodeError):
                continue
            stored_path = state.get("workbook_path")
            if not stored_path:
                continue
            try:
                same_workbook_path = Path(stored_path).resolve() == Path(self.project.workbook_path).resolve()
            except OSError:
                same_workbook_path = False
            if not same_workbook_path:
                continue
            if self.legacy_state_matches_project(state):
                return path
        return None

    def load_state(self, reset: bool = False) -> Dict[str, Any]:
        source_path = None if reset else self.state_source_path()
        if reset or source_path is None:
            return self._initial_state()
        try:
            state = json.loads(source_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            return self._initial_state()
        stored_learning_fingerprint = state.get("learning_fingerprint")
        if (
            stored_learning_fingerprint
            and stored_learning_fingerprint != self.project.learning_fingerprint
        ) or (
            not stored_learning_fingerprint
            and not self.legacy_state_matches_project(state)
        ):
            self.backup_incompatible_state(state, source_path)
            return self._initial_state()
        protocol_only_change = state.get("workbook_fingerprint") != self.project.workbook_fingerprint
        state["workbook_path"] = self.project.workbook_path
        state["workbook_fingerprint"] = self.project.workbook_fingerprint
        state["learning_fingerprint"] = self.project.learning_fingerprint
        if state.get("profile_version") in {7, 8, 9}:
            previous_profile = state.get("profile_version")
            state["profile_version"] = PROFILE_VERSION
            previous_protocol = state.get("active_protocol_profile_id") or state.get("active_requirement_profile_id")
            state.setdefault(
                "active_protocol_profile_id",
                previous_protocol
                if previous_protocol in self.project.requirement_profile_by_id()
                else self.project.default_requirement_profile().id
                if self.project.default_requirement_profile() is not None
                else LEGACY_REQUIREMENT_PROFILE_ID,
            )
            for item in state.get("preference_evidence", []):
                item["learning_context"] = "generic_product_effectiveness"
            state.setdefault("range_update_history", [])
            state.setdefault("learning_events", []).append(
                {
                    "id": f"E-{len(state.get('learning_events', [])) + 1:03d}",
                    "created_at": now_iso(),
                    "type": "protocol_decoupling_upgrade",
                    "title": "新技术协议已与学习模型解耦",
                    "detail": (
                        f"已从状态版本 {previous_profile} 升级；原专家判断全部保留，"
                        "BT/UTA现按通用产品效能统一训练，新技术协议只作为100分评分参考。"
                    ),
                }
            )
            self.retrain_feasibility_model(state)
            self.retrain_preference_models(state)
            self.save_state(state)
        elif state.get("profile_version") != PROFILE_VERSION:
            self.backup_incompatible_state(state)
            return self._initial_state()
        state.setdefault("feasibility_evidence", [])
        state.setdefault("preference_evidence", [])
        state.setdefault("review_evidence", [])
        state.setdefault("review_history", [])
        state.setdefault("uta_segments", DEFAULT_UTA_SEGMENTS)
        state.setdefault("data_mode", "live_expert")
        state.setdefault("demo_profile", None)
        state.setdefault("range_update_history", [])
        state.setdefault(
            "active_protocol_profile_id",
            self.project.default_requirement_profile().id
            if self.project.default_requirement_profile() is not None
            else LEGACY_REQUIREMENT_PROFILE_ID,
        )
        if state.get("active_protocol_profile_id") not in self.project.requirement_profile_by_id():
            state["active_protocol_profile_id"] = (
                self.project.default_requirement_profile().id
                if self.project.default_requirement_profile() is not None
                else LEGACY_REQUIREMENT_PROFILE_ID
            )
        for item in state.get("preference_evidence", []):
            item.setdefault("learning_context", "generic_product_effectiveness")
        if not state.get("feasibility_weights"):
            self.retrain_feasibility_model(state)
            self.save_state(state)
        if not state.get("bt_model") or not state.get("uta_model"):
            self.retrain_preference_models(state)
            self.save_state(state)
        if protocol_only_change or source_path != self.state_path:
            state.setdefault("learning_events", []).append(
                {
                    "id": f"E-{len(state.get('learning_events', [])) + 1:03d}",
                    "created_at": now_iso(),
                    "type": "protocol_workbook_changed",
                    "title": "新技术协议已更新，学习状态继续沿用",
                    "detail": (
                        "检测到属性、历史方案和耦合知识未改变；已保留全部专家判断及BT/UTA参数，"
                        "只更新100分评分参考。"
                    ),
                }
            )
            self.save_state(state)
        return state

    def save_state(self, state: Dict[str, Any]) -> None:
        state["workbook_path"] = self.project.workbook_path
        state["workbook_fingerprint"] = self.project.workbook_fingerprint
        state["learning_fingerprint"] = self.project.learning_fingerprint
        state["updated_at"] = now_iso()
        self.state_path.parent.mkdir(parents=True, exist_ok=True)
        self.state_path.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")

    @staticmethod
    def find_scheme(state: Dict[str, Any], scheme_id: str) -> Optional[Dict[str, Any]]:
        for item in state.get("known_schemes", []):
            if item.get("id") == scheme_id:
                return item
        return None

    def learning_completion(self, state: Dict[str, Any]) -> Dict[str, Any]:
        uta = state.get("uta_model", {})
        evidence_pairs = int(uta.get("evidence_pairs", 0))
        exact_pairs = evidence_pairs
        warm_pairs = 0
        validation_values: List[float] = []
        for value in (
            uta.get("test_accuracy"),
            (uta.get("cross_validation") or {}).get("accuracy"),
            (uta.get("whole_scheme_holdout") or {}).get("accuracy"),
        ):
            if value is not None:
                validation_values.append(float(value))
        validation_accuracy = min(validation_values) if validation_values else None
        consistency_value = uta.get("judgment_consistency")
        if consistency_value is None:
            consistency_value = uta.get("training_accuracy")
        if consistency_value is None and evidence_pairs:
            high_tolerance = int(uta.get("high_tolerance_pairs", 0))
            consistency_value = clamp(1.0 - high_tolerance / max(evidence_pairs, 1), 0.0, 1.0)
        consistency = float(consistency_value) if consistency_value is not None else None
        minimum_pairs = max(20, 2 * int(uta.get("attribute_count", len(self.attributes))))
        recommended_pairs = max(minimum_pairs, int(uta.get("recommended_training_pairs", minimum_pairs)))
        open_reviews = sum(1 for item in state.get("review_evidence", []) if item.get("status") == "open")
        reasons: List[str] = []
        if exact_pairs < minimum_pairs:
            reasons.append(f"通用效能有效比较 {exact_pairs} 对，至少需要 {minimum_pairs} 对")
        if validation_accuracy is None:
            reasons.append("还没有形成可用的独立验证正确率，目标为 80%")
        elif validation_accuracy < 0.80:
            reasons.append(f"保守验证正确率 {validation_accuracy:.0%}，未达到 80%")
        if consistency is None:
            reasons.append("还没有足够判断来检查一致性，目标为 80%")
        elif consistency < 0.80 or not uta.get("m1_consistent", False):
            reasons.append(f"判断一致度约 {consistency:.0%}，或 M1 一致性检查未通过")
        if open_reviews:
            reasons.append(f"还有 {open_reviews} 项判断待复核")

        standard_complete = not reasons
        high_confidence = bool(
            standard_complete
            and validation_accuracy is not None
            and validation_accuracy >= 0.90
            and consistency is not None
            and consistency >= 0.90
            and exact_pairs >= recommended_pairs
        )
        if high_confidence:
            status = "high_confidence_complete"
            title = "效能模型已高置信完成"
            message = "正确率、一致性和有效比较数均达到高置信标准；更换新技术协议无需重新训练。"
        elif standard_complete:
            status = "standard_complete"
            title = "效能模型已达到可用标准"
            message = "已达到 80% 正确率与一致性门槛，可以停止；继续判断会进一步提高置信度。"
        else:
            status = "continue_learning"
            title = "效能模型仍需继续学习"
            message = "；".join(reasons) + "。"
        additional_pairs = 0 if standard_complete else max(5, minimum_pairs - exact_pairs)
        if validation_accuracy is not None and validation_accuracy < 0.80:
            additional_pairs = max(additional_pairs, 10)
        return {
            "status": status,
            "complete": standard_complete,
            "high_confidence": high_confidence,
            "title": title,
            "message": message,
            "validation_accuracy": None if validation_accuracy is None else round(validation_accuracy, 3),
            "validation_components": {
                "test_accuracy": uta.get("test_accuracy"),
                "cross_validation_accuracy": (uta.get("cross_validation") or {}).get("accuracy"),
                "whole_scheme_holdout_accuracy": (uta.get("whole_scheme_holdout") or {}).get("accuracy"),
                "aggregation": "取全部可用独立验证正确率的最低值",
            },
            "accuracy_target": 0.80,
            "high_confidence_accuracy_target": 0.90,
            "consistency": None if consistency is None else round(consistency, 3),
            "consistency_definition": "全体有效判断在最小容忍UTA模型下的拟合正确率",
            "consistency_target": 0.80,
            "exact_context_pairs": exact_pairs,
            "legacy_warm_start_pairs": warm_pairs,
            "generic_preference_pairs": evidence_pairs,
            "protocol_used_for_training": False,
            "minimum_pairs": minimum_pairs,
            "recommended_pairs": recommended_pairs,
            "suggested_additional_effective_pairs": additional_pairs,
            "open_reviews": open_reviews,
            "reasons": reasons,
        }

    def public_scheme(self, scheme: Dict[str, Any], state: Dict[str, Any]) -> Dict[str, Any]:
        return {**scheme, "prediction": self.evaluate(scheme["params"], state)}

    def public_review(self, state: Dict[str, Any], review: Dict[str, Any]) -> Dict[str, Any]:
        output = dict(review)
        records: List[Dict[str, Any]] = []
        evidence_ids = list(review.get("evidence_ids") or [])
        for evidence_id in evidence_ids:
            preference = self.find_evidence(state.get("preference_evidence", []), str(evidence_id))
            feasibility = self.find_evidence(state.get("feasibility_evidence", []), str(evidence_id))
            if preference:
                records.append(
                    {
                        "id": preference.get("id"),
                        "kind": "preference",
                        "created_at": preference.get("created_at"),
                        "status": preference.get("status"),
                        "relation": preference.get("relation"),
                        "scheme_a": preference.get("scheme_a"),
                        "scheme_b": preference.get("scheme_b"),
                        "params_a": preference.get("params_a"),
                        "params_b": preference.get("params_b"),
                        "requirement_profile_id": preference.get("requirement_profile_id"),
                    }
                )
            elif feasibility:
                records.append(
                    {
                        "id": feasibility.get("id"),
                        "kind": "feasibility",
                        "created_at": feasibility.get("created_at"),
                        "status": feasibility.get("status"),
                        "label": feasibility.get("label"),
                        "scheme_id": feasibility.get("scheme_id"),
                        "params": feasibility.get("params"),
                        "reason_codes": feasibility.get("reason_codes"),
                        "reason_text": feasibility.get("reason_text"),
                    }
                )
        related = [
            item.get("id")
            for item in state.get("review_evidence", [])
            if item.get("id") != review.get("id")
            and set(item.get("evidence_ids") or []).intersection(evidence_ids)
        ]
        output["evidence_records"] = records
        output["dependent_review_ids"] = related
        if output.get("status") != "open" and output.get("resolution"):
            output["allowed_actions"] = ["undo"]
        return output

    def summarize(self, state: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        state = state or self.load_state()
        pair = state["active_pair"]
        scheme_a = self.find_scheme(state, pair["a"])
        scheme_b = self.find_scheme(state, pair["b"])
        if not scheme_a or not scheme_b:
            state = self._initial_state()
            pair = state["active_pair"]
            scheme_a = self.find_scheme(state, pair["a"])
            scheme_b = self.find_scheme(state, pair["b"])
        assert scheme_a and scheme_b
        active_requirement = self.active_requirement_profile(state)
        generated_count = sum(1 for item in state["known_schemes"] if item.get("source") != "existing_sample")
        open_reviews = [item for item in state.get("review_evidence", []) if item.get("status") == "open"]
        recent_resolved = [item for item in state.get("review_evidence", []) if item.get("status") != "open"][-20:][::-1]
        numeric_keys = [item.key for item in self.attributes if item.is_numeric]
        if self.project.couplings:
            default_slice_x = self.project.couplings[0].source_key
            default_slice_y = self.project.couplings[0].target_key
        else:
            default_slice_x = numeric_keys[0] if numeric_keys else None
            default_slice_y = numeric_keys[1] if len(numeric_keys) > 1 else None
        return {
            "project": {
                "name": self.project.project_name,
                "workbook_path": self.project.workbook_path,
                "fingerprint": self.project.workbook_fingerprint,
                "learning_fingerprint": self.project.learning_fingerprint,
                "warnings": self.project.warnings,
                "stage": 10,
                "model_note": "通用BT/UTA与新技术协议已解耦；协议只作为100分参考。",
            },
            "dataMode": state.get("data_mode", "live_expert"),
            "demoProfile": state.get("demo_profile"),
            "params": self.param_profile(),
            "couplings": [
                {
                    "source_key": item.source_key,
                    "source_label": item.source_label,
                    "target_key": item.target_key,
                    "target_label": item.target_label,
                    "direction": item.direction,
                    "status": item.status,
                }
                for item in self.project.couplings
            ],
            "couplingModels": self.coupling_system.summaries(),
            "feasibilityModel": state.get("feasibility_model", {}),
            "feasibilityWeights": state.get("feasibility_weights", {}),
            "btModel": state.get("bt_model", {}),
            "utaModel": state.get("uta_model", {}),
            "learningCompletion": self.learning_completion(state),
            "requirementProfiles": [
                {
                    "id": item.id,
                    "name": item.name,
                    "requirement_count": len(item.requirements),
                }
                for item in self.project.requirement_profiles
            ],
            "protocolProfiles": [
                {
                    "id": item.id,
                    "name": item.name,
                    "attribute_count": len(item.requirements),
                }
                for item in self.project.requirement_profiles
            ],
            "activeRequirementProfileId": self.active_requirement_profile_id(state),
            "activeProtocolProfileId": self.active_requirement_profile_id(state),
            "activeRequirement": (
                {
                    "id": active_requirement.id,
                    "name": active_requirement.name,
                    "requirements": [asdict(item) for item in active_requirement.requirements],
                }
                if active_requirement is not None
                else None
            ),
            "activeProtocol": (
                {
                    "id": active_requirement.id,
                    "name": active_requirement.name,
                    "values": {
                        item.attribute_key: item.target_value
                        if item.target_value is not None
                        else item.minimum
                        if item.minimum is not None
                        else item.maximum
                        for item in active_requirement.requirements
                    },
                }
                if active_requirement is not None
                else None
            ),
            "preferenceSettings": {
                "utaSegments": state.get("uta_segments", DEFAULT_UTA_SEGMENTS),
                "minSegments": MIN_UTA_SEGMENTS,
                "maxSegments": MAX_UTA_SEGMENTS,
            },
            "visualization": {
                "defaultSliceX": default_slice_x,
                "defaultSliceY": default_slice_y,
                "numericAttributes": numeric_keys,
            },
            "reasonOptions": self.reason_options,
            "labels": {
                "status": STATUS_LABELS,
                "feasibility": FEASIBILITY_LABELS,
                "preference": PREFERENCE_LABELS,
                "source": SOURCE_LABELS,
            },
            "stats": {
                "attributes": len(self.attributes),
                "existingSchemes": len(state["existing_samples"]),
                "knownSchemes": len(state["known_schemes"]),
                "generatedSchemes": generated_count,
                "couplings": len(self.project.couplings),
                "interactions": len(state["interactions"]),
                "feasibilityEvidence": len(state.get("feasibility_evidence", [])),
                "preferenceEvidence": len(state.get("preference_evidence", [])),
                "activePreferenceEvidence": len(OnlineBTModel.active_evidence(state.get("preference_evidence", []))),
                "utaTrainingPairs": state.get("uta_model", {}).get("training_pairs", 0),
                "utaTestPairs": state.get("uta_model", {}).get("test_pairs", 0),
                "m3MinimumConflicts": state.get("uta_model", {}).get("m3", {}).get("minimum_conflicts", 0),
                "openReviews": len([item for item in state.get("review_evidence", []) if item.get("status") == "open"]),
                "currentBestId": state["current_best_id"],
            },
            "pair": {"a": self.public_scheme(scheme_a, state), "b": self.public_scheme(scheme_b, state)},
            "learningEvents": state.get("learning_events", [])[-20:][::-1],
            "recentInteractions": state.get("interactions", [])[-8:][::-1],
            "reviewEvidence": [self.public_review(state, item) for item in open_reviews + recent_resolved],
            "reviewHistory": state.get("review_history", [])[-20:][::-1],
        }

    @staticmethod
    def next_evidence_id(items: List[Dict[str, Any]], prefix: str) -> str:
        return f"{prefix}-{len(items) + 1:03d}"

    def sync_preference_reviews(self, state: Dict[str, Any]) -> None:
        """Create traceable reviews for structural conflicts and UTA diagnostics."""
        reviews = state.setdefault("review_evidence", [])
        existing_by_key = {item.get("review_key"): item for item in reviews if item.get("review_key")}
        active = OnlineBTModel.active_evidence(state.get("preference_evidence", []))
        order = {item.get("id"): index for index, item in enumerate(active)}
        detected_keys: set[str] = set()
        structural_ids: set[str] = set()

        def add_review(
            review_key: str,
            review_type: str,
            evidence_ids: List[str],
            target_evidence_id: str,
            message: str,
        ) -> None:
            detected_keys.add(review_key)
            existing = existing_by_key.get(review_key)
            if existing:
                if existing.get("status") == "open":
                    existing["message"] = message
                    existing["updated_at"] = now_iso()
                    existing["diagnostic_active"] = True
                    existing.pop("dependency_note", None)
                return
            review = {
                "id": self.next_evidence_id(reviews, "R"),
                "created_at": now_iso(),
                "type": review_type,
                "review_key": review_key,
                "evidence_ids": evidence_ids,
                "target_evidence_id": target_evidence_id,
                "status": "open",
                "allowed_actions": ["keep", "tie", "reverse", "withdraw"],
                "message": message,
                "diagnostic_active": True,
            }
            reviews.append(review)
            existing_by_key[review_key] = review

        pair_groups: Dict[Tuple[str, str], List[Dict[str, Any]]] = {}
        for item in active:
            if item.get("relation") not in {"A", "B"}:
                continue
            pair_key = tuple(sorted((str(item.get("scheme_a")), str(item.get("scheme_b")))))
            pair_groups.setdefault(pair_key, []).append(item)
        for pair, items in pair_groups.items():
            winners = {item.get("winner_id") for item in items}
            if len(winners) < 2:
                continue
            evidence_ids = sorted(str(item["id"]) for item in items)
            structural_ids.update(evidence_ids)
            target = max(items, key=lambda item: order.get(item.get("id"), -1))
            review_key = f"preference_pair_conflict:{pair[0]}:{pair[1]}:{','.join(evidence_ids)}"
            add_review(
                review_key,
                "preference_pair_conflict",
                evidence_ids,
                str(target["id"]),
                f"方案对 {pair[0]} / {pair[1]} 出现相反胜负判断。默认复核最近证据 {target['id']}。",
            )

        adjacency: Dict[str, List[Tuple[str, str]]] = {}
        strict_items: Dict[str, Dict[str, Any]] = {}
        for item in active:
            if item.get("relation") not in {"A", "B"}:
                continue
            winner = str(item.get("winner_id"))
            loser = str(item.get("loser_id"))
            evidence_id = str(item.get("id"))
            adjacency.setdefault(winner, []).append((loser, evidence_id))
            strict_items[evidence_id] = item
        seen_cycles: set[Tuple[str, ...]] = set()
        for evidence_id, item in strict_items.items():
            winner = str(item["winner_id"])
            loser = str(item["loser_id"])
            queue: List[Tuple[str, List[str], List[str]]] = [(loser, [loser], [])]
            found: Optional[Tuple[List[str], List[str]]] = None
            while queue and found is None:
                node, path_nodes, path_evidence = queue.pop(0)
                for target, edge_id in adjacency.get(node, []):
                    if edge_id == evidence_id:
                        continue
                    next_nodes = path_nodes + [target]
                    next_evidence = path_evidence + [edge_id]
                    if target == winner:
                        found = (next_nodes, next_evidence)
                        break
                    if target not in path_nodes and len(next_nodes) <= 8:
                        queue.append((target, next_nodes, next_evidence))
            if not found:
                continue
            cycle_ids = tuple(sorted({evidence_id, *found[1]}))
            if len(cycle_ids) < 3 or cycle_ids in seen_cycles:
                continue
            seen_cycles.add(cycle_ids)
            structural_ids.update(cycle_ids)
            cycle_nodes = [winner, *found[0][:-1], winner]
            target_id = max(cycle_ids, key=lambda value: order.get(value, -1))
            review_key = f"preference_cycle:{','.join(cycle_ids)}"
            add_review(
                review_key,
                "preference_cycle",
                list(cycle_ids),
                target_id,
                f"检测到偏好环 {' > '.join(cycle_nodes)}。默认复核最近证据 {target_id}。",
            )

        for diagnostic in state.get("uta_model", {}).get("tolerances", []):
            evidence_id = str(diagnostic.get("evidence_id"))
            if evidence_id in structural_ids:
                continue
            high_tolerance = bool(diagnostic.get("high_tolerance"))
            m3_conflict = bool(diagnostic.get("m3_conflict"))
            if not high_tolerance and not m3_conflict:
                continue
            slack = diagnostic.get("slack")
            review_type = "uta_m3_conflict" if m3_conflict else "uta_high_tolerance"
            review_key = f"{review_type}:{evidence_id}"
            reason = "M3 最小冲突集命中" if m3_conflict else f"M2 容忍量 {slack} 超过阈值"
            add_review(
                review_key,
                review_type,
                [evidence_id],
                evidence_id,
                f"偏好证据 {evidence_id} 的 {reason}，建议复核该次判断。",
            )

        auto_types = {
            "preference_pair_conflict",
            "preference_cycle",
            "uta_high_tolerance",
            "uta_m3_conflict",
        }
        for review in reviews:
            if (
                review.get("type") in auto_types
                and review.get("status") == "open"
                and review.get("review_key") not in detected_keys
            ):
                review["diagnostic_active"] = False
                review["dependency_note"] = (
                    "关联证据更新后，算法暂时不再检出该冲突；"
                    "本复核项未删除，仍需专家独立确认。"
                )
                review["updated_at"] = now_iso()

    @staticmethod
    def find_evidence(items: List[Dict[str, Any]], evidence_id: str) -> Optional[Dict[str, Any]]:
        return next((item for item in items if item.get("id") == evidence_id), None)

    def review_action(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        state = self.load_state()
        review_id = str(payload.get("reviewId") or "")
        action = str(payload.get("action") or "")
        if action not in {"keep", "tie", "reverse", "withdraw", "undo"}:
            raise ProjectDataError("复核操作无效。")
        review = next((item for item in state.get("review_evidence", []) if item.get("id") == review_id), None)
        if not review:
            raise ProjectDataError("复核项不存在。")
        if action == "undo":
            if review.get("status") == "open" or not review.get("resolution"):
                raise ProjectDataError("该复核项没有可撤销的处理。")
            resolution = review["resolution"]
            target_id = str(resolution.get("target_evidence_id") or "")
            preference = self.find_evidence(state.get("preference_evidence", []), target_id)
            feasibility = self.find_evidence(state.get("feasibility_evidence", []), target_id)
            evidence = preference or feasibility
            previous = dict(resolution.get("previous") or {})
            if evidence:
                for key, value in previous.items():
                    evidence[key] = value
                evidence.setdefault("revision_history", []).append(
                    {"created_at": now_iso(), "review_id": review_id, "action": "undo", "restored": previous}
                )
            review["status"] = "open"
            review["reopened_at"] = now_iso()
            review["last_undone_resolution"] = resolution
            review.pop("resolution", None)
            review.pop("resolved_at", None)
            state.setdefault("review_history", []).append(
                {
                    "created_at": now_iso(),
                    "review_id": review_id,
                    "review_type": review.get("type"),
                    "action": "undo",
                    "target_evidence_id": target_id,
                    "previous": previous,
                }
            )
            if preference:
                self.retrain_preference_models(state)
            elif feasibility:
                self.retrain_feasibility_model(state)
            state["learning_events"].append(
                {
                    "id": f"E-{len(state['learning_events']) + 1:03d}",
                    "created_at": now_iso(),
                    "type": "review_reopened",
                    "title": "已撤销上次复核处理",
                    "detail": f"{review_id} 已恢复为待复核，证据 {target_id or '无'} 已恢复到处理前状态。",
                }
            )
            self.save_state(state)
            return self.summarize(state)
        if review.get("status") != "open":
            raise ProjectDataError("该复核项已经处理。")
        if action not in review.get("allowed_actions", ["keep", "withdraw"]):
            raise ProjectDataError("该复核项不支持此操作。")
        target_id = str(review.get("target_evidence_id") or "")
        preference = self.find_evidence(state.get("preference_evidence", []), target_id)
        feasibility = self.find_evidence(state.get("feasibility_evidence", []), target_id)
        evidence = preference or feasibility
        if action != "keep" and not evidence:
            raise ProjectDataError("待处理的原始证据不存在。")

        previous: Dict[str, Any] = {}
        if evidence:
            previous = {
                "status": evidence.get("status"),
                "relation": evidence.get("relation"),
                "label": evidence.get("label"),
                "winner_id": evidence.get("winner_id"),
                "loser_id": evidence.get("loser_id"),
            }
        if action == "withdraw":
            assert evidence is not None
            evidence["status"] = "withdrawn"
        elif action == "tie":
            if not preference:
                raise ProjectDataError("只有偏好证据可以改为差不多。")
            preference["relation"] = "tie"
            preference["winner_id"] = None
            preference["loser_id"] = None
            preference["status"] = "revised"
        elif action == "reverse":
            if not preference or preference.get("relation") not in {"A", "B"}:
                raise ProjectDataError("只有严格偏好证据可以反向。")
            preference["relation"] = "B" if preference["relation"] == "A" else "A"
            preference["winner_id"], preference["loser_id"] = preference["loser_id"], preference["winner_id"]
            preference["status"] = "revised"

        if evidence and action != "keep":
            evidence.setdefault("revision_history", []).append(
                {"created_at": now_iso(), "review_id": review_id, "action": action, "previous": previous}
            )
        review["status"] = f"resolved_{action}"
        review["resolved_at"] = now_iso()
        review["resolution"] = {"action": action, "target_evidence_id": target_id, "previous": previous}
        state.setdefault("review_history", []).append(
            {
                "created_at": now_iso(),
                "review_id": review_id,
                "review_type": review.get("type"),
                "action": action,
                "target_evidence_id": target_id,
                "previous": previous,
            }
        )
        if preference and action != "keep":
            self.retrain_preference_models(state)
        elif feasibility and action == "withdraw":
            self.retrain_feasibility_model(state)
        state["learning_events"].append(
            {
                "id": f"E-{len(state['learning_events']) + 1:03d}",
                "created_at": now_iso(),
                "type": "review_resolved",
                "title": "复核项已处理",
                "detail": f"{review_id} 对证据 {target_id or '无'} 执行 {action}；相关模型已按当前有效证据更新。",
            }
        )
        self.save_state(state)
        return self.summarize(state)

    def record_feasibility_evidence(
        self,
        state: Dict[str, Any],
        interaction_id: str,
        scheme: Dict[str, Any],
        side: str,
        label_value: str,
        reason_codes: List[str],
        reason_text: str,
        coupling_feedback: Optional[Dict[str, Any]] = None,
    ) -> Optional[Dict[str, Any]]:
        if label_value not in {"feasible", "infeasible"}:
            return None
        evidence = {
            "id": self.next_evidence_id(state["feasibility_evidence"], "F"),
            "created_at": now_iso(),
            "interaction_id": interaction_id,
            "scheme_id": scheme["id"],
            "scheme_side": side,
            "label": label_value,
            "reason_codes": list(reason_codes),
            "reason_text": reason_text,
            "coupling_feedback": coupling_feedback,
            "params": dict(scheme["params"]),
            "source": "expert",
            "status": "active",
        }
        attribute_feedback = self.prior_feasibility_model.infer_attribute_feedback(evidence)
        if attribute_feedback:
            spec = self.attribute_by_key[attribute_feedback["attribute_key"]]
            attribute_feedback.update(
                {
                    "attribute_label": spec.label,
                    "value": scheme["params"].get(spec.key),
                    "unit": spec.unit,
                    "collection_mode": attribute_feedback.get("scope", "structured_feedback"),
                }
            )
            evidence["attribute_feedback"] = attribute_feedback
        state["feasibility_evidence"].append(evidence)
        labels = {
            item["label"]
            for item in state["feasibility_evidence"]
            if item.get("scheme_id") == scheme["id"] and item.get("status") == "active"
        }
        if labels == {"feasible", "infeasible"}:
            existing = next(
                (
                    item
                    for item in state["review_evidence"]
                    if item.get("type") == "feasibility_conflict"
                    and item.get("scheme_id") == scheme["id"]
                    and item.get("status") == "open"
                ),
                None,
            )
            related_ids = [
                item["id"] for item in state["feasibility_evidence"] if item.get("scheme_id") == scheme["id"]
            ]
            if existing:
                existing["evidence_ids"] = related_ids
                existing["target_evidence_id"] = related_ids[-1]
                existing["updated_at"] = now_iso()
            else:
                state["review_evidence"].append(
                    {
                        "id": self.next_evidence_id(state["review_evidence"], "R"),
                        "created_at": now_iso(),
                        "type": "feasibility_conflict",
                        "scheme_id": scheme["id"],
                        "evidence_ids": related_ids,
                        "target_evidence_id": related_ids[-1],
                        "status": "open",
                        "allowed_actions": ["keep", "withdraw"],
                        "message": f"方案 {scheme['id']} 同时出现可行与不可行判断，需要专家复核。",
                    }
                )
        return evidence

    def normalize_coupling_feedback(
        self,
        raw: Any,
        scheme: Dict[str, Any],
        label_value: str,
        reason_codes: List[str],
    ) -> Optional[Dict[str, Any]]:
        provided = raw if isinstance(raw, dict) else {}
        explanation = scheme.get("generation_explanation") or {}
        target_key = str(provided.get("target_key") or "").strip() or None
        if not target_key:
            coupling_codes = [code[len("coupling_") :] for code in reason_codes if code.startswith("coupling_")]
            target_key = coupling_codes[0] if coupling_codes else explanation.get("probe_target")
        if target_key not in self.coupling_system.target_keys():
            return None
        issue = str(provided.get("issue") or "").strip()
        if label_value == "feasible":
            issue = "feasible_anchor"
        elif issue not in {"target_low", "target_high", "mismatch"}:
            issue = "mismatch"
        probe_source = explanation.get("probe_source")
        source_keys = [probe_source] if probe_source else [
            edge.source_key for edge in self.coupling_system.edges_by_target.get(target_key, [])
        ]
        return {
            "target_key": target_key,
            "target_label": self.attribute_by_key[target_key].label,
            "issue": issue,
            "source_keys": source_keys,
            "source_labels": [self.attribute_by_key[key].label for key in source_keys],
            "collection_mode": "auto_suggested_quick_feedback",
            "confidence": 0.90 if provided.get("target_key") else 0.72,
        }

    def record_preference_evidence(
        self,
        state: Dict[str, Any],
        interaction_id: str,
        scheme_a: Dict[str, Any],
        scheme_b: Dict[str, Any],
        feasibility_a: str,
        feasibility_b: str,
        preference: str,
    ) -> Optional[Dict[str, Any]]:
        if feasibility_a != "feasible" or feasibility_b != "feasible" or preference not in {"A", "B", "tie"}:
            return None
        evidence = {
            "id": self.next_evidence_id(state["preference_evidence"], "P"),
            "created_at": now_iso(),
            "interaction_id": interaction_id,
            "scheme_a": scheme_a["id"],
            "scheme_b": scheme_b["id"],
            "relation": preference,
            "winner_id": scheme_a["id"] if preference == "A" else scheme_b["id"] if preference == "B" else None,
            "loser_id": scheme_b["id"] if preference == "A" else scheme_a["id"] if preference == "B" else None,
            "params_a": dict(scheme_a["params"]),
            "params_b": dict(scheme_b["params"]),
            "confidence": 1.0,
            "expert_id": "default_expert",
            "revision_history": [],
            "status": "pending_bt_uta",
            "learning_context": "generic_product_effectiveness",
        }
        state["preference_evidence"].append(evidence)
        return evidence

    def submit(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        state = self.load_state()
        pair = state["active_pair"]
        scheme_a = self.find_scheme(state, pair["a"])
        scheme_b = self.find_scheme(state, pair["b"])
        if not scheme_a or not scheme_b:
            raise ProjectDataError("当前 A/B 方案不存在，请重置项目状态。")
        feasibility_a = str(payload.get("feasibilityA", "uncertain"))
        feasibility_b = str(payload.get("feasibilityB", "uncertain"))
        preference = str(payload.get("preference", "unknown"))
        if feasibility_a not in FEASIBILITY_LABELS or feasibility_b not in FEASIBILITY_LABELS:
            raise ProjectDataError("可行性选项无效。")
        if preference not in PREFERENCE_LABELS:
            raise ProjectDataError("偏好选项无效。")
        interaction_id = f"H-{len(state['interactions']) + 1:03d}"
        valid_reason_codes = set(self.reason_options)
        reason_codes_a = sorted(set(payload.get("reasonCodesA") or []) & valid_reason_codes)
        reason_codes_b = sorted(set(payload.get("reasonCodesB") or []) & valid_reason_codes)
        reason_a = str(payload.get("reasonA") or "").strip()
        reason_b = str(payload.get("reasonB") or "").strip()
        coupling_feedback_a = self.normalize_coupling_feedback(
            payload.get("couplingFeedbackA"), scheme_a, feasibility_a, reason_codes_a
        )
        coupling_feedback_b = self.normalize_coupling_feedback(
            payload.get("couplingFeedbackB"), scheme_b, feasibility_b, reason_codes_b
        )
        if feasibility_a == "infeasible" and coupling_feedback_a:
            reason_codes_a = sorted(set(reason_codes_a + [f"coupling_{coupling_feedback_a['target_key']}"]))
        if feasibility_b == "infeasible" and coupling_feedback_b:
            reason_codes_b = sorted(set(reason_codes_b + [f"coupling_{coupling_feedback_b['target_key']}"]))
        interaction = {
            "id": interaction_id,
            "created_at": now_iso(),
            "scheme_a": scheme_a["id"],
            "scheme_b": scheme_b["id"],
            "feasibility_a": feasibility_a,
            "feasibility_b": feasibility_b,
            "preference": preference,
            "reason_codes_a": reason_codes_a,
            "reason_codes_b": reason_codes_b,
            "reason_a": reason_a,
            "reason_b": reason_b,
            "coupling_feedback_a": coupling_feedback_a,
            "coupling_feedback_b": coupling_feedback_b,
            "feasibility_evidence_ids": [],
            "preference_evidence_id": None,
            "learning_status": "可行性证据更新 P_feasible；有效偏好立即更新 BT，并重算 LP-UTA 检查点",
        }
        state["interactions"].append(interaction)
        reviews_before = len(state["review_evidence"])
        evidence_a = self.record_feasibility_evidence(
            state,
            interaction_id,
            scheme_a,
            "A",
            feasibility_a,
            reason_codes_a,
            reason_a,
            coupling_feedback_a,
        )
        evidence_b = self.record_feasibility_evidence(
            state,
            interaction_id,
            scheme_b,
            "B",
            feasibility_b,
            reason_codes_b,
            reason_b,
            coupling_feedback_b,
        )
        interaction["feasibility_evidence_ids"] = [
            item["id"] for item in (evidence_a, evidence_b) if item is not None
        ]
        preference_evidence = self.record_preference_evidence(
            state, interaction_id, scheme_a, scheme_b, feasibility_a, feasibility_b, preference
        )
        if preference_evidence:
            interaction["preference_evidence_id"] = preference_evidence["id"]
            self.retrain_preference_models(state)
        if evidence_a or evidence_b:
            self.retrain_feasibility_model(state)
        if feasibility_a == "feasible" and preference == "A":
            state["current_best_id"] = scheme_a["id"]
        elif feasibility_b == "feasible" and preference == "B":
            state["current_best_id"] = scheme_b["id"]
        elif feasibility_b == "feasible" and feasibility_a != "feasible":
            state["current_best_id"] = scheme_b["id"]

        base = self.find_scheme(state, state["current_best_id"]) or scheme_a
        candidate = self.select_candidate(state, base)
        state["known_schemes"].append(candidate)
        state["active_pair"] = {"a": base["id"], "b": candidate["id"]}
        state["learning_events"].append(
            {
                "id": f"E-{len(state['learning_events']) + 1:03d}",
                "created_at": now_iso(),
                "type": "feedback_saved",
                "title": "专家证据已拆分并保存",
                "detail": (
                    f"A={FEASIBILITY_LABELS[feasibility_a]}，B={FEASIBILITY_LABELS[feasibility_b]}，"
                    f"偏好={PREFERENCE_LABELS[preference]}；形成 {len(interaction['feasibility_evidence_ids'])} 条可行性证据"
                    f"和 {1 if preference_evidence else 0} 条偏好证据。"
                ),
            }
        )
        if evidence_a or evidence_b:
            model_summary = state["feasibility_model"]
            state["learning_events"].append(
                {
                    "id": f"E-{len(state['learning_events']) + 1:03d}",
                    "created_at": now_iso(),
                    "type": "feasibility_model_updated",
                    "title": "学习型可行性模型已更新",
                    "detail": (
                        f"当前训练样本 {model_summary.get('training_samples', 0)} 条，其中专家证据 "
                        f"{model_summary.get('expert_samples', 0)} 条；正例 {model_summary.get('positive_samples', 0)}、"
                        f"负例 {model_summary.get('negative_samples', 0)}。"
                    ),
                }
            )
        if preference_evidence:
            bt_summary = state["bt_model"]
            uta_summary = state["uta_model"]
            state["learning_events"].append(
                {
                    "id": f"E-{len(state['learning_events']) + 1:03d}",
                    "created_at": now_iso(),
                    "type": "preference_models_updated",
                    "title": "在线 BT 与 LP-UTA 已更新",
                    "detail": (
                        f"BT 已学习 {bt_summary.get('training_pairs', 0)} 对；UTA 当前 {uta_summary.get('segments')} 段、"
                        f"{uta_summary.get('utility_increment_variables')} 个增量变量，M1 "
                        f"{'一致' if uta_summary.get('m1_consistent') else '不一致'}，"
                        f"训练/测试 {uta_summary.get('training_pairs', 0)}/{uta_summary.get('test_pairs', 0)}，"
                        f"状态 {uta_summary.get('status')}。"
                    ),
                }
            )
        if len(state["review_evidence"]) > reviews_before:
            new_reviews = state["review_evidence"][reviews_before:]
            state["learning_events"].append(
                {
                    "id": f"E-{len(state['learning_events']) + 1:03d}",
                    "created_at": now_iso(),
                    "type": "review_created",
                    "title": "发现需要复核的专家证据",
                    "detail": f"新增 {len(new_reviews)} 项：" + "；".join(item["message"] for item in new_reviews[:3]),
                }
            )
        self.save_state(state)
        return self.summarize(state)

    def update_preference_settings(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        state = self.load_state()
        segments = validate_segment_count(payload.get("utaSegments"))
        previous = int(state.get("uta_segments", DEFAULT_UTA_SEGMENTS))
        if segments != previous:
            state["uta_segments"] = segments
            self.retrain_preference_models(state)
            summary = state["uta_model"]
            state["learning_events"].append(
                {
                    "id": f"E-{len(state['learning_events']) + 1:03d}",
                    "created_at": now_iso(),
                    "type": "uta_segments_changed",
                    "title": "UTA 分段数已调整",
                    "detail": (
                        f"统一分段数由 {previous} 改为 {segments}；全部属性已按相同分段重建。"
                        f"当前 {summary.get('utility_increment_variables')} 个效用增量变量，"
                        f"建议至少 {summary.get('recommended_training_pairs')} 条训练偏好。原始专家证据均已保留。"
                    ),
                }
            )
            self.save_state(state)
        return self.summarize(state)

    def update_active_protocol(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        """Switch only the final scoring reference; generic learning is untouched."""
        state = self.load_state()
        protocol_id = str(payload.get("protocolId") or "").strip()
        profiles = self.project.requirement_profile_by_id()
        if protocol_id not in profiles:
            raise ProjectDataError(f"未找到新技术协议“{protocol_id}”。")
        previous_id = self.active_requirement_profile_id(state)
        if protocol_id != previous_id:
            state["active_protocol_profile_id"] = protocol_id
            state.setdefault("learning_events", []).append(
                {
                    "id": f"E-{len(state.get('learning_events', [])) + 1:03d}",
                    "created_at": now_iso(),
                    "type": "active_protocol_changed",
                    "title": "100分评分参考已切换",
                    "detail": (
                        f"评分参考由 {previous_id} 切换为 {protocol_id}。"
                        "物理可行模型、耦合模型、BT/UTA参数和全部专家判断均未改变；"
                        "未知方案将在下次评估时按新协议重新计分。"
                    ),
                }
            )
            self.save_state(state)
        return self.summarize(state)

    def update_generation_ranges(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        state = self.load_state()
        raw_ranges = payload.get("ranges")
        if not isinstance(raw_ranges, dict):
            raise ProjectDataError("生成范围格式无效。")
        ranges: Dict[str, Tuple[float, float]] = {}
        previous: Dict[str, List[float]] = {}
        for key, raw in raw_ranges.items():
            if key not in self.attribute_by_key:
                raise ProjectDataError(f"未知属性ID：{key}。")
            spec = self.attribute_by_key[key]
            if not spec.is_numeric or not spec.participates_generation:
                continue
            if not isinstance(raw, dict):
                raise ProjectDataError(f"属性“{spec.label}”的范围格式无效。")
            try:
                lo = float(raw.get("min"))
                hi = float(raw.get("max"))
            except (TypeError, ValueError) as exc:
                raise ProjectDataError(f"属性“{spec.label}”的生成上下限必须是数值。") from exc
            if not math.isfinite(lo) or not math.isfinite(hi) or lo >= hi:
                raise ProjectDataError(f"属性“{spec.label}”的生成下限必须小于生成上限。")
            assert spec.feasible_min is not None and spec.feasible_max is not None
            if lo > spec.feasible_min or hi < spec.feasible_max:
                raise ProjectDataError(
                    f"属性“{spec.label}”的新生成范围必须包含样本经验范围 "
                    f"[{spec.feasible_min:g}, {spec.feasible_max:g}]。"
                )
            ranges[key] = (lo, hi)
            previous[key] = [float(spec.generation_min), float(spec.generation_max)]
        if not ranges:
            raise ProjectDataError("没有可保存的数值属性生成范围。")

        old_state_path = self.state_path
        old_fingerprint = self.project.workbook_fingerprint
        backup = write_generation_ranges(self.project.workbook_path, ranges)
        new_project = load_project_workbook(self.project.workbook_path)
        self.configure_project(new_project)

        existing = self.existing_schemes()
        generated = [
            item for item in state.get("known_schemes", []) if item.get("source") != "existing_sample"
        ]
        state["existing_samples"] = existing
        state["known_schemes"] = existing + generated
        state["workbook_fingerprint"] = new_project.workbook_fingerprint
        state["learning_fingerprint"] = new_project.learning_fingerprint
        state["workbook_path"] = new_project.workbook_path
        state["profile_version"] = PROFILE_VERSION
        state.setdefault("range_update_history", []).append(
            {
                "created_at": now_iso(),
                "old_fingerprint": old_fingerprint,
                "new_fingerprint": new_project.workbook_fingerprint,
                "old_state_file": str(old_state_path),
                "workbook_backup": str(backup),
                "previous_ranges": previous,
                "new_ranges": {key: list(value) for key, value in ranges.items()},
            }
        )
        self.retrain_feasibility_model(state)
        self.retrain_preference_models(state)
        base = self.find_scheme(state, state.get("current_best_id")) or existing[0]
        candidate = self.select_candidate(state, base)
        state["known_schemes"].append(candidate)
        state["active_pair"] = {"a": base["id"], "b": candidate["id"]}
        state.setdefault("learning_events", []).append(
            {
                "id": f"E-{len(state.get('learning_events', [])) + 1:03d}",
                "created_at": now_iso(),
                "type": "generation_ranges_changed",
                "title": "生成范围已写回 Excel",
                "detail": (
                    f"已更新 {len(ranges)} 个属性，样本经验范围保持自动读取；"
                    "原判断已迁移到新工作簿指纹，并按新范围生成下一组方案。"
                ),
            }
        )
        self.save_state(state)
        return self.summarize(state)

    def feasibility_slice(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        state = self.load_state()
        numeric_specs = {item.key: item for item in self.attributes if item.is_numeric}
        x_key = str(payload.get("xKey") or "")
        y_key = str(payload.get("yKey") or "")
        if x_key not in numeric_specs or y_key not in numeric_specs:
            raise ProjectDataError("可行域切片坐标必须选择数值属性。")
        if x_key == y_key:
            raise ProjectDataError("可行域切片的横轴和纵轴不能相同。")
        try:
            resolution = int(payload.get("resolution", 20))
        except (TypeError, ValueError) as exc:
            raise ProjectDataError("切片分辨率必须是整数。") from exc
        resolution = max(12, min(30, resolution))
        pair = state.get("active_pair") or {}
        base = None
        for evidence in reversed(state.get("feasibility_evidence", [])):
            if evidence.get("label") == "feasible" and evidence.get("status", "active") == "active":
                base = self.find_scheme(state, str(evidence.get("scheme_id")))
                if base:
                    break
        base = (
            base
            or self.find_scheme(state, state.get("current_best_id"))
            or self.find_scheme(state, str(pair.get("a")))
        )
        if not base:
            raise ProjectDataError("找不到可行域切片的基准方案。")
        x_spec = numeric_specs[x_key]
        y_spec = numeric_specs[y_key]
        assert x_spec.generation_min is not None and x_spec.generation_max is not None
        assert y_spec.generation_min is not None and y_spec.generation_max is not None
        feasibility_model = self.feasibility_model_from_state(state)
        cells: List[Dict[str, Any]] = []
        for y_index in range(resolution):
            y_ratio = y_index / max(resolution - 1, 1)
            y_value = y_spec.generation_min + y_ratio * (y_spec.generation_max - y_spec.generation_min)
            for x_index in range(resolution):
                x_ratio = x_index / max(resolution - 1, 1)
                x_value = x_spec.generation_min + x_ratio * (x_spec.generation_max - x_spec.generation_min)
                params = dict(base["params"])
                params[x_key] = int(round(x_value)) if x_spec.data_type == "integer" else x_value
                params[y_key] = int(round(y_value)) if y_spec.data_type == "integer" else y_value
                hard_violation = any(
                    item.is_numeric
                    and (
                        float(params[item.key]) < float(item.generation_min)
                        or float(params[item.key]) > float(item.generation_max)
                    )
                    for item in self.attributes
                )
                experience_extrapolation = any(
                    item.is_numeric
                    and (
                        float(params[item.key]) < float(item.feasible_min)
                        or float(params[item.key]) > float(item.feasible_max)
                    )
                    for item in self.attributes
                )
                coupling_assessments = self.coupling_system.assess(params)
                outside_assessments = [
                    item for item in coupling_assessments if item.status in {"below_band", "above_band"}
                ]
                coupling_outside = len(outside_assessments)
                coupling_severity = max([float(item.severity) for item in outside_assessments] or [0.0])
                probability = feasibility_model.probability(params)
                mature_boundary_violation = bool(
                    feasibility_model.boundary_violations(params, mature_only=True)
                )
                status = self.classify_feasibility_status(
                    probability,
                    hard_violation=hard_violation,
                    mature_boundary_violation=mature_boundary_violation,
                    outside_sample_experience=experience_extrapolation,
                    coupling_outside=bool(coupling_outside),
                    coupling_severity=coupling_severity,
                )
                cells.append(
                    {
                        "x_index": x_index,
                        "y_index": y_index,
                        "x": round(float(params[x_key]), x_spec.precision),
                        "y": round(float(params[y_key]), y_spec.precision),
                        "probability": round(probability, 4),
                        "status": status,
                        "coupling_outside": coupling_outside,
                        "coupling_severity": round(coupling_severity, 4),
                        "outside_sample_experience": experience_extrapolation,
                        "mature_boundary_violation": mature_boundary_violation,
                    }
                )
        points = [
            {
                "source": "existing_sample",
                "scheme_id": item["id"],
                "x": item["params"][x_key],
                "y": item["params"][y_key],
                "label": "existing",
            }
            for item in state.get("existing_samples", [])
        ]
        points.extend(
            {
                "source": "expert_evidence",
                "scheme_id": item["scheme_id"],
                "evidence_id": item["id"],
                "x": item["params"][x_key],
                "y": item["params"][y_key],
                "label": item["label"],
                "reason_codes": item.get("reason_codes", []),
            }
            for item in state.get("feasibility_evidence", [])
            if item.get("status") != "withdrawn"
        )
        return {
            "base_scheme_id": base["id"],
            "x_axis": {
                "key": x_key,
                "label": x_spec.label,
                "unit": x_spec.unit,
                "min": x_spec.generation_min,
                "max": x_spec.generation_max,
                "precision": x_spec.precision,
            },
            "y_axis": {
                "key": y_key,
                "label": y_spec.label,
                "unit": y_spec.unit,
                "min": y_spec.generation_min,
                "max": y_spec.generation_max,
                "precision": y_spec.precision,
            },
            "resolution": resolution,
            "cells": cells,
            "points": points,
            "fixed_attributes": [
                {"key": item.key, "label": item.label, "unit": item.unit, "value": base["params"][item.key]}
                for item in self.attributes
                if item.key not in {x_key, y_key}
            ],
            "model_note": "切片优先固定其余属性为最近一次专家确认可行的方案；样本范围外只标记经验外推，只有生成范围外才属于硬违反。",
        }

    def evaluate_unknown(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        params = self.parse_params(payload)
        state = self.load_state()
        result = self.evaluate(params, state)
        record = {
            "id": f"U-{len(state['unknown_evaluations']) + 1:03d}",
            "created_at": now_iso(),
            "params": params,
            "result": result,
        }
        state["unknown_evaluations"].append(record)
        state["unknown_evaluations"] = state["unknown_evaluations"][-100:]
        self.save_state(state)
        return record

    @staticmethod
    def serializable_cell(value: Any) -> Any:
        if isinstance(value, (dict, list, tuple)):
            return json.dumps(value, ensure_ascii=False)
        return value

    @classmethod
    def csv_text(cls, rows: List[Dict[str, Any]]) -> str:
        if not rows:
            return "状态\r\n暂无数据\r\n"
        headers: List[str] = []
        for row in rows:
            for key in row:
                if key not in headers:
                    headers.append(key)
        output = StringIO(newline="")
        writer = csv.DictWriter(output, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({key: cls.serializable_cell(row.get(key)) for key in headers})
        return output.getvalue()

    def export_datasets(self, state: Dict[str, Any]) -> Dict[str, List[Dict[str, Any]]]:
        schemes = [
            {
                "方案编号": item["id"],
                "来源": item.get("source"),
                "来源说明": item.get("source_detail"),
                **{self.attribute_by_key[key].label: value for key, value in item["params"].items()},
            }
            for item in state.get("known_schemes", [])
        ]
        interactions = [
            {
                "交互编号": item.get("id"),
                "时间": item.get("created_at"),
                "方案A": item.get("scheme_a"),
                "方案B": item.get("scheme_b"),
                "A可行性": item.get("feasibility_a"),
                "B可行性": item.get("feasibility_b"),
                "偏好": item.get("preference"),
                "来源": item.get("source", "human_expert"),
                "A原因标签": item.get("reason_codes_a", []),
                "B原因标签": item.get("reason_codes_b", []),
                "A原因文本": item.get("reason_a"),
                "B原因文本": item.get("reason_b"),
                "A耦合快速反馈": item.get("coupling_feedback_a"),
                "B耦合快速反馈": item.get("coupling_feedback_b"),
            }
            for item in state.get("interactions", [])
        ]
        feasibility = [
            {
                "证据编号": item.get("id"),
                "时间": item.get("created_at"),
                "方案编号": item.get("scheme_id"),
                "判断": item.get("label"),
                "状态": item.get("status"),
                "来源": item.get("source"),
                "专家编号": item.get("expert_id"),
                "原因标签": item.get("reason_codes", []),
                "原因文本": item.get("reason_text"),
                "耦合前沿反馈": item.get("coupling_feedback"),
                **{self.attribute_by_key[key].label: value for key, value in item.get("params", {}).items()},
            }
            for item in state.get("feasibility_evidence", [])
        ]
        preference = [
            {
                "证据编号": item.get("id"),
                "时间": item.get("created_at"),
                "方案A": item.get("scheme_a"),
                "方案B": item.get("scheme_b"),
                "关系": item.get("relation"),
                "胜者": item.get("winner_id"),
                "败者": item.get("loser_id"),
                "置信度": item.get("confidence"),
                "状态": item.get("status"),
                "来源": item.get("source", "human_expert"),
                "专家编号": item.get("expert_id"),
                "修订历史": item.get("revision_history", []),
            }
            for item in state.get("preference_evidence", [])
        ]
        reviews = [
            {
                "时间": item.get("created_at"),
                "复核编号": item.get("review_id"),
                "复核类型": item.get("review_type"),
                "操作": item.get("action"),
                "目标证据": item.get("target_evidence_id"),
                "原值": item.get("previous"),
            }
            for item in state.get("review_history", [])
        ]
        unknown = [
            {
                "评估编号": item.get("id"),
                "时间": item.get("created_at"),
                "来源": item.get("source", "user_input"),
                "可行状态": item.get("result", {}).get("status"),
                "P可行": item.get("result", {}).get("learned_feasibility_probability"),
                "可行性置信说明": item.get("result", {}).get("feasibility_confidence"),
                "效能分": item.get("result", {}).get("effectiveness_score"),
                "效能来源": item.get("result", {}).get("effectiveness_source"),
                "效能置信说明": item.get("result", {}).get("effectiveness_confidence"),
                "UTA区间": item.get("result", {}).get("uta_score_interval"),
                "评估依据": item.get("result", {}).get("assessment_basis"),
                **{self.attribute_by_key[key].label: value for key, value in item.get("params", {}).items()},
            }
            for item in state.get("unknown_evaluations", [])
        ]
        curves = []
        for curve in state.get("uta_model", {}).get("marginal_curves", []):
            for index, point in enumerate(curve.get("points", [])):
                curves.append(
                    {
                        "属性ID": curve.get("key"),
                        "属性": curve.get("label"),
                        "单位": curve.get("unit"),
                        "偏好方向": curve.get("preference_direction"),
                        "边际规律": curve.get("marginal_trend"),
                        "属性权重": curve.get("attribute_weight"),
                        "断点序号": index,
                        "原始值": point.get("raw_value"),
                        "效用方向位置": point.get("benefit"),
                        "累计效用": point.get("utility"),
                    }
                )
        couplings = [
            {
                "源属性": item.source_label,
                "目标属性": item.target_label,
                "方向": item.direction,
                "关系类型": item.relation_type,
                "置信状态": item.status,
                "说明": item.description,
            }
            for item in self.project.couplings
        ]
        return {
            "schemes": schemes,
            "interactions": interactions,
            "feasibility_evidence": feasibility,
            "preference_evidence": preference,
            "review_history": reviews,
            "unknown_evaluations": unknown,
            "uta_marginal_curves": curves,
            "couplings": couplings,
        }

    def report_markdown(self, state: Dict[str, Any]) -> str:
        feasibility = state.get("feasibility_model", {})
        uta = state.get("uta_model", {})
        mode = state.get("data_mode", "live_expert")
        lines = [
            f"# {self.project.project_name} 效能评估报告",
            "",
            f"生成时间：{now_iso()}",
            f"数据模式：{mode}",
        ]
        if mode == "fixed_demo_simulation":
            lines.extend(["", "> 注意：当前报告使用系统模拟专家展示数据，不得作为真实工程结论。"])
        lines.extend(
            [
                "",
                "## 项目概览",
                "",
                f"- 属性：{len(self.attributes)} 个",
                f"- Excel 已有方案：{len(state.get('existing_samples', []))} 个",
                f"- 有向耦合：{len(self.project.couplings)} 条",
                f"- 可行性专家证据：{feasibility.get('expert_samples', 0)} 条",
                f"- 有效偏好：{uta.get('evidence_pairs', 0)} 对",
                "",
                "## 模型状态",
                "",
                f"- UTA 分段数：{uta.get('segments', DEFAULT_UTA_SEGMENTS)}",
                f"- UTA 状态：{uta.get('status', 'no_data')}",
                f"- M1 一致性：{uta.get('m1_label', '待检验')}",
                f"- M2 总容忍量：{uta.get('m2_total_slack', 0)}",
                f"- 训练/测试准确率：{uta.get('training_accuracy')} / {uta.get('test_accuracy')}",
                f"- 交叉验证准确率：{uta.get('cross_validation', {}).get('accuracy')}",
                f"- 整方案留出准确率：{uta.get('whole_scheme_holdout', {}).get('accuracy')}",
                f"- 泛化状态：{uta.get('generalization_status')}",
                "",
                "## 未知方案评估",
                "",
            ]
        )
        evaluations = state.get("unknown_evaluations", [])
        if not evaluations:
            lines.append("暂无未知方案评估记录。")
        else:
            lines.extend(["| 编号 | 可行状态 | P可行 | 效能分 | 效能来源 |", "|---|---|---:|---:|---|"])
            for item in evaluations:
                result = item.get("result", {})
                lines.append(
                    f"| {item.get('id')} | {result.get('status')} | {result.get('learned_feasibility_probability')} "
                    f"| {result.get('effectiveness_score')} | {result.get('effectiveness_source')} |"
                )
        lines.extend(
            [
                "",
                "## 使用边界",
                "",
                "- 条件经验带和二维切片是当前数据下的学习轮廓，不是已知物理硬定律。",
                "- 数据量、独立验证或专家证据不足时，效能分只代表阶段性趋势。",
                "- 自由文本原因目前保存但不会未经确认自动改变物理约束。",
            ]
        )
        return "\n".join(lines) + "\n"

    def export_workbook_bytes(self, state: Dict[str, Any], datasets: Dict[str, List[Dict[str, Any]]]) -> bytes:
        workbook = Workbook()
        workbook.remove(workbook.active)
        overview = [
            {
                "项目": self.project.project_name,
                "Excel": self.project.workbook_path,
                "工作簿指纹": self.project.workbook_fingerprint,
                "数据模式": state.get("data_mode"),
                "属性数": len(self.attributes),
                "已有方案数": len(state.get("existing_samples", [])),
                "耦合数": len(self.project.couplings),
                "有效偏好数": state.get("uta_model", {}).get("evidence_pairs", 0),
                "UTA状态": state.get("uta_model", {}).get("status"),
                "泛化状态": state.get("uta_model", {}).get("generalization_status"),
            }
        ]
        sheet_specs = [
            ("概览", overview),
            ("方案", datasets["schemes"]),
            ("专家交互", datasets["interactions"]),
            ("可行性证据", datasets["feasibility_evidence"]),
            ("偏好证据", datasets["preference_evidence"]),
            ("复核历史", datasets["review_history"]),
            ("未知方案评估", datasets["unknown_evaluations"]),
            ("UTA边际效用", datasets["uta_marginal_curves"]),
            ("耦合关系", datasets["couplings"]),
        ]
        for title, rows in sheet_specs:
            sheet = workbook.create_sheet(title)
            if not rows:
                sheet.append(["状态"])
                sheet.append(["暂无数据"])
                continue
            headers: List[str] = []
            for row in rows:
                for key in row:
                    if key not in headers:
                        headers.append(key)
            sheet.append(headers)
            for row in rows:
                sheet.append([self.serializable_cell(row.get(key)) for key in headers])
            for cell in sheet[1]:
                cell.font = Font(color="FFFFFF", bold=True)
                cell.fill = PatternFill("solid", fgColor="0F766E")
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for column in sheet.columns:
                max_length = max(len(str(cell.value or "")) for cell in list(column)[:200])
                sheet.column_dimensions[column[0].column_letter].width = min(max(max_length + 2, 10), 38)
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def export_bundle(self) -> Tuple[bytes, str]:
        state = self.load_state()
        datasets = self.export_datasets(state)
        package = BytesIO()
        with zipfile.ZipFile(package, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            archive.writestr("01_project_and_state.json", json.dumps(self.export(), ensure_ascii=False, indent=2))
            archive.writestr("02_summary_report.md", self.report_markdown(state))
            archive.writestr("03_effectiveness_report.xlsx", self.export_workbook_bytes(state, datasets))
            for index, (name, rows) in enumerate(datasets.items(), start=4):
                archive.writestr(f"{index:02d}_{name}.csv", "\ufeff" + self.csv_text(rows))
        filename = f"effectiveness_report_{self.project.workbook_fingerprint}.zip"
        return package.getvalue(), filename

    def export(self) -> Dict[str, Any]:
        state = self.load_state()
        return {
            "exported_at": now_iso(),
            "project": self.project.to_dict(),
            "state": state,
            "method_status": {
                "completed_stage": 8,
                "current_scoring": "online_bt_and_additive_piecewise_lp_uta",
                "active": [
                    "three_stage_design_order",
                    "monotonic_coupling_surrogate",
                    "conditional_experience_band",
                    "coupling_aware_generation",
                    "direction_only_coupling_fallback",
                    "expert_monotone_frontier_evidence",
                    "automatic_unit_free_scaling",
                    "sample_extrema_experience_core",
                    "latin_hypercube_global_exploration",
                    "maximin_history_distance_exploration",
                    "adaptive_multiscale_local_search",
                    "generation_range_excel_writeback",
                    "expert_learned_feasibility",
                    "separated_evidence_ledger",
                    "online_constrained_feature_bt",
                    "lp_uta_m1_consistency",
                    "lp_uta_m2_minimum_slack",
                    "deterministic_train_test_holdout",
                    "m3_minimum_conflict_localization",
                    "pair_cross_validation",
                    "whole_scheme_holdout",
                    "uta_score_stability_interval",
                    "actionable_review_workflow",
                    "uta_marginal_utility_visualization",
                    "scheme_contribution_comparison",
                    "score_stability_interval_visualization",
                    "conditional_feasibility_slice",
                    "fixed_labeled_demo_state",
                    "automatic_state_backup",
                    "json_csv_excel_markdown_export_bundle",
                    "repeatable_end_to_end_acceptance",
                ],
                "not_yet_active": ["multi_expert_reliability", "document_knowledge_extraction", "production_authentication"],
            },
        }


HTML = r"""<!doctype html>
<html lang="zh-CN"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>参数方案比较与效能评估</title>
<style>
*{box-sizing:border-box}body{margin:0;font-family:Segoe UI,Microsoft YaHei,Arial,sans-serif;background:#f4f7fb;color:#172033}header{background:#fff;border-bottom:1px solid #dbe4ee;padding:16px 24px;position:sticky;top:0;z-index:5}h1{margin:0;font-size:22px;letter-spacing:0}.sub{margin-top:5px;color:#64748b;font-size:13px;line-height:1.5;overflow-wrap:anywhere}.header-row{max-width:1500px;margin:auto;display:flex;align-items:flex-start;justify-content:space-between;gap:16px}.header-actions{display:flex;gap:8px;flex-wrap:wrap}main{max-width:1500px;margin:0 auto;padding:18px;display:grid;grid-template-columns:minmax(0,1fr) 390px;gap:16px}.stats{display:grid;grid-template-columns:repeat(auto-fit,minmax(112px,1fr));gap:9px;margin-bottom:12px}.stat,.panel,.scheme{background:#fff;border:1px solid #dbe4ee;border-radius:8px;padding:13px}.num{font-size:21px;font-weight:800;color:#0f766e}.cap,.meta{font-size:12px;color:#64748b;line-height:1.5}.pair{display:grid;grid-template-columns:minmax(0,1fr) minmax(0,1fr);gap:13px}.scheme{display:flex;flex-direction:column}.scheme h2,.panel h2{margin:0 0 8px;font-size:18px;letter-spacing:0}.scheme-top{flex:0 0 auto}.notice{border:1px solid #dbe4ee;background:#f8fafc;border-radius:7px;padding:8px 9px;margin:8px 0;font-size:12px;line-height:1.55}.notice.warn{border-color:#f3d39b;background:#fffbeb;color:#78350f}.risk{display:flex;gap:7px;align-items:center;margin:8px 0 10px;flex-wrap:wrap}.badge{border-radius:999px;padding:4px 8px;font-size:12px;font-weight:700}.ok{background:#dcfce7;color:#166534}.bad{background:#fee2e2;color:#991b1b}.mid{background:#fef3c7;color:#92400e}table{width:100%;border-collapse:collapse;font-size:13px;table-layout:fixed}td{padding:7px 6px;border-bottom:1px solid #edf2f7;text-align:right;vertical-align:middle}td:first-child{text-align:left;color:#334155;width:67%}.param-table tr{height:58px}.p-label{display:block}.p-unit{display:block;color:#64748b;font-size:11px;margin-top:2px}.feedback{margin-top:13px}.feedback-grid{display:grid;grid-template-columns:1fr 1fr;gap:14px}.feedback-side{min-width:0}.choice,.checks,.actions{display:flex;gap:7px;flex-wrap:wrap;margin:7px 0 9px}.choice label{border:1px solid #cbd5e1;border-radius:7px;padding:7px 9px;cursor:pointer;background:#fff;font-size:13px}.checks label{font-size:12px;background:#f8fafc;border:1px solid #dbe4ee;border-radius:999px;padding:5px 8px;cursor:help}.pref-row{padding-top:9px;border-top:1px solid #edf2f7;margin-top:10px}textarea{width:100%;min-height:68px;resize:vertical;border:1px solid #cbd5e1;border-radius:7px;padding:8px;font-family:inherit;line-height:1.5}button{border:none;border-radius:7px;background:#0f766e;color:white;padding:9px 12px;font-weight:800;cursor:pointer}button.secondary{background:#334155}button.warn{background:#b45309}select{border:1px solid #cbd5e1;border-radius:7px;background:#fff;padding:6px 8px;color:#172033}.setting-row{display:flex;align-items:center;justify-content:space-between;gap:9px;margin:8px 0}.review-actions{display:flex;gap:5px;flex-wrap:wrap;margin-top:7px}.review-actions button{padding:5px 7px;font-size:11px}.side{display:flex;flex-direction:column;gap:13px}.eval-grid{display:grid;grid-template-columns:1fr 1fr;gap:8px}.field label{display:block;font-size:12px;color:#64748b;margin-bottom:3px}.field input{width:100%;border:1px solid #cbd5e1;border-radius:7px;padding:7px}.eval-result{background:#f8fafc;border:1px solid #dbe4ee;border-radius:7px;padding:11px;font-size:13px;line-height:1.55;margin-top:10px;max-height:430px;overflow:auto}.list{margin:6px 0 0;padding-left:18px}.recent,.event,.coupling{font-size:12px;color:#64748b;line-height:1.5;border-bottom:1px solid #edf2f7;padding:7px 0}.event b,.coupling b{color:#172033}.empty{color:#64748b;font-size:12px}.error{color:#991b1b;font-weight:700}.section-title{font-weight:800;margin-top:10px}details.panel{padding:0}details.panel summary{cursor:pointer;padding:13px;font-weight:800}details.panel .details-body{padding:0 13px 13px}.stage{display:inline-block;background:#e0f2fe;color:#075985;border-radius:999px;padding:3px 7px;font-size:11px;font-weight:800;margin-left:5px}@media(max-width:1120px){main{grid-template-columns:1fr}.pair{grid-template-columns:1fr}.feedback-grid{grid-template-columns:1fr}.stats{grid-template-columns:repeat(3,1fr)}}@media(max-width:620px){header{padding:13px}.header-row{display:block}.header-actions{margin-top:9px}main{padding:10px}.stats{grid-template-columns:repeat(2,1fr)}.eval-grid{grid-template-columns:1fr}}
.visual-panel{margin-top:13px}.viz-head{display:flex;align-items:center;justify-content:space-between;gap:10px;flex-wrap:wrap}.viz-tabs{display:flex;gap:5px;flex-wrap:wrap}.viz-tabs button{background:#e2e8f0;color:#334155;padding:7px 9px}.viz-tabs button[aria-selected="true"]{background:#0f766e;color:#fff}.viz-controls{display:flex;gap:10px;align-items:end;flex-wrap:wrap;margin:10px 0}.viz-controls .field{min-width:150px;flex:0 1 220px}.viz-controls select{width:100%}.chart-grid{display:grid;grid-template-columns:minmax(0,1.25fr) minmax(0,1fr);gap:16px}.chart-host{min-width:0}.chart-title{font-size:13px;font-weight:800;margin:8px 0 5px}.chart-svg{display:block;width:100%;height:auto;overflow:visible}.viz-note{font-size:12px;color:#64748b;line-height:1.55;margin:6px 0}.viz-metrics{display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:10px;margin:8px 0 12px}.viz-metric{border-bottom:2px solid #dbe4ee;padding:4px 0 8px}.viz-metric b{display:block;font-size:18px;color:#0f766e}.viz-legend{display:flex;gap:12px;flex-wrap:wrap;font-size:11px;color:#64748b;margin:6px 0}.legend-mark{display:inline-block;width:10px;height:10px;margin-right:4px;vertical-align:-1px}.viz-empty{min-height:190px;display:flex;align-items:center;justify-content:center;color:#64748b;font-size:13px}.slice-detail{font-size:12px;color:#334155;min-height:20px;margin-top:5px}.hidden{display:none}.infeasible-area{margin-top:8px;padding-top:7px;border-top:1px solid #edf2f7}.scheme-details{margin:8px 0 4px;font-size:12px;color:#64748b}.scheme-details summary{cursor:pointer;color:#475569}.plain-reference{margin:8px 0;font-size:13px;line-height:1.55}.plain-reference b{color:#0f766e}@media(max-width:780px){.chart-grid{grid-template-columns:1fr}.viz-metrics{grid-template-columns:1fr 1fr}}@media(max-width:480px){.viz-metrics{grid-template-columns:1fr}.viz-tabs{width:100%}.viz-tabs button{flex:1 1 auto}}
 body{font-size:17px;line-height:1.55}h1{font-size:27px}.sub{font-size:15px}.header-row,main{max-width:1660px}main{grid-template-columns:minmax(0,1fr) 430px;gap:18px}.stat,.panel,.scheme{padding:16px}.num{font-size:25px}.cap,.meta{font-size:14px}.scheme h2,.panel h2{font-size:21px}.notice{font-size:14px;padding:10px 11px}.badge{font-size:14px;padding:5px 9px}table{font-size:16px}td{padding:9px 8px}.param-table tr{height:70px}.p-unit{font-size:13px}.choice label{font-size:15px;padding:9px 11px}.checks label{font-size:14px;padding:7px 10px}textarea{min-height:82px;font-size:15px;padding:10px}button{font-size:15px;padding:11px 14px}select{font-size:15px;padding:8px 10px}.review-actions button{font-size:13px;padding:7px 9px}.field label{font-size:14px}.field input{font-size:15px;padding:9px}.eval-result{font-size:15px;line-height:1.65}.recent,.event,.coupling{font-size:14px}.empty{font-size:14px}details.panel summary{font-size:16px;padding:16px}.stage{font-size:13px}.chart-title{font-size:16px}.viz-note{font-size:14px}.viz-metric b{font-size:22px}.viz-legend{font-size:13px}.viz-empty{font-size:15px}.slice-detail{font-size:14px}.scheme-details{font-size:14px}.plain-reference{font-size:15px}.quick-coupling{display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-top:8px}.quick-coupling .field{min-width:0}.quick-coupling select{width:100%}.range-row{border-bottom:1px solid #edf2f7;padding:9px 0}.range-name{font-weight:800}.range-core{font-size:13px;color:#64748b}.range-inputs{display:grid;grid-template-columns:1fr 1fr;gap:7px;margin-top:6px}.range-inputs input{width:100%;font-size:14px;padding:7px;border:1px solid #cbd5e1;border-radius:7px}@media(max-width:1280px){main{grid-template-columns:1fr}}@media(max-width:620px){body{font-size:16px}h1{font-size:24px}.param-table tr{height:68px}.quick-coupling{grid-template-columns:1fr}}
</style></head><body>
<header><div class="header-row"><div><h1>方案对比评估 <span class="stage">专家判断界面</span></h1><div class="sub" id="projectMeta">正在读取项目...</div></div><div class="header-actions"><button class="secondary" id="refreshBtn">刷新</button><button class="secondary" id="demoBtn">载入演示数据</button><button class="secondary" id="exportBtn">导出数据</button><button id="bundleBtn">导出报告</button><button class="warn" id="resetBtn">清空判断</button></div></div></header>
<main><section><div id="modeBanner"></div><div class="stats" id="stats"></div><div class="pair"><article class="scheme" id="schemeA"></article><article class="scheme" id="schemeB"></article></div>
<form class="panel feedback" id="feedbackForm"><h2>请完成本轮判断</h2><div class="feedback-grid"><div class="feedback-side"><div><b>方案 A 能否实现？</b></div><div class="choice" id="feasibilityA"></div><div class="infeasible-area hidden" id="reasonAreaA"><div class="meta">系统能自动带出的信息已经预选，正确时无需额外操作。</div><div class="quick-coupling" id="quickCouplingA"></div><details class="scheme-details"><summary>补充其他原因（可选）</summary><div class="checks" id="reasonsA"></div><textarea id="reasonA" placeholder="还需要说明时再填写。"></textarea></details></div></div><div class="feedback-side"><div><b>方案 B 能否实现？</b></div><div class="choice" id="feasibilityB"></div><div class="infeasible-area hidden" id="reasonAreaB"><div class="meta">系统能自动带出的信息已经预选，正确时无需额外操作。</div><div class="quick-coupling" id="quickCouplingB"></div><details class="scheme-details"><summary>补充其他原因（可选）</summary><div class="checks" id="reasonsB"></div><textarea id="reasonB" placeholder="还需要说明时再填写。"></textarea></details></div></div></div><div class="pref-row hidden" id="preferenceArea"><div><b>两个方案都能实现，您更推荐哪个？</b></div><div class="choice" id="preference"></div></div><div class="actions"><button type="submit">提交判断，查看下一组</button><span class="meta" id="msg"></span></div></form><details class="panel visual-panel"><summary>查看系统学到了什么（可选）</summary><div class="details-body"><div class="viz-tabs" role="tablist" aria-label="结果图表"><button type="button" data-viz-view="uta" aria-selected="true">为什么分数不同</button><button type="button" data-viz-view="stability" aria-selected="false">评分是否稳定</button><button type="button" data-viz-view="feasibility" aria-selected="false">参数可行范围</button></div><div id="vizControls" class="viz-controls"></div><div id="vizContent"></div></div></details></section>
<aside class="side"><section class="panel"><h2>当前任务</h2><div id="projectInfo"></div></section><section class="panel"><h2>评估新方案</h2><div class="eval-grid" id="evalGrid"></div><div class="actions"><button id="evalBtn">开始评估</button></div><div id="evalResult"></div></section><details class="panel"><summary>调整方案生成范围</summary><div class="details-body"><div class="meta">样本最小值到最大值由系统自动读取。这里只调整允许生成方案的外层范围，保存后会直接写入 Excel。</div><div id="rangeEditor"></div><div class="actions"><button type="button" id="saveRangesBtn">保存到 Excel</button><span class="meta" id="rangeMsg"></span></div></div></details><details class="panel"><summary>需要重新确认的判断 <span id="reviewCount"></span></summary><div class="details-body" id="reviews"></div></details><details class="panel"><summary>最近提交</summary><div class="details-body" id="recent"></div></details><details class="panel"><summary>研究人员：模型与算法详情</summary><div class="details-body" id="modelDetails"></div></details><details class="panel"><summary>研究人员：后台记录 <span id="eventCount"></span></summary><div class="details-body" id="events"></div></details></aside></main>
<script>
let appState=null;
let visualView='comparison';
let selectedCurveKey=null;
let sliceXKey=null;
let sliceYKey=null;
let sliceCache=null;
let sliceRequest=0;
const esc=v=>String(v??'').replace(/[&<>'"]/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;',"'":'&#39;','"':'&quot;'}[c]));
const entries=()=>Object.entries(appState.params||{});
const label=(group,key)=>appState.labels?.[group]?.[key]||key;
const utaStatusLabel=status=>({no_data:'暂无偏好数据',insufficient_data:'数据不足',preliminary:'初步模型',validated:'已通过留出验证',needs_review:'需要复核'})[status]||status;
const generalizationLabel=status=>({insufficient_evidence:'证据不足，不能判断泛化',pair_validation_only:'仅完成比较对验证',whole_scheme_supported:'整方案留出支持',unstable_on_holdout:'留出验证不稳定'})[status]||status;
const reviewActionLabel={keep:'维持原判断',tie:'改为差不多',reverse:'改成相反结果',withdraw:'仅作废这次判断',undo:'撤销上次处理'};
function percentText(value){return value==null?'暂无':`${Math.round(Number(value)*100)}%`;}
function plainFeasibility(probability,status){if(status==='infeasible_by_range'||status==='likely_infeasible_learned')return '可能无法实现';if(status==='likely_feasible_extrapolation')return '较可能实现，但属于经验外推';if(status==='uncertain_feasibility'||status==='coupling_outside_experience'||status==='coupling_boundary')return '还需要专家确认';if(Number(probability)>=0.75)return '较可能实现';return '初步看起来可以实现';}
function renderCompletion(){if(!appState)return;const parent=document.getElementById('projectInfo');if(!parent)return;let host=document.getElementById('completionPanel');if(!host){host=document.createElement('div');host.id='completionPanel';parent.appendChild(host);}const c=appState.learningCompletion||{},v=c.validation_components||{},accuracy=percentText(c.validation_accuracy),consistency=percentText(c.consistency),protocol=appState.activeProtocol,cls=c.complete?'notice':'notice warn',validationDetail=`测试集 ${percentText(v.test_accuracy)} · 交叉验证 ${percentText(v.cross_validation_accuracy)} · 整方案留出 ${percentText(v.whole_scheme_holdout_accuracy)}；保守值取其中最低项`;host.innerHTML=`<div class="${cls}"><b>${esc(c.title||'通用效能模型正在学习')}</b><br>${esc(c.message||'')}<br><span class="meta">通用有效比较：${esc(c.generic_preference_pairs??c.exact_context_pairs??0)} 对<br>保守验证正确率：${esc(accuracy)}（目标 80%） · 判断拟合一致度：${esc(consistency)}（目标 80%）<br>${esc(validationDetail)}${c.complete?'':`<br>建议下一检查点：再获得约 ${esc(c.suggested_additional_effective_pairs||5)} 对有效比较`}<br>当前评分参考：${esc(protocol?.name||'未设置')}（协议本身=100分）。更换协议只会改变评分，不会重新训练。</span></div>`;}
function plainModelState(uta,openReviews){setTimeout(renderCompletion,0);const c=appState?.learningCompletion||{};if(c.complete)return c.title;if(openReviews>0)return '有判断需要重新确认';if(uta.status==='validated')return '已有较多判断，结果相对稳定';if((uta.evidence_pairs||0)===0)return '刚开始学习，所有分数仅供参考';return '正在学习，判断越多结果越稳定';}
function confidenceWord(text){const value=String(text||'');if(value.startsWith('较高'))return '较可信';if(value.startsWith('中等'))return '可以参考';if(value.startsWith('初步'))return '初步结果';return '仅供参考';}
function renderEvaluationReport(record){const r=record.result||{},req=r.requirement_assessment||{},attrs=req.attributes||[],direction={higher_better:'越大越好',lower_better:'越小越好',target:'达到指定值即可',at_least:'越大越好',at_most:'越小越好'},below=attrs.filter(x=>Number(x.relative_score_percent)<100).map(x=>`<li><b>${esc(x.attribute_label)}</b>：${esc(x.explanation)}</li>`).join(''),ledger=[...attrs].sort((a,b)=>Number(b.weighted_score)-Number(a.weighted_score)).map(x=>`<tr><td>${esc(x.attribute_label)}<br><span class="meta">第 ${esc(x.design_stage)} 批 · 权重 ${(Number(x.weight)*100).toFixed(1)}%</span></td><td>${esc(x.value)} ${esc(x.unit)}</td><td>${esc(x.reference_value)} ${esc(x.unit)}<br><span class="meta">${esc(direction[x.requirement_type]||x.requirement_label)}</span></td><td>${Number(x.relative_score_percent).toFixed(1)} 分</td><td>${(Number(x.weighted_score)*100).toFixed(2)} 分</td></tr>`).join(''),boundaries=(r.learned_boundary_violations||[]).map(x=>`<li>${esc(x.message)}（边界置信度 ${Math.round(Number(x.confidence)*100)}%）</li>`).join(''),mechanisms=(r.coupling_assessments||[]).flatMap(x=>(x.source_contributions||[]).slice(0,3).map(s=>`<li>${esc(s.source_label)}对${esc(x.target_label)}为${s.direction==='positive'?'正向':'负向'}影响；按当前单调代理，每增加 1 ${esc(s.source_unit)}，${esc(x.target_label)}约变化 ${esc(s.physical_coefficient)} ${esc(s.target_unit)}。</li>`)).join('');if(!req.profile_id)return renderEvaluation(record);return `${renderEvaluation(record)}<div class="eval-result"><div class="section-title">相对新技术协议的效能结论</div><div class="plain-reference"><b>${esc(r.reuse_recommendation||req.decision_label||'暂无')}</b><br>当前方案：<b>${esc(req.coverage_percent)} 分</b>；“${esc(req.profile_name)}”本身固定为 <b>100 分</b>。<br><span class="meta">评分权重来源：${esc(req.weight_source)}。新协议只提供参考值，没有参与BT/UTA学习。</span></div>${below?`<div class="section-title">低于100分参考的属性</div><ul class="list">${below}</ul>`:'<div class="meta">各属性均达到或超过方向允许的100分参考。</div>'}${boundaries?`<div class="section-title">专家可行边界</div><ul class="list">${boundaries}</ul>`:''}${mechanisms?`<div class="section-title">耦合机理依据</div><ul class="list">${mechanisms}</ul>`:''}<details class="scheme-details" open><summary>查看每项相对得分和总分贡献</summary><table><thead><tr><th>属性</th><th>当前方案</th><th>协议100分参考</th><th>属性相对分</th><th>总分贡献</th></tr></thead><tbody>${ledger}</tbody></table></details></div>`;}
function formatValue(value,spec){const n=Number(value);if(spec.data_type==='categorical'||!Number.isFinite(n))return esc(value);return n.toFixed(Number(spec.precision??3));}
function badge(status){const cls=status==='infeasible_by_range'||status==='likely_infeasible_learned'?'bad':status==='uncertain_feasibility'||status==='coupling_outside_experience'||status==='coupling_boundary'?'mid':'ok';return `<span class="badge ${cls}">${esc(label('status',status))}</span>`;}
function couplingRows(items){return (items||[]).map(x=>`<div class="meta ${x.status==='below_band'||x.status==='above_band'?'error':''}">· ${esc(x.target_label)}：实际 ${esc(x.actual)}，条件带 [${esc(x.lower)}, ${esc(x.upper)}] ${esc(x.target_unit)}（${esc(x.status)}）</div>`).join('');}
function renderScheme(s){const p=s.prediction||{};const rows=entries().map(([key,spec])=>`<tr><td><span class="p-label">${esc(spec.label)}</span><span class="p-unit">第 ${esc(spec.design_stage)} 批设计 · 单位：${esc(spec.unit)}</span></td><td>${formatValue(s.params[key],spec)}</td></tr>`).join('');const violations=(p.hard_violations||[]).map(v=>`<div class="meta error">· ${esc(v.title)}：${esc(v.message)}</div>`).join('');const coupling=couplingRows(p.coupling_assessments);const seq=(s.generation_explanation?.design_sequence||[]).map(x=>`第${x.stage}批：${(x.attributes||[]).join('、')}`).join('；');const selection=s.generation_explanation?.selection_strategy?`候选选择：${esc(s.generation_explanation.selection_strategy)}；候选数 ${esc(s.generation_explanation.candidate_pool_size)}，距最近历史方案 ${esc(s.generation_explanation.normalized_novelty)}，不确定性 ${esc(s.generation_explanation.bt_uncertainty)}。`:'';const technicalGeneration=s.generation_explanation?`<div><b>生成详情</b><br>${esc(s.generation_explanation.detail)}${seq?`<br>设计顺序：${esc(seq)}`:''}${selection?`<br>${selection}`:''}</div>`:'';const purpose=s.generation_explanation?.purpose||'';const interval=p.uta_score_interval?`，区间 ${esc(p.uta_score_interval[0])}～${esc(p.uta_score_interval[1])}`:'';const uta=p.uta_score===null||p.uta_score===undefined?'尚无效能学习结果':`分段效用分 ${esc(p.uta_score)}（${esc(utaStatusLabel(p.uta_status))}${interval}）`;const score=p.uta_score??p.bt_score;return `<div class="scheme-top"><h2>方案 ${esc(s.id)}</h2>${purpose?`<div class="notice">${esc(purpose)}</div>`:''}<div class="plain-reference">系统参考：<b>${esc(plainFeasibility(p.learned_feasibility_probability,p.status))}</b><br>通用效能参考分：<b>${esc(score)}</b> <span class="meta">不使用新技术协议；判断较少时只看趋势</span></div>${violations}<details class="scheme-details"><summary>查看系统分析</summary><div class="notice"><div>来源：${esc(label('source',s.source))} · ${esc(s.source_detail||'')}</div>${technicalGeneration}<div>P可行 ${esc(p.learned_feasibility_probability)}，风险 ${esc(p.risk)}，BT参考分 ${esc(p.bt_score)}，${uta}。</div>${coupling}</div></details></div><table class="param-table"><tbody>${rows}</tbody></table>`;}
function alignSchemeCards(){if(matchMedia('(max-width:1120px)').matches)return;const cards=[document.getElementById('schemeA'),document.getElementById('schemeB')];if(cards.some(c=>!c))return;const tops=cards.map(c=>c.querySelector('.scheme-top'));if(tops.some(t=>!t))return;tops.forEach(t=>t.style.minHeight='0px');const topHeight=Math.max(...tops.map(t=>t.getBoundingClientRect().height));tops.forEach(t=>t.style.minHeight=`${Math.ceil(topHeight)}px`);const groups=cards.map(c=>Array.from(c.querySelectorAll('.param-table tr')));if(groups.some(g=>g.length===0))return;for(let i=0;i<Math.min(...groups.map(g=>g.length));i++){groups.forEach(g=>g[i].style.height='auto');const h=Math.max(...groups.map(g=>g[i].getBoundingClientRect().height),70);groups.forEach(g=>g[i].style.height=`${Math.ceil(h)}px`);}}
function choice(container,name,items,selected){document.getElementById(container).innerHTML=items.map(([value,text])=>`<label><input type="radio" name="${esc(name)}" value="${esc(value)}" ${value===selected?'checked':''}> ${esc(text)}</label>`).join('');}
function reasons(container,name){document.getElementById(container).innerHTML=Object.entries(appState.reasonOptions||{}).map(([code,item])=>`<label title="${esc(item.description)}"><input type="checkbox" name="${esc(name)}" value="${esc(code)}"> ${esc(item.title)}</label>`).join('');}
function renderQuickCoupling(side,scheme){const host=document.getElementById(`quickCoupling${side}`),targets=[];for(const item of appState.couplings||[]){if(!targets.some(x=>x.key===item.target_key))targets.push({key:item.target_key,label:item.target_label});}if(!targets.length){host.innerHTML='<div class="meta">当前项目没有配置耦合关系，可直接补充文字原因。</div>';return;}const explanation=scheme.generation_explanation||{},suggested=explanation.probe_target||'',suggestedSide=explanation.probe_side==='below'?'target_low':explanation.probe_side==='above'?'target_high':'mismatch';const targetOptions=[['','系统暂未定位'],...targets.map(x=>[x.key,x.label])].map(([value,text])=>`<option value="${esc(value)}" ${value===suggested?'selected':''}>${esc(text)}</option>`).join('');const issueOptions=[['mismatch','参数搭配不合理'],['target_low','目标参数偏低'],['target_high','目标参数偏高']].map(([value,text])=>`<option value="${value}" ${value===suggestedSide?'selected':''}>${text}</option>`).join('');host.innerHTML=`<div class="field"><label for="couplingTarget${side}">主要是哪项配合问题？</label><select id="couplingTarget${side}">${targetOptions}</select></div><div class="field"><label for="couplingIssue${side}">大概是什么情况？</label><select id="couplingIssue${side}">${issueOptions}</select></div>`;}
function readCouplingFeedback(side){const target=document.getElementById(`couplingTarget${side}`)?.value||'';if(!target)return null;return {target_key:target,issue:document.getElementById(`couplingIssue${side}`)?.value||'mismatch'};}
function syncReasonArea(feasibilityName,areaId,reasonName,textId){const selected=document.querySelector(`input[name="${feasibilityName}"]:checked`)?.value,area=document.getElementById(areaId),show=selected==='infeasible';area.classList.toggle('hidden',!show);if(!show){document.querySelectorAll(`input[name="${reasonName}"]`).forEach(input=>input.checked=false);document.getElementById(textId).value='';}}
function syncPreferenceArea(){const a=document.querySelector('input[name="feasibilityA"]:checked')?.value,b=document.querySelector('input[name="feasibilityB"]:checked')?.value,show=a==='feasible'&&b==='feasible';document.getElementById('preferenceArea').classList.toggle('hidden',!show);if(!show)document.querySelectorAll('input[name="preference"]').forEach(input=>input.checked=false);}
function bindReasonAreas(){[['feasibilityA','reasonAreaA','reasonCodesA','reasonA'],['feasibilityB','reasonAreaB','reasonCodesB','reasonB']].forEach(args=>{document.querySelectorAll(`input[name="${args[0]}"]`).forEach(input=>input.onchange=()=>{syncReasonArea(...args);syncPreferenceArea();});syncReasonArea(...args);});syncPreferenceArea();}
function renderRangeEditor(){const rows=entries().filter(([,spec])=>spec.data_type!=='categorical'&&spec.participates_generation).map(([key,spec])=>`<div class="range-row"><div class="range-name">${esc(spec.label)} <span class="meta">${esc(spec.unit)}</span></div><div class="range-core">样本经验范围：${esc(spec.feasible_min)} ～ ${esc(spec.feasible_max)}；自动尺度：${esc(spec.scale_profile?.mode||'linear')}</div><div class="range-inputs"><input id="rangeMin_${esc(key)}" type="number" step="any" value="${esc(spec.min)}" aria-label="${esc(spec.label)}生成下限"><input id="rangeMax_${esc(key)}" type="number" step="any" value="${esc(spec.max)}" aria-label="${esc(spec.label)}生成上限"></div></div>`).join('');document.getElementById('rangeEditor').innerHTML=rows||'<div class="empty">没有可调整的数值属性。</div>';document.getElementById('saveRangesBtn').onclick=async e=>{if(!confirm('确定把新的生成范围写入当前 Excel 吗？软件会自动备份工作簿并保留全部判断。'))return;const ranges={};entries().forEach(([key,spec])=>{if(spec.data_type!=='categorical'&&spec.participates_generation)ranges[key]={min:document.getElementById(`rangeMin_${key}`).value,max:document.getElementById(`rangeMax_${key}`).value};});const button=e.currentTarget,msg=document.getElementById('rangeMsg');button.disabled=true;msg.textContent='正在保存...';try{render(await api('/api/generation-ranges',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({ranges})}));document.getElementById('rangeMsg').textContent='已写入 Excel，并生成了新的方案。';}catch(err){msg.textContent=err.message;button.disabled=false;}};}
function checked(name){return Array.from(document.querySelectorAll(`input[name="${name}"]:checked`)).map(x=>x.value);}
const nfmt=(value,digits=2)=>Number.isFinite(Number(value))?Number(value).toFixed(digits):'暂无';
function parameterComparisonSvg(){const a=appState.pair.a,b=appState.pair.b,specs=entries().filter(([,spec])=>spec.data_type!=='categorical'),width=900,left=190,right=150,top=55,rowH=58,plot=width-left-right,height=top+specs.length*rowH+42,rows=specs.map(([key,spec],index)=>{const lo=Number(spec.min),hi=Number(spec.max),span=Math.max(hi-lo,1e-9),az=Math.max(0,Math.min(1,(Number(a.params[key])-lo)/span)),bz=Math.max(0,Math.min(1,(Number(b.params[key])-lo)/span)),ax=left+az*plot,bx=left+bz*plot,y=top+index*rowH,connection=`<line x1="${Math.min(ax,bx)}" y1="${y}" x2="${Math.max(ax,bx)}" y2="${y}" stroke="#64748b" stroke-width="4"/>`;return `<text x="0" y="${y-3}" font-size="14" fill="#17324d">${esc(spec.label)}</text><text x="0" y="${y+16}" font-size="11" fill="#64748b">第 ${esc(spec.design_stage)} 批 · ${esc(spec.unit)}</text><line x1="${left}" y1="${y}" x2="${left+plot}" y2="${y}" stroke="#cbd5e1" stroke-width="2"/>${connection}<circle cx="${ax}" cy="${y}" r="8" fill="#0f766e"><title>方案 A：${formatValue(a.params[key],spec)} ${esc(spec.unit)}</title></circle><circle cx="${bx}" cy="${y}" r="8" fill="#2563eb"><title>方案 B：${formatValue(b.params[key],spec)} ${esc(spec.unit)}</title></circle><text x="${width-right+12}" y="${y-4}" font-size="12" fill="#0f766e">A ${formatValue(a.params[key],spec)}</text><text x="${width-right+12}" y="${y+14}" font-size="12" fill="#2563eb">B ${formatValue(b.params[key],spec)}</text>`;}).join('');return `<div class="viz-note">每一行使用该参数自己的生成范围归一化，不受单位和数量级影响。A/B 判断学习的是通用产品效能，不显示也不使用新技术协议。</div><svg class="chart-svg" viewBox="0 0 ${width} ${height}" role="img" aria-label="方案 A 和方案 B 原始参数对比"><title>方案 A 和方案 B 原始参数对比</title><circle cx="${left}" cy="22" r="7" fill="#0f766e"/><text x="${left+12}" y="27" font-size="13" fill="#334155">方案 A</text><circle cx="${left+100}" cy="22" r="7" fill="#2563eb"/><text x="${left+112}" y="27" font-size="13" fill="#334155">方案 B</text>${rows}<text x="${left}" y="${height-9}" font-size="11" fill="#64748b">生成下限</text><text x="${left+plot}" y="${height-9}" text-anchor="end" font-size="11" fill="#64748b">生成上限</text></svg>`;}
function ensureComparisonTab(){const tabs=document.querySelector('.viz-tabs');if(!tabs||tabs.querySelector('[data-viz-view="comparison"]'))return;const button=document.createElement('button');button.type='button';button.dataset.vizView='comparison';button.textContent='A/B 参数对比';button.onclick=()=>{visualView='comparison';renderVisuals();};tabs.insertBefore(button,tabs.firstChild);}
function renderComparisonVisual(){document.getElementById('vizControls').innerHTML='';document.getElementById('vizContent').innerHTML=parameterComparisonSvg();}
function contributionSvg(contributionsA,contributionsB){if(!contributionsA.length&&!contributionsB.length)return '<div class="viz-empty">判断次数还不够，暂时无法解释各参数对分数的影响。</div>';const mapA=Object.fromEntries(contributionsA.map(x=>[x.key,x]));const mapB=Object.fromEntries(contributionsB.map(x=>[x.key,x]));const specs=entries().filter(([key])=>mapA[key]||mapB[key]);const width=760,rowH=42,left=145,right=62,top=32,height=top+specs.length*rowH+30,plot=width-left-right;const maxValue=Math.max(1,...specs.flatMap(([key])=>[Number(mapA[key]?.score_points||0),Number(mapB[key]?.score_points||0)]));const rows=specs.map(([key,spec],index)=>{const a=Number(mapA[key]?.score_points||0),b=Number(mapB[key]?.score_points||0),y=top+index*rowH,aw=plot*a/maxValue,bw=plot*b/maxValue;return `<text x="0" y="${y+16}" font-size="13" fill="#334155">${esc(spec.label)}</text><rect x="${left}" y="${y+2}" width="${aw}" height="11" rx="2" fill="#0f766e"><title>方案 A：${nfmt(a,3)} 分</title></rect><rect x="${left}" y="${y+20}" width="${bw}" height="11" rx="2" fill="#2563eb"><title>方案 B：${nfmt(b,3)} 分</title></rect><text x="${Math.min(left+aw+4,width-right+3)}" y="${y+12}" font-size="11" fill="#475569">${nfmt(a,2)}</text><text x="${Math.min(left+bw+4,width-right+3)}" y="${y+30}" font-size="11" fill="#475569">${nfmt(b,2)}</text>`;}).join('');return `<svg class="chart-svg" viewBox="0 0 ${width} ${height}" role="img" aria-label="方案 A 和方案 B 各参数对分数的影响"><title>方案 A 和方案 B 各参数对分数的影响</title><line x1="${left}" y1="20" x2="${left}" y2="${height-22}" stroke="#cbd5e1"/>${rows}<text x="${left}" y="14" font-size="12" fill="#64748b">0</text><text x="${width-right}" y="14" text-anchor="end" font-size="12" fill="#64748b">${nfmt(maxValue,2)} 分</text></svg>`;}
function marginalCurveSvg(curve){if(!curve)return '<div class="viz-empty">暂时没有可显示的变化趋势。</div>';const width=560,height=340,left=60,right=24,top=26,bottom=78,plotW=width-left-right,plotH=height-top-bottom,points=curve.points||[],maxUtility=Math.max(Number(curve.attribute_weight||0),0.0001);const coords=points.map((point,index)=>({x:left+(points.length===1?0:index/(points.length-1))*plotW,y:top+(1-Number(point.utility)/maxUtility)*plotH,point}));const polyline=coords.map(x=>`${x.x},${x.y}`).join(' ');const ticks=coords.map(({x,y,point})=>`<line x1="${x}" y1="${top+plotH}" x2="${x}" y2="${top+plotH+6}" stroke="#94a3b8"/><text x="${x}" y="${top+plotH+23}" text-anchor="middle" font-size="11" fill="#64748b">${esc(point.raw_value)}</text><circle cx="${x}" cy="${y}" r="5" fill="#0f766e"><title>参数值 ${esc(point.raw_value)} ${esc(curve.unit)}，对总分影响 ${nfmt(point.utility,4)}</title></circle>`).join('');return `<svg class="chart-svg" viewBox="0 0 ${width} ${height}" role="img" aria-label="${esc(curve.label)}变化对参考分的影响"><title>${esc(curve.label)}变化对参考分的影响</title><line x1="${left}" y1="${top}" x2="${left}" y2="${top+plotH}" stroke="#94a3b8"/><line x1="${left}" y1="${top+plotH}" x2="${left+plotW}" y2="${top+plotH}" stroke="#94a3b8"/><line x1="${left}" y1="${top}" x2="${left+plotW}" y2="${top}" stroke="#e2e8f0"/><text x="${left-8}" y="${top+4}" text-anchor="end" font-size="11" fill="#64748b">${nfmt(maxUtility,3)}</text><text x="${left-8}" y="${top+plotH+4}" text-anchor="end" font-size="11" fill="#64748b">0</text><polyline points="${polyline}" fill="none" stroke="#0f766e" stroke-width="3"/>${ticks}<text x="${left+plotW/2}" y="${height-10}" text-anchor="middle" font-size="12" fill="#475569">原始参数值 / ${esc(curve.unit)}</text><text transform="translate(14 ${top+plotH/2}) rotate(-90)" text-anchor="middle" font-size="12" fill="#475569">对总分的影响</text></svg>`;}
function renderUtaVisual(){const uta=appState.utaModel||{},curves=uta.marginal_curves||[];if(!selectedCurveKey||!curves.some(x=>x.key===selectedCurveKey))selectedCurveKey=curves[0]?.key||null;const options=curves.map(x=>`<option value="${esc(x.key)}" ${x.key===selectedCurveKey?'selected':''}>${esc(x.label)}</option>`).join('');document.getElementById('vizControls').innerHTML=options?`<div class="field"><label for="curveAttribute">选择要查看的参数</label><select id="curveAttribute">${options}</select></div><div class="viz-note">系统已参考 ${esc(uta.evidence_pairs||0)} 次有效比较</div>`:'';if(options)document.getElementById('curveAttribute').onchange=e=>{selectedCurveKey=e.target.value;renderUtaVisual();};if(!uta.evidence_pairs){document.getElementById('vizContent').innerHTML='<div class="viz-empty">完成几次有效比较后，这里会显示每个参数对参考分的影响。</div>';return;}const curve=curves.find(x=>x.key===selectedCurveKey);const a=appState.pair.a.prediction?.uta_contributions||[],b=appState.pair.b.prediction?.uta_contributions||[];document.getElementById('vizContent').innerHTML=`<div class="viz-note">这些图只解释系统当前学到的趋势，判断次数较少时请勿当作最终结论。</div><div class="chart-grid"><div class="chart-host"><div class="chart-title">A 和 B 的分数主要来自哪些参数</div><div class="viz-legend"><span><i class="legend-mark" style="background:#0f766e"></i>方案 A</span><span><i class="legend-mark" style="background:#2563eb"></i>方案 B</span></div>${contributionSvg(a,b)}</div><div class="chart-host"><div class="chart-title">${esc(curve?.label||'')}变化时，参考分如何变化</div>${marginalCurveSvg(curve)}</div></div>`;}
function stabilitySvg(items){if(!items.length)return '<div class="viz-empty">至少需要 12 条结构完整的偏好证据，才能形成多模型评分稳定区间。</div>';const data=items.slice(0,10),width=820,rowH=40,left=110,right=50,top=40,height=top+data.length*rowH+38,plot=width-left-right;const grid=[0,25,50,75,100].map(value=>{const x=left+plot*value/100;return `<line x1="${x}" y1="${top-17}" x2="${x}" y2="${height-30}" stroke="#e2e8f0"/><text x="${x}" y="16" text-anchor="middle" font-size="11" fill="#64748b">${value}</text>`;}).join('');const rows=data.map((item,index)=>{const y=top+index*rowH,x1=left+plot*Number(item.p10)/100,x2=left+plot*Number(item.p90)/100,xm=left+plot*Number(item.mean_score)/100;return `<text x="0" y="${y+5}" font-size="12" fill="#334155">${esc(item.scheme_id)}</text><line x1="${x1}" y1="${y}" x2="${x2}" y2="${y}" stroke="#2563eb" stroke-width="5" stroke-linecap="round"><title>P10 ${esc(item.p10)}，P90 ${esc(item.p90)}</title></line><circle cx="${xm}" cy="${y}" r="6" fill="#0f766e"><title>均值 ${esc(item.mean_score)}，区间宽度 ${esc(item.width)}</title></circle><text x="${Math.min(x2+8,width-right+5)}" y="${y+5}" font-size="11" fill="#475569">${esc(item.mean_score)}</text>`;}).join('');return `<svg class="chart-svg" viewBox="0 0 ${width} ${height}" role="img" aria-label="已有方案 UTA 评分稳定区间"><title>已有方案 UTA 评分稳定区间</title>${grid}${rows}<text x="${left+plot/2}" y="${height-6}" text-anchor="middle" font-size="12" fill="#475569">效能评分</text></svg>`;}
function renderStabilityVisual(){const uta=appState.utaModel||{},cv=uta.cross_validation||{},holdout=uta.whole_scheme_holdout||{},a=appState.pair.a.prediction||{},b=appState.pair.b.prediction||{};document.getElementById('vizControls').innerHTML='';const aInterval=a.uta_score_interval?`${a.uta_score_interval[0]}～${a.uta_score_interval[1]}`:'暂无',bInterval=b.uta_score_interval?`${b.uta_score_interval[0]}～${b.uta_score_interval[1]}`:'暂无';document.getElementById('vizContent').innerHTML=`<div class="viz-metrics"><div class="viz-metric"><span class="meta">交叉验证正确率</span><b>${esc(percentText(cv.accuracy))}</b><span class="meta">多次换训练/验证分组后的预测表现</span></div><div class="viz-metric"><span class="meta">整方案留出正确率</span><b>${esc(percentText(holdout.accuracy))}</b><span class="meta">检验没有见过的新方案</span></div><div class="viz-metric"><span class="meta">判断拟合一致度</span><b>${esc(percentText(uta.judgment_consistency))}</b><span class="meta">模型对全部有效判断的拟合程度</span></div></div><div class="viz-note">当前方案 A 的可能分数范围：${esc(aInterval)}；方案 B：${esc(bInterval)}。</div><div class="chart-title">已有方案在不同训练结果下的分数范围</div><div class="viz-legend"><span><i class="legend-mark" style="background:#2563eb"></i>可能范围</span><span><i class="legend-mark" style="background:#0f766e;border-radius:50%"></i>平均分</span></div>${stabilitySvg(uta.rank_stability||[])}`;}
function sliceColor(probability,status){if(status==='infeasible_by_range')return '#991b1b';if(status==='likely_infeasible_learned')return '#dc2626';if(status==='uncertain_feasibility')return '#d97706';if(status==='likely_feasible_extrapolation')return '#0284c7';return '#16a34a';}
function showSliceCell(x,y,probability,status,couplingOutside,outsideSample){const target=document.getElementById('sliceDetail');if(target)target.textContent=`参数值 ${x} / ${y}：预测可行概率约 ${Math.round(Number(probability)*100)}%，${plainFeasibility(probability,status)}${outsideSample?'；超出历史样本经验范围':''}${couplingOutside?`；有 ${couplingOutside} 处参数配合需要检查`:''}。`;}
function feasibilitySliceSvg(data){const width=820,height=570,left=85,right=30,top=28,bottom=84,plotW=width-left-right,plotH=height-top-bottom,res=Number(data.resolution),cellW=plotW/res,cellH=plotH/res;const cells=(data.cells||[]).map(cell=>{const x=left+Number(cell.x_index)*cellW,y=top+(res-1-Number(cell.y_index))*cellH,fill=sliceColor(cell.probability,cell.status);return `<rect x="${x}" y="${y}" width="${cellW+0.4}" height="${cellH+0.4}" fill="${fill}" onclick="showSliceCell(${Number(cell.x)},${Number(cell.y)},${Number(cell.probability)},'${esc(cell.status)}',${Number(cell.coupling_outside)},${cell.outside_sample_experience?1:0})"><title>${esc(data.x_axis.label)} ${esc(cell.x)}，${esc(data.y_axis.label)} ${esc(cell.y)}，预测可行概率 ${esc(cell.probability)}，${esc(label('status',cell.status))}</title></rect>`;}).join('');const xScale=value=>left+(Number(value)-Number(data.x_axis.min))/Math.max(Number(data.x_axis.max)-Number(data.x_axis.min),1e-9)*plotW,yScale=value=>top+plotH-(Number(value)-Number(data.y_axis.min))/Math.max(Number(data.y_axis.max)-Number(data.y_axis.min),1e-9)*plotH;const existing=(data.points||[]).filter(p=>p.source==='existing_sample').map(point=>`<circle cx="${xScale(point.x)}" cy="${yScale(point.y)}" r="3" fill="none" stroke="#0f172a" stroke-width="1.2" opacity=".55"><title>已有方案 ${esc(point.scheme_id)}</title></circle>`).join('');const expert=(data.points||[]).filter(p=>p.source==='expert_evidence').map(point=>{const x=xScale(point.x),y=yScale(point.y);if(point.label==='feasible')return `<circle cx="${x}" cy="${y}" r="6" fill="#2563eb" stroke="#fff" stroke-width="1.5"><title>专家判可行：${esc(point.scheme_id)} / ${esc(point.evidence_id)}</title></circle>`;return `<path d="M ${x-6} ${y-6} L ${x+6} ${y+6} M ${x+6} ${y-6} L ${x-6} ${y+6}" stroke="#111827" stroke-width="3"><title>专家判不可行：${esc(point.scheme_id)} / ${esc(point.evidence_id)}</title></path>`;}).join('');return `<svg class="chart-svg" viewBox="0 0 ${width} ${height}" role="img" aria-label="${esc(data.x_axis.label)}与${esc(data.y_axis.label)}条件可行域切片"><title>${esc(data.x_axis.label)}与${esc(data.y_axis.label)}条件可行域切片</title>${cells}${existing}${expert}<rect x="${left}" y="${top}" width="${plotW}" height="${plotH}" fill="none" stroke="#475569"/><text x="${left}" y="${top+plotH+24}" font-size="11" fill="#64748b">${esc(data.x_axis.min)}</text><text x="${left+plotW}" y="${top+plotH+24}" text-anchor="end" font-size="11" fill="#64748b">${esc(data.x_axis.max)}</text><text x="${left-9}" y="${top+plotH}" text-anchor="end" font-size="11" fill="#64748b">${esc(data.y_axis.min)}</text><text x="${left-9}" y="${top+9}" text-anchor="end" font-size="11" fill="#64748b">${esc(data.y_axis.max)}</text><text x="${left+plotW/2}" y="${height-16}" text-anchor="middle" font-size="13" fill="#334155">${esc(data.x_axis.label)} / ${esc(data.x_axis.unit)}</text><text transform="translate(19 ${top+plotH/2}) rotate(-90)" text-anchor="middle" font-size="13" fill="#334155">${esc(data.y_axis.label)} / ${esc(data.y_axis.unit)}</text></svg>`;}
function renderSliceData(data){const existing=(data.points||[]).filter(x=>x.source==='existing_sample').length,feasible=(data.points||[]).filter(x=>x.source==='expert_evidence'&&x.label==='feasible').length,infeasible=(data.points||[]).filter(x=>x.source==='expert_evidence'&&x.label==='infeasible').length;document.getElementById('vizContent').innerHTML=`<div class="viz-note">只改变横向和纵向两个参数，其余参数固定为方案 ${esc(data.base_scheme_id)}。样本范围外表示经验外推，不等于物理不可行。</div><div class="viz-legend"><span><i class="legend-mark" style="background:#16a34a"></i>较可能可行</span><span><i class="legend-mark" style="background:#0284c7"></i>较可能可行但属于外推</span><span><i class="legend-mark" style="background:#d97706"></i>需要确认</span><span><i class="legend-mark" style="background:#dc2626"></i>已学不可行</span><span>○ 已有方案 ${existing}</span><span style="color:#2563eb">● 专家认为可以 ${feasible}</span><span>× 专家认为不可以 ${infeasible}</span></div>${feasibilitySliceSvg(data)}<div class="slice-detail" id="sliceDetail">点击图中的色块，可以查看这个参数组合的参考判断。</div>`;}
async function loadFeasibilitySlice(force=false){if(!appState||!sliceXKey||!sliceYKey)return;const fm=appState.feasibilityModel||{};const signature=`${appState.pair.a.id}|${fm.expert_samples||0}|${fm.positive_samples||0}|${fm.negative_samples||0}|${sliceXKey}|${sliceYKey}`;if(!force&&sliceCache?.signature===signature){renderSliceData(sliceCache.data);return;}const request=++sliceRequest;document.getElementById('vizContent').innerHTML='<div class="viz-empty">正在计算参数可行范围...</div>';try{const data=await api('/api/feasibility-slice',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({xKey:sliceXKey,yKey:sliceYKey,resolution:20})});if(request!==sliceRequest||visualView!=='feasibility')return;sliceCache={signature,data};renderSliceData(data);}catch(err){if(request===sliceRequest)document.getElementById('vizContent').innerHTML=`<div class="viz-empty error">${esc(err.message)}</div>`;}}
function updateSliceAxes(changed){const x=document.getElementById('sliceX'),y=document.getElementById('sliceY');if(!x||!y)return;if(x.value===y.value){const alternatives=Array.from((changed==='x'?y:x).options).map(o=>o.value).filter(value=>value!==(changed==='x'?x:y).value);if(changed==='x')y.value=alternatives[0];else x.value=alternatives[0];}sliceXKey=x.value;sliceYKey=y.value;sliceCache=null;loadFeasibilitySlice(true);}
function renderFeasibilityVisual(){const numeric=appState.visualization?.numericAttributes||[];if(numeric.length<2){document.getElementById('vizControls').innerHTML='';document.getElementById('vizContent').innerHTML='<div class="viz-empty">至少需要两个数值参数才能绘图。</div>';return;}sliceXKey=sliceXKey&&numeric.includes(sliceXKey)?sliceXKey:appState.visualization.defaultSliceX;sliceYKey=sliceYKey&&numeric.includes(sliceYKey)&&sliceYKey!==sliceXKey?sliceYKey:appState.visualization.defaultSliceY;if(sliceXKey===sliceYKey)sliceYKey=numeric.find(x=>x!==sliceXKey);const options=selected=>numeric.map(key=>`<option value="${esc(key)}" ${key===selected?'selected':''}>${esc(appState.params[key].label)}</option>`).join('');document.getElementById('vizControls').innerHTML=`<div class="field"><label for="sliceX">横向参数</label><select id="sliceX">${options(sliceXKey)}</select></div><div class="field"><label for="sliceY">纵向参数</label><select id="sliceY">${options(sliceYKey)}</select></div><button type="button" id="sliceRefresh">重新计算</button>`;document.getElementById('sliceX').onchange=()=>updateSliceAxes('x');document.getElementById('sliceY').onchange=()=>updateSliceAxes('y');document.getElementById('sliceRefresh').onclick=()=>loadFeasibilitySlice(true);loadFeasibilitySlice();}
function renderVisuals(){document.querySelectorAll('[data-viz-view]').forEach(button=>button.setAttribute('aria-selected',button.dataset.vizView===visualView?'true':'false'));if(visualView==='stability')renderStabilityVisual();else if(visualView==='feasibility')renderFeasibilityVisual();else renderUtaVisual();}
renderVisuals=function(){ensureComparisonTab();document.querySelectorAll('[data-viz-view]').forEach(button=>button.setAttribute('aria-selected',button.dataset.vizView===visualView?'true':'false'));if(visualView==='comparison')renderComparisonVisual();else if(visualView==='stability')renderStabilityVisual();else if(visualView==='feasibility')renderFeasibilityVisual();else renderUtaVisual();};
function renderEvalInputs(){const base=appState.pair.a.params;document.getElementById('evalGrid').innerHTML=entries().map(([key,spec])=>`<div class="field"><label for="eval_${esc(key)}">${esc(spec.label)} (${esc(spec.unit)})</label><input id="eval_${esc(key)}" type="${spec.data_type==='categorical'?'text':'number'}" ${spec.data_type==='categorical'?'':`step="${spec.precision===0?'1':'any'}" min="${spec.min}" max="${spec.max}"`} value="${esc(base[key])}"></div>`).join('');}
function renderEvaluation(record){const r=record.result||{};const violations=(r.hard_violations||[]).map(v=>`<li>${esc(v.title)}：${esc(v.message)}</li>`).join('');const coupling=(r.coupling_assessments||[]).map(x=>`<li>${esc(x.message)}</li>`).join('');const risks=(r.feasibility_risk_contributors||[]).slice(0,3).map(x=>`<li>${esc(x.label)}</li>`).join('');const technicalRisks=(r.feasibility_risk_contributors||[]).map(x=>`<li>${esc(x.label)}：风险贡献 ${esc(x.contribution)}</li>`).join('');const contributions=(r.top_contributors||[]).map(v=>`<li>${esc(v.label)}：标准化表现 ${esc(v.value)}</li>`).join('');const basis=(r.assessment_basis||[]).map(x=>`<li>${esc(x)}</li>`).join('');const interval=r.uta_score_interval?`${esc(r.uta_score_interval[0])}～${esc(r.uta_score_interval[1])}`:'暂无';const probability=Math.round(Number(r.learned_feasibility_probability||0)*100);const problems=violations||risks,hasProtocol=Boolean(r.requirement_assessment?.profile_id),scoreNote=hasProtocol?'（所选新技术协议=100分，可低于或高于100）':'（通用效能参考）';return `<div class="eval-result"><div class="section-title">系统参考结论</div><div class="plain-reference">能否实现：<b>${esc(plainFeasibility(r.learned_feasibility_probability,r.status))}</b><br>可行把握：<b>${esc(probability)}%</b><br>效能分：<b>${esc(r.effectiveness_score)}</b> ${scoreNote}<br>结果可信程度：<b>${esc(confidenceWord(r.effectiveness_confidence))}</b></div>${problems?`<div class="section-title">建议重点检查</div><ul class="list">${problems}</ul>`:'<div class="meta">暂未发现明显问题。</div>'}<details class="scheme-details"><summary>查看计算依据</summary><div class="notice"><div>可行性说明：${esc(r.feasibility_confidence)}</div><div>效能说明：${esc(r.effectiveness_confidence)}</div><div>P可行 ${esc(r.learned_feasibility_probability)}，综合风险 ${esc(r.risk)}，物理状态 ${esc(label('status',r.physics_status))}。</div><div>通用BT分 ${esc(r.bt_score)}，通用UTA分 ${esc(r.uta_score??'暂无')}，UTA区间 ${interval}；协议评分与训练模型相互独立。</div>${technicalRisks?`<ul class="list">${technicalRisks}</ul>`:''}${coupling?`<ul class="list">${coupling}</ul>`:''}${contributions?`<ul class="list">${contributions}</ul>`:''}${basis?`<div><b>使用依据</b></div><ul class="list">${basis}</ul>`:''}</div></details></div>`;}
function renderProject(){const p=appState.project,s=appState.stats,uta=appState.utaModel||{},fm=appState.feasibilityModel||{},bt=appState.btModel||{},cv=uta.cross_validation||{},holdout=uta.whole_scheme_holdout||{},m3=uta.m3||{};document.getElementById('projectMeta').textContent=`${p.name} · ${s.attributes} 项参数`;const stateText=plainModelState(uta,s.openReviews);document.getElementById('projectInfo').innerHTML=`<div><b>${esc(p.name)}</b></div><div class="meta">本轮比较 ${esc(appState.pair.a.id)} 和 ${esc(appState.pair.b.id)}</div><div class="notice"><b>${esc(stateText)}</b><br>${s.interactions?`已完成 ${esc(s.interactions)} 轮判断。`:'请从左侧方案 A 和方案 B 开始第一轮判断。'}</div>${s.openReviews?`<div class="notice warn">有 ${esc(s.openReviews)} 个判断需要重新确认。</div>`:''}`;const warnings=(p.warnings||[]).map(x=>`<div class="notice warn">${esc(x)}</div>`).join('');const stages=[1,2,3].map(stage=>`${stage}批 ${(entries().filter(([,spec])=>Number(spec.design_stage)===stage)).length}项`).join(' · ');const settings=appState.preferenceSettings||{},options=Array.from({length:Number(settings.maxSegments)-Number(settings.minSegments)+1},(_,i)=>Number(settings.minSegments)+i).map(n=>`<option value="${n}" ${n===Number(settings.utaSegments)?'selected':''}>${n} 段</option>`).join('');const models=(appState.couplingModels||[]).map(model=>model.model_status==='direction_only'?`<div class="coupling"><b>${esc(model.target_label)}方向耦合</b><br>${esc(model.source_effects.length)} 个源属性；当前只有影响方向，已有 ${esc(model.sample_count)} 条样本，正在通过专家判断学习强度。</div>`:`<div class="coupling"><b>${esc(model.target_label)}条件代理</b><br>${esc(model.source_effects.length)} 个源属性 · R² ${esc(model.r2)} · RMSE ${esc(model.rmse)} ${esc(model.target_unit)} · ${esc(model.confidence)}置信</div>`).join('');document.getElementById('modelDetails').innerHTML=`<div class="meta">Excel：${esc(p.workbook_path)}<br>指纹：${esc(p.fingerprint)}<br>设计顺序：${esc(stages)}</div><div class="setting-row"><label for="utaSegments"><b>UTA 统一分段数</b></label><select id="utaSegments">${options}</select></div><div class="coupling"><b>可行性模型</b><br>专家证据 ${esc(fm.expert_samples||0)}，单调前沿证据 ${esc(fm.frontier_evidence||0)}，正/负例 ${esc(fm.positive_samples||0)}/${esc(fm.negative_samples||0)}。</div><div class="coupling"><b>Bradley-Terry</b><br>有效偏好 ${esc(bt.training_pairs||0)} 对，训练准确率 ${esc(bt.training_accuracy??'暂无')}，log loss ${esc(bt.log_loss??'暂无')}。</div><div class="coupling"><b>LP-UTA</b><br>${esc(uta.attribute_count||0)} 属性 × ${esc(uta.segments||settings.utaSegments)} 段；M1 ${esc(uta.m1_label??'待检验')}；M2 总容忍 ${esc(uta.m2_total_slack??0)}；M3 最少冲突 ${esc(m3.minimum_conflicts??0)}。<br>训练/测试准确率 ${esc(uta.training_accuracy??'暂无')}/${esc(uta.test_accuracy??'暂无')}；交叉验证 ${esc(cv.accuracy??'暂无')}；整方案留出 ${esc(holdout.accuracy??'暂无')}。</div>${models||'<div class="empty">没有耦合代理</div>'}${warnings}`;document.getElementById('utaSegments').onchange=async e=>{e.target.disabled=true;try{render(await api('/api/preference-settings',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({utaSegments:Number(e.target.value)})}));}catch(err){document.getElementById('projectMeta').innerHTML=`<span class="error">${esc(err.message)}</span>`;e.target.disabled=false;}};}
const renderProjectCore=renderProject;
renderProject=function(){renderProjectCore();const profiles=appState.protocolProfiles||[],active=appState.activeProtocolProfileId,projectInfo=document.getElementById('projectInfo'),modelDetails=document.getElementById('modelDetails');projectInfo.insertAdjacentHTML('beforeend',`<div class="meta">A/B判断只学习通用产品效能，不使用新技术协议。协议仅在“未知方案评估”时提供100分参考。</div>`);if(!profiles.length){modelDetails.insertAdjacentHTML('afterbegin','<div class="notice warn"><b>未设置新技术协议。</b> 当前只能查看通用效能参考分。</div>');return;}const options=profiles.map(x=>`<option value="${esc(x.id)}" ${x.id===active?'selected':''}>${esc(x.name)}（${esc(x.id)}）</option>`).join('');modelDetails.insertAdjacentHTML('afterbegin',`<div class="setting-row"><label for="activeProtocol"><b>未知方案评分参考</b><br><span class="meta">协议本身=100分；只切换评分，不重新训练</span></label><select id="activeProtocol">${options}</select></div>`);document.getElementById('activeProtocol').onchange=async e=>{e.target.disabled=true;try{render(await api('/api/active-protocol',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({protocolId:e.target.value})}));}catch(err){document.getElementById('projectMeta').innerHTML=`<span class="error">${esc(err.message)}</span>`;e.target.disabled=false;}};};
function renderReview(x){const typeLabel={feasibility_conflict:'同一方案的可行性判断前后不一致',preference_pair_conflict:'同一对方案出现了相反选择',preference_cycle:'多次方案选择互相矛盾',uta_high_tolerance:'这次选择与整体判断趋势差别较大',uta_m3_conflict:'系统定位到一条可能有误的判断'}[x.type]||'这次判断需要重新确认';const actions=x.status==='open'?(x.allowed_actions||[]).map(action=>`<button type="button" class="${action==='withdraw'?'warn':'secondary'}" onclick="resolveReview('${esc(x.id)}','${esc(action)}')">${esc(reviewActionLabel[action]||action)}</button>`).join(''):'';const status=x.status==='open'?'等待确认':'已经处理';return `<div class="event"><b>${esc(typeLabel)}</b><br><span class="meta">${esc(status)}。请根据工程经验重新确认，不需要考虑系统算法。</span>${actions?`<div class="review-actions">${actions}</div>`:''}<details class="scheme-details"><summary>查看记录编号</summary>${esc(x.id)}；关联判断 ${esc((x.evidence_ids||[]).join('、'))}；默认处理 ${esc(x.target_evidence_id||'无')}。<br>${esc(x.message)}</details></div>`;}
renderReview=function(x){const typeLabel={feasibility_conflict:'同一方案的可行性判断前后不一致',preference_pair_conflict:'同一对方案出现相反选择',preference_cycle:'多次方案选择形成矛盾环',uta_high_tolerance:'这次选择与整体趋势差别较大',uta_m3_conflict:'系统定位到一条可能有误的判断'}[x.type]||'这次判断需要重新确认',records=(x.evidence_records||[]).map(record=>{if(record.kind==='preference'){const rows=entries().filter(([key])=>record.params_a?.[key]!==record.params_b?.[key]).map(([key,spec])=>`<tr><td>${esc(spec.label)}</td><td>${formatValue(record.params_a?.[key],spec)}</td><td>${formatValue(record.params_b?.[key],spec)}</td></tr>`).join('');return `<div class="notice"><b>${esc(record.id)}：${esc(record.scheme_a)} / ${esc(record.scheme_b)}</b><br>原选择：${esc(label('preference',record.relation))}；证据状态：${esc(record.status)}${rows?`<table><thead><tr><th>变化参数</th><th>A</th><th>B</th></tr></thead><tbody>${rows}</tbody></table>`:''}</div>`;}return `<div class="notice"><b>${esc(record.id)}：方案 ${esc(record.scheme_id)}</b><br>原判断：${esc(label('feasibility',record.label))}；原因：${esc(record.reason_text||'未填写')}</div>`;}).join(''),actions=(x.allowed_actions||[]).map(action=>`<button type="button" class="${action==='withdraw'?'warn':'secondary'}" onclick="resolveReview('${esc(x.id)}','${esc(action)}')">${esc(reviewActionLabel[action]||action)}</button>`).join(''),status=x.status==='open'?'等待专家处理':'已处理',dependency=x.dependency_note?`<div class="notice warn">${esc(x.dependency_note)}</div>`:'',related=(x.dependent_review_ids||[]).length?`<div class="meta">这次判断还关联复核项：${esc(x.dependent_review_ids.join('、'))}。处理本项不会删除其他项。</div>`:'';return `<div class="event"><b>${esc(typeLabel)}</b><br><span class="meta">${status}。每个复核项独立处理，原始记录始终保留。</span>${dependency}${records}<div class="review-actions">${actions}</div>${related}<details class="scheme-details"><summary>查看检测依据</summary>${esc(x.message)}<br>复核编号 ${esc(x.id)}；关联证据 ${esc((x.evidence_ids||[]).join('、'))}；本项默认作用于 ${esc(x.target_evidence_id||'无')}。</details></div>`;};
async function resolveReview(reviewId,action){if(action!=='keep'&&!confirm(`确定要“${reviewActionLabel[action]}”吗？原判断会保留在修订历史中。`))return;try{render(await api('/api/review-action',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({reviewId,action})}));}catch(err){alert(err.message);}}
function renderModeBanner(){const target=document.getElementById('modeBanner');if(appState.dataMode==='fixed_demo_simulation'){target.innerHTML='<div class="notice warn"><b>当前是演示模式。</b> 页面中的判断由软件模拟生成，只用于查看功能。正式使用前请点击“清空判断”。</div>';}else{target.innerHTML='<div class="notice"><b>当前可以直接开始判断。</b> 请比较左右两个方案，并在下方提交您的意见。</div>';}}
function render(state){appState=state;renderModeBanner();renderProject();const s=state.stats;document.getElementById('stats').innerHTML=[['已有方案',s.existingSchemes],['已完成判断',s.interactions],['有效比较',s.activePreferenceEvidence],['需要复核',s.openReviews]].map(x=>`<div class="stat"><div class="num">${esc(x[1])}</div><div class="cap">${esc(x[0])}</div></div>`).join('');document.getElementById('schemeA').innerHTML=renderScheme(state.pair.a);document.getElementById('schemeB').innerHTML=renderScheme(state.pair.b);choice('feasibilityA','feasibilityA',[['feasible','可以实现'],['infeasible','不能实现'],['uncertain','暂时看不准']],'');choice('feasibilityB','feasibilityB',[['feasible','可以实现'],['infeasible','不能实现'],['uncertain','暂时看不准']],'');choice('preference','preference',[['A','方案 A 更好'],['B','方案 B 更好'],['tie','两个方案差不多'],['unknown','暂时无法比较']],'');reasons('reasonsA','reasonCodesA');reasons('reasonsB','reasonCodesB');renderQuickCoupling('A',state.pair.a);renderQuickCoupling('B',state.pair.b);bindReasonAreas();renderEvalInputs();renderRangeEditor();document.getElementById('eventCount').textContent=`(${state.learningEvents.length})`;document.getElementById('events').innerHTML=state.learningEvents.map(e=>`<div class="event"><b>${esc(e.title)}</b><br>${esc(e.detail)}<br><span class="meta">${esc(e.created_at)}</span></div>`).join('')||'<div class="empty">暂无后台记录</div>';document.getElementById('reviewCount').textContent=`(${s.openReviews})`;document.getElementById('reviews').innerHTML=(state.reviewEvidence||[]).map(renderReview).join('')||'<div class="empty">目前没有需要重新确认的判断。</div>';document.getElementById('recent').innerHTML=state.recentInteractions.map(x=>`<div class="recent"><b>${esc(x.scheme_a)} / ${esc(x.scheme_b)}</b><br>A：${esc(label('feasibility',x.feasibility_a))}，B：${esc(label('feasibility',x.feasibility_b))}；${esc(label('preference',x.preference))}</div>`).join('')||'<div class="empty">还没有提交判断。</div>';renderVisuals();requestAnimationFrame(alignSchemeCards);}
async function api(path,options){const res=await fetch(path,options);const data=await res.json();if(!res.ok)throw new Error(data.error||'请求失败');return data;}
async function load(){try{render(await api('/api/state'));}catch(e){document.getElementById('projectMeta').innerHTML=`<span class="error">${esc(e.message)}</span>`;}}
document.getElementById('feedbackForm').addEventListener('submit',async e=>{e.preventDefault();const message=document.getElementById('msg'),a=document.querySelector('input[name="feasibilityA"]:checked'),b=document.querySelector('input[name="feasibilityB"]:checked');if(!a||!b){message.textContent='请先判断方案 A 和方案 B 能否实现。';return;}let preference='unknown';if(a.value==='feasible'&&b.value==='feasible'){const selected=document.querySelector('input[name="preference"]:checked');if(!selected){message.textContent='两个方案都可以实现，请再选择您更推荐哪个。';return;}preference=selected.value;}const payload={feasibilityA:a.value,feasibilityB:b.value,preference,reasonCodesA:checked('reasonCodesA'),reasonCodesB:checked('reasonCodesB'),reasonA:document.getElementById('reasonA').value,reasonB:document.getElementById('reasonB').value,couplingFeedbackA:readCouplingFeedback('A'),couplingFeedbackB:readCouplingFeedback('B')};try{render(await api('/api/submit',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(payload)}));document.getElementById('msg').textContent='已保存，下一组已经准备好。';}catch(err){message.textContent=err.message;}});
document.getElementById('evalBtn').onclick=async()=>{const payload={};entries().forEach(([key])=>payload[key]=document.getElementById(`eval_${key}`).value);try{document.getElementById('evalResult').innerHTML=renderEvaluationReport(await api('/api/evaluate',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(payload)}));}catch(e){document.getElementById('evalResult').innerHTML=`<div class="eval-result error">${esc(e.message)}</div>`;}};
document.querySelectorAll('[data-viz-view]').forEach(button=>button.onclick=()=>{visualView=button.dataset.vizView;renderVisuals();});
function downloadBlob(blob,filename){const url=URL.createObjectURL(blob),link=document.createElement('a');link.href=url;link.download=filename;link.click();URL.revokeObjectURL(url);}
document.getElementById('refreshBtn').onclick=load;
document.getElementById('demoBtn').onclick=async e=>{if(!confirm('演示数据由软件模拟生成，会替换当前判断。确定载入吗？'))return;const button=e.currentTarget,old=button.textContent;button.disabled=true;button.textContent='正在准备...';try{render(await api('/api/load-demo',{method:'POST'}));}catch(err){alert(err.message);}finally{button.disabled=false;button.textContent=old;}};
document.getElementById('resetBtn').onclick=async()=>{if(!confirm('确定清空当前所有判断并重新开始吗？'))return;render(await api('/api/reset',{method:'POST'}));};
document.getElementById('exportBtn').onclick=async()=>{const data=await api('/api/export');downloadBlob(new Blob([JSON.stringify(data,null,2)],{type:'application/json;charset=utf-8'}),`project_export_${appState.project.fingerprint}.json`);};
document.getElementById('bundleBtn').onclick=async e=>{const button=e.currentTarget,old=button.textContent;button.disabled=true;button.textContent='正在生成报告...';try{const response=await fetch('/api/export-bundle');if(!response.ok)throw new Error('报告生成失败');const disposition=response.headers.get('Content-Disposition')||'',match=disposition.match(/filename="([^"]+)"/);downloadBlob(await response.blob(),match?.[1]||`effectiveness_report_${appState.project.fingerprint}.zip`);}catch(err){alert(err.message);}finally{button.disabled=false;button.textContent=old;}};
window.addEventListener('resize',()=>requestAnimationFrame(alignSchemeCards));load();
</script></body></html>"""


class ProjectRequestHandler(BaseHTTPRequestHandler):
    server: Any

    def log_message(self, format: str, *args: Any) -> None:
        return

    @property
    def app(self) -> ProjectApp:
        return self.server.app

    def do_GET(self) -> None:
        path = urlparse(self.path).path
        try:
            if path == "/":
                html_response(self, HTML)
            elif path == "/api/state":
                json_response(self, self.app.summarize())
            elif path == "/api/export":
                json_response(self, self.app.export())
            elif path == "/api/export-bundle":
                data, filename = self.app.export_bundle()
                binary_response(self, data, "application/zip", filename)
            elif path == "/api/health":
                json_response(self, {"status": "ok", "project": self.app.project.project_name})
            else:
                json_response(self, {"error": "未找到该接口。"}, 404)
        except (ProjectDataError, ValueError) as exc:
            json_response(self, {"error": str(exc)}, 400)
        except Exception as exc:
            json_response(self, {"error": f"服务器错误：{exc}"}, 500)

    def do_POST(self) -> None:
        path = urlparse(self.path).path
        try:
            payload = read_body(self)
            if path == "/api/submit":
                json_response(self, self.app.submit(payload))
            elif path == "/api/evaluate":
                json_response(self, self.app.evaluate_unknown(payload))
            elif path == "/api/preference-settings":
                json_response(self, self.app.update_preference_settings(payload))
            elif path == "/api/active-protocol":
                json_response(self, self.app.update_active_protocol(payload))
            elif path == "/api/generation-ranges":
                json_response(self, self.app.update_generation_ranges(payload))
            elif path == "/api/review-action":
                json_response(self, self.app.review_action(payload))
            elif path == "/api/feasibility-slice":
                json_response(self, self.app.feasibility_slice(payload))
            elif path == "/api/load-demo":
                json_response(self, self.app.prepare_demo_state())
            elif path == "/api/reset":
                json_response(self, self.app.summarize(self.app.load_state(reset=True)))
            else:
                json_response(self, {"error": "未找到该接口。"}, 404)
        except (ProjectDataError, ValueError, json.JSONDecodeError) as exc:
            json_response(self, {"error": str(exc)}, 400)
        except Exception as exc:
            json_response(self, {"error": f"服务器错误：{exc}"}, 500)


def run_server(app: ProjectApp, host: str, port: int) -> None:
    server = ThreadingHTTPServer((host, port), ProjectRequestHandler)
    server.app = app
    print(f"项目：{app.project.project_name}")
    print(f"Excel：{app.project.workbook_path}")
    print(f"状态：{app.state_path}")
    print(f"打开：http://{host}:{port}")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        server.server_close()


def choose_workbook(state_dir: Path) -> Path:
    """Open a native Windows file picker; CLI users can still pass --workbook."""
    recent_file = state_dir.expanduser().resolve() / "recent_workbook.json"
    initial_dir = str(Path.home())
    try:
        recent = json.loads(recent_file.read_text(encoding="utf-8"))
        recent_path = Path(str(recent.get("workbook") or ""))
        if recent_path.parent.exists():
            initial_dir = str(recent_path.parent)
    except (OSError, ValueError, TypeError, json.JSONDecodeError):
        pass
    try:
        import tkinter as tk
        from tkinter import filedialog
    except ImportError as exc:
        raise ProjectDataError(
            "当前 Python 缺少 Tkinter，无法打开 Excel 选择窗口。"
            "请安装 Tkinter，或用 --workbook 指定文件。"
        ) from exc
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    try:
        selected = filedialog.askopenfilename(
            parent=root,
            title="请选择方案数据 Excel",
            initialdir=initial_dir,
            filetypes=[("Excel 工作簿", "*.xlsx *.xlsm"), ("所有文件", "*.*")],
        )
    finally:
        root.destroy()
    if not selected:
        raise ProjectDataError("已取消选择 Excel，程序未启动。")
    workbook = Path(selected).expanduser().resolve()
    recent_file.parent.mkdir(parents=True, exist_ok=True)
    recent_file.write_text(
        json.dumps({"workbook": str(workbook), "selected_at": now_iso()}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return workbook


def main() -> None:
    parser = argparse.ArgumentParser(description="Workbook-driven A/B parameter comparison application.")
    parser.add_argument("--workbook", type=Path, default=None)
    parser.add_argument("--demo", action="store_true", help="直接打开内置演示 Excel，不弹出文件选择框。")
    parser.add_argument("--state-dir", type=Path, default=DEFAULT_STATE_DIR)
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", type=int, default=DEFAULT_PORT)
    parser.add_argument("--reset", action="store_true")
    args = parser.parse_args()
    try:
        workbook = DEFAULT_WORKBOOK if args.demo else args.workbook or choose_workbook(args.state_dir)
        app = ProjectApp(workbook, state_dir=args.state_dir)
        if args.reset:
            app.load_state(reset=True)
    except ProjectDataError as exc:
        raise SystemExit(f"项目载入失败：{exc}") from exc
    run_server(app, args.host, args.port)


if __name__ == "__main__":
    main()
