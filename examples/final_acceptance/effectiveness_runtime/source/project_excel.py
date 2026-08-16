# -*- coding: utf-8 -*-
"""Generic Excel project loader for parameter comparison projects."""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import os
import re
import shutil
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook


DATA_SHEET = "方案数据"
ATTRIBUTE_SHEET = "属性配置"
COUPLING_SHEET = "耦合关系"
PROTOCOL_SHEET = "新技术协议"
PROJECT_SHEET = "项目信息"
LEGACY_REQUIREMENT_SHEET = "评价基准"

RESERVED_HEADERS = {
    "方案编号": "id",
    "方案id": "id",
    "id": "id",
    "来源": "source",
    "source": "source",
    "已知可行性": "known_feasibility",
    "可行性标签": "known_feasibility",
    "备注": "notes",
    "notes": "notes",
}

ATTRIBUTE_HEADER_ALIASES = {
    "属性名": "label",
    "属性名称": "label",
    "属性id": "key",
    "内部id": "key",
    "单位": "unit",
    "数据类型": "data_type",
    "设计顺序": "design_stage",
    "设计阶段": "design_stage",
    "显示精度": "precision",
    "生成下限": "generation_min",
    "生成上限": "generation_max",
    "可行下限": "feasible_min",
    "可行上限": "feasible_max",
    "偏好方向": "preference_direction",
    "边际规律": "marginal_trend",
    "参与效能": "participates_utility",
    "参与生成": "participates_generation",
    "说明": "description",
}

COUPLING_HEADER_ALIASES = {
    "源属性": "source",
    "目标属性": "target",
    "方向": "direction",
    "关系类型": "relation_type",
    "先验系数": "coefficient_prior",
    "置信状态": "status",
    "说明": "description",
}

REQUIREMENT_HEADER_ALIASES = {
    "基准编号": "profile_id",
    "评价基准编号": "profile_id",
    "基准名称": "profile_name",
    "评价基准名称": "profile_name",
    "属性名": "label",
    "属性名称": "label",
    "要求类型": "requirement_type",
    "评价方式": "requirement_type",
    "要求值": "target_value",
    "目标值": "target_value",
    "要求下限": "minimum",
    "要求上限": "maximum",
    "容差": "tolerance",
    "硬性要求": "hard_requirement",
    "超额加分": "overachievement_bonus",
    "直接复用阈值": "direct_reuse_threshold",
    "改进复用阈值": "improvement_threshold",
    "重新研制阈值": "redesign_threshold",
    "说明": "description",
}

PROTOCOL_METADATA_HEADERS = {
    "协议编号": "profile_id",
    "技术协议编号": "profile_id",
    "协议名称": "profile_name",
    "技术协议名称": "profile_name",
    "说明": "description",
}

DIRECTION_ALIASES = {
    "正向": "positive",
    "正": "positive",
    "positive": "positive",
    "+": "positive",
    "负向": "negative",
    "负": "negative",
    "negative": "negative",
    "-": "negative",
}

PREFERENCE_ALIASES = {
    "越大越好": "higher_better",
    "递增": "higher_better",
    "higher_better": "higher_better",
    "越小越好": "lower_better",
    "递减": "lower_better",
    "lower_better": "lower_better",
    "区间型": "interval",
    "区间偏好": "interval",
    "interval": "interval",
    "不参与": "neutral",
    "无偏好": "neutral",
    "neutral": "neutral",
    "待确认": "unspecified",
    "": "unspecified",
}

REQUIREMENT_TYPE_ALIASES = {
    "达到下限即可": "at_least",
    "不低于下限": "at_least",
    "不低于": "at_least",
    "at_least": "at_least",
    "不超过上限即可": "at_most",
    "不高于上限": "at_most",
    "不超过": "at_most",
    "at_most": "at_most",
    "落入区间即可": "within_range",
    "区间内即可": "within_range",
    "区间要求": "within_range",
    "within_range": "within_range",
    "接近目标值最好": "target",
    "接近目标": "target",
    "目标值": "target",
    "target": "target",
    "越大越好": "higher_better",
    "higher_better": "higher_better",
    "越小越好": "lower_better",
    "lower_better": "lower_better",
}

DATA_TYPE_ALIASES = {
    "连续": "continuous",
    "连续型": "continuous",
    "float": "continuous",
    "continuous": "continuous",
    "整数": "integer",
    "整数型": "integer",
    "int": "integer",
    "integer": "integer",
    "类别": "categorical",
    "分类型": "categorical",
    "category": "categorical",
    "categorical": "categorical",
}


class ProjectDataError(ValueError):
    """Raised when a workbook cannot be interpreted safely."""


@dataclass
class AttributeSpec:
    key: str
    label: str
    unit: str = "-"
    data_type: str = "continuous"
    design_stage: int = 2
    precision: int = 3
    generation_min: Optional[float] = None
    generation_max: Optional[float] = None
    feasible_min: Optional[float] = None
    feasible_max: Optional[float] = None
    preference_direction: str = "unspecified"
    marginal_trend: str = "待确认"
    participates_utility: bool = True
    participates_generation: bool = True
    description: str = ""
    inferred_fields: List[str] = field(default_factory=list)

    @property
    def is_numeric(self) -> bool:
        return self.data_type in {"continuous", "integer"}


@dataclass
class CouplingSpec:
    source_key: str
    source_label: str
    target_key: str
    target_label: str
    direction: str
    relation_type: str = "monotonic"
    coefficient_prior: Optional[float] = None
    status: str = "待确认"
    description: str = ""


@dataclass
class RequirementSpec:
    attribute_key: str
    attribute_label: str
    requirement_type: str
    target_value: Optional[float] = None
    minimum: Optional[float] = None
    maximum: Optional[float] = None
    tolerance: float = 0.0
    hard_requirement: bool = False
    overachievement_bonus: bool = False
    description: str = ""


@dataclass
class RequirementProfile:
    id: str
    name: str
    requirements: List[RequirementSpec]
    direct_reuse_threshold: float = 0.95
    improvement_threshold: float = 0.80
    redesign_threshold: float = 0.60

    def requirement_by_key(self) -> Dict[str, RequirementSpec]:
        return {item.attribute_key: item for item in self.requirements}


@dataclass
class SchemeRecord:
    id: str
    source: str
    known_feasibility: str
    notes: str
    params: Dict[str, Any]
    excel_row: int


@dataclass
class ProjectDataset:
    product_code: str
    project_name: str
    workbook_path: str
    workbook_fingerprint: str
    learning_fingerprint: str
    attributes: List[AttributeSpec]
    schemes: List[SchemeRecord]
    couplings: List[CouplingSpec]
    requirement_profiles: List[RequirementProfile] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)

    def attribute_by_key(self) -> Dict[str, AttributeSpec]:
        return {item.key: item for item in self.attributes}

    def requirement_profile_by_id(self) -> Dict[str, RequirementProfile]:
        return {item.id: item for item in self.requirement_profiles}

    def default_requirement_profile(self) -> Optional[RequirementProfile]:
        return self.requirement_profiles[0] if self.requirement_profiles else None

    def to_dict(self) -> Dict[str, Any]:
        return asdict(self)

    def summary(self) -> Dict[str, Any]:
        return {
            "product_code": self.product_code,
            "project_name": self.project_name,
            "workbook_path": self.workbook_path,
            "workbook_fingerprint": self.workbook_fingerprint,
            "learning_fingerprint": self.learning_fingerprint,
            "attribute_count": len(self.attributes),
            "scheme_count": len(self.schemes),
            "coupling_count": len(self.couplings),
            "requirement_profile_count": len(self.requirement_profiles),
            "requirement_profiles": [
                {"id": item.id, "name": item.name, "requirement_count": len(item.requirements)}
                for item in self.requirement_profiles
            ],
            "attribute_labels": [item.label for item in self.attributes],
            "coupling_edges": [
                {
                    "source": item.source_label,
                    "target": item.target_label,
                    "direction": item.direction,
                    "status": item.status,
                }
                for item in self.couplings
            ],
            "warnings": list(self.warnings),
        }


def normalize_header(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "").strip()).lower()


def display_text(value: Any, default: str = "") -> str:
    text = str(value or "").strip()
    return text if text else default


def optional_float(value: Any, field_name: str) -> Optional[float]:
    if value is None or str(value).strip() == "":
        return None
    try:
        result = float(value)
    except (TypeError, ValueError) as exc:
        raise ProjectDataError(f"{field_name} 必须是数值，当前值为 {value!r}。") from exc
    if not math.isfinite(result):
        raise ProjectDataError(f"{field_name} 必须是有限数值。")
    return result


def parse_fraction(value: Any, field_name: str, default: float) -> float:
    parsed = optional_float(value, field_name)
    if parsed is None:
        return default
    if parsed > 1.0:
        parsed /= 100.0
    if not 0.0 <= parsed <= 1.0:
        raise ProjectDataError(f"{field_name} 必须位于 0 到 1（或 0 到 100%）之间。")
    return parsed


def parse_bool(value: Any, default: bool = True) -> bool:
    text = normalize_header(value)
    if text in {"是", "1", "true", "yes", "y", "参与"}:
        return True
    if text in {"否", "0", "false", "no", "n", "不参与"}:
        return False
    return default


def parse_precision(value: Any) -> int:
    if value is None or str(value).strip() == "":
        return 3
    try:
        precision = int(value)
    except (TypeError, ValueError) as exc:
        raise ProjectDataError(f"显示精度必须是 0-8 的整数，当前值为 {value!r}。") from exc
    if not 0 <= precision <= 8:
        raise ProjectDataError("显示精度必须位于 0-8。")
    return precision


def parse_design_stage(value: Any) -> Optional[int]:
    if value is None or str(value).strip() == "":
        return None
    text = normalize_header(value)
    aliases = {"第一批": 1, "第一阶段": 1, "1": 1, "第二批": 2, "第二阶段": 2, "2": 2, "第三批": 3, "第三阶段": 3, "3": 3}
    stage = aliases.get(text)
    if stage is None:
        raise ProjectDataError(f"设计顺序必须是 1、2、3 或第一批/第二批/第三批，当前值为 {value!r}。")
    return stage


def stable_default_key(index: int) -> str:
    return f"attr_{index:02d}"


def validate_key(value: str, label: str) -> str:
    key = value.strip()
    if not re.fullmatch(r"[A-Za-z][A-Za-z0-9_]*", key):
        raise ProjectDataError(f"属性“{label}”的属性ID {key!r} 无效，只能使用英文字母、数字和下划线。")
    return key


def row_dicts(sheet: Any, aliases: Dict[str, str]) -> Iterable[Tuple[int, Dict[str, Any]]]:
    rows = sheet.iter_rows(values_only=True)
    try:
        raw_headers = next(rows)
    except StopIteration:
        return
    mapped: List[Optional[str]] = []
    for value in raw_headers:
        mapped.append(aliases.get(normalize_header(value)))
    for excel_row, values in enumerate(rows, start=2):
        if not any(value is not None and str(value).strip() for value in values):
            continue
        record: Dict[str, Any] = {}
        for key, value in zip(mapped, values):
            if key:
                record[key] = value
        yield excel_row, record


def read_project_identity(workbook: Any, fallback_name: str) -> Tuple[str, str, List[str]]:
    if PROJECT_SHEET not in workbook.sheetnames:
        return fallback_name, fallback_name, [
            f"缺少“{PROJECT_SHEET}”工作表，成品代号和名称暂使用文件名。"
        ]
    values: Dict[str, str] = {}
    for row in workbook[PROJECT_SHEET].iter_rows(min_row=1, max_col=2, values_only=True):
        key = normalize_header(row[0] if row else None)
        value = display_text(row[1] if len(row) > 1 else None)
        if key in {"成品代号", "项目代号", "product_code", "project_code"}:
            values["product_code"] = value
        elif key in {"成品名称", "项目名称", "product_name", "project_name"}:
            values["project_name"] = value
    product_code = values.get("product_code") or fallback_name
    project_name = values.get("project_name") or fallback_name
    if not re.fullmatch(r"[A-Za-z][A-Za-z0-9_.-]*", product_code):
        raise ProjectDataError(
            f"“{PROJECT_SHEET}”中的成品代号 {product_code!r} 无效，"
            "只能使用英文字母开头的字母、数字、下划线、点或短横线。"
        )
    warnings: List[str] = []
    if "product_code" not in values:
        warnings.append(f"“{PROJECT_SHEET}”未填写成品代号，暂使用文件名。")
    if "project_name" not in values:
        warnings.append(f"“{PROJECT_SHEET}”未填写成品名称，暂使用文件名。")
    return product_code, project_name, warnings


def infer_numeric_range(values: Sequence[float]) -> Tuple[float, float]:
    lo = min(values)
    hi = max(values)
    span = hi - lo
    pad = 0.05 * span if span > 1e-12 else max(abs(lo) * 0.05, 1.0)
    return lo - pad, hi + pad


def workbook_fingerprint(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()[:16]


def learning_fingerprint(
    attributes: Sequence[AttributeSpec],
    schemes: Sequence[SchemeRecord],
    couplings: Sequence[CouplingSpec],
) -> str:
    """Hash only reusable learning inputs; protocol rows are deliberately excluded."""
    payload = {
        "attributes": [asdict(item) for item in attributes],
        "schemes": [
            {
                "id": item.id,
                "source": item.source,
                "known_feasibility": item.known_feasibility,
                "notes": item.notes,
                "params": item.params,
            }
            for item in schemes
        ],
        "couplings": [asdict(item) for item in couplings],
    }
    encoded = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(encoded.encode("utf-8")).hexdigest()[:16]


def read_attribute_config(workbook: Any) -> Tuple[Dict[str, Dict[str, Any]], List[str]]:
    warnings: List[str] = []
    if ATTRIBUTE_SHEET not in workbook.sheetnames:
        warnings.append(f"缺少“{ATTRIBUTE_SHEET}”工作表，属性范围和偏好方向将从数据推断或标为待确认。")
        return {}, warnings
    configs: Dict[str, Dict[str, Any]] = {}
    for excel_row, record in row_dicts(workbook[ATTRIBUTE_SHEET], ATTRIBUTE_HEADER_ALIASES):
        label = display_text(record.get("label"))
        if not label:
            raise ProjectDataError(f"“{ATTRIBUTE_SHEET}”第 {excel_row} 行缺少属性名。")
        normalized = normalize_header(label)
        if normalized in configs:
            raise ProjectDataError(f"“{ATTRIBUTE_SHEET}”中属性“{label}”重复。")
        configs[normalized] = record
    return configs, warnings


def read_data_headers(sheet: Any) -> Tuple[List[str], List[Tuple[int, str]], Dict[int, str]]:
    rows = sheet.iter_rows(min_row=1, max_row=1, values_only=True)
    raw = list(next(rows, ()))
    while raw and (raw[-1] is None or not str(raw[-1]).strip()):
        raw.pop()
    if not raw:
        raise ProjectDataError(f"“{DATA_SHEET}”第一行没有表头。")
    seen: Dict[str, int] = {}
    attributes: List[Tuple[int, str]] = []
    metadata: Dict[int, str] = {}
    display_headers: List[str] = []
    for index, value in enumerate(raw):
        label = display_text(value)
        if not label:
            raise ProjectDataError(f"“{DATA_SHEET}”第一行第 {index + 1} 列表头为空。")
        normalized = normalize_header(label)
        if normalized in seen:
            raise ProjectDataError(f"“{DATA_SHEET}”第一行存在重复表头“{label}”。")
        seen[normalized] = index
        display_headers.append(label)
        reserved = RESERVED_HEADERS.get(normalized)
        if reserved:
            metadata[index] = reserved
        else:
            attributes.append((index, label))
    if not attributes:
        raise ProjectDataError(f"“{DATA_SHEET}”没有识别到属性列。")
    return display_headers, attributes, metadata


def read_raw_data(
    sheet: Any,
    attribute_columns: List[Tuple[int, str]],
    metadata_columns: Dict[int, str],
) -> Tuple[List[Dict[str, Any]], Dict[str, List[Any]]]:
    rows: List[Dict[str, Any]] = []
    values_by_label = {label: [] for _, label in attribute_columns}
    max_column = max([index for index, _ in attribute_columns] + list(metadata_columns.keys())) + 1
    for excel_row, values in enumerate(sheet.iter_rows(min_row=2, max_col=max_column, values_only=True), start=2):
        if not any(value is not None and str(value).strip() for value in values):
            continue
        metadata: Dict[str, Any] = {}
        for index, key in metadata_columns.items():
            metadata[key] = values[index] if index < len(values) else None
        raw_params: Dict[str, Any] = {}
        for index, label in attribute_columns:
            value = values[index] if index < len(values) else None
            if value is None or str(value).strip() == "":
                raise ProjectDataError(f"“{DATA_SHEET}”第 {excel_row} 行属性“{label}”为空。")
            raw_params[label] = value
            values_by_label[label].append(value)
        rows.append({"excel_row": excel_row, "metadata": metadata, "params": raw_params})
    if not rows:
        raise ProjectDataError(f"“{DATA_SHEET}”没有数据行。")
    return rows, values_by_label


def build_attribute_specs(
    attribute_columns: List[Tuple[int, str]],
    configs: Dict[str, Dict[str, Any]],
    values_by_label: Dict[str, List[Any]],
) -> Tuple[List[AttributeSpec], List[str]]:
    specs: List[AttributeSpec] = []
    warnings: List[str] = []
    used_keys: Dict[str, str] = {}
    present_labels = {normalize_header(label) for _, label in attribute_columns}
    for config_label in configs:
        if config_label not in present_labels:
            warnings.append(f"属性配置中的“{display_text(configs[config_label].get('label'))}”未出现在方案数据第一行，已忽略。")

    for order, (_, label) in enumerate(attribute_columns, start=1):
        config = configs.get(normalize_header(label), {})
        key_value = display_text(config.get("key"), stable_default_key(order))
        key = validate_key(key_value, label)
        if key in used_keys:
            raise ProjectDataError(f"属性“{label}”与“{used_keys[key]}”使用了相同属性ID“{key}”。")
        used_keys[key] = label
        data_type_raw = normalize_header(config.get("data_type") or "连续")
        data_type = DATA_TYPE_ALIASES.get(data_type_raw)
        if not data_type:
            raise ProjectDataError(f"属性“{label}”的数据类型 {config.get('data_type')!r} 不支持。")

        parsed_values: List[Any] = []
        if data_type in {"continuous", "integer"}:
            for value in values_by_label[label]:
                number = optional_float(value, f"属性“{label}”的数据")
                assert number is not None
                if data_type == "integer" and abs(number - round(number)) > 1e-9:
                    raise ProjectDataError(f"属性“{label}”配置为整数，但存在值 {value!r}。")
                parsed_values.append(int(round(number)) if data_type == "integer" else number)
        else:
            parsed_values = [display_text(value) for value in values_by_label[label]]

        generation_min = optional_float(config.get("generation_min"), f"属性“{label}”生成下限")
        generation_max = optional_float(config.get("generation_max"), f"属性“{label}”生成上限")
        inferred_fields: List[str] = []
        design_stage = parse_design_stage(config.get("design_stage"))
        if design_stage is None:
            design_stage = 2
            inferred_fields.append("design_stage")
        if data_type in {"continuous", "integer"}:
            inferred_lo, inferred_hi = infer_numeric_range([float(value) for value in parsed_values])
            sample_min = min(float(value) for value in parsed_values)
            sample_max = max(float(value) for value in parsed_values)
            if generation_min is None:
                generation_min = inferred_lo
                inferred_fields.append("generation_min")
            if generation_max is None:
                generation_max = inferred_hi
                inferred_fields.append("generation_max")
            # The current feasible range is an empirical core, not a hard wall.
            # It always follows the observed sample extrema so sparse projects can
            # explore outside the sample cloud within the wider generation range.
            feasible_min = sample_min
            feasible_max = sample_max
            if generation_min >= generation_max:
                raise ProjectDataError(f"属性“{label}”生成下限必须小于生成上限。")
            if feasible_min < generation_min or feasible_max > generation_max:
                warnings.append(f"属性“{label}”的样本经验范围超出生成范围，请扩大生成范围。")
        else:
            feasible_min = None
            feasible_max = None

        preference_raw = normalize_header(config.get("preference_direction"))
        preference_direction = PREFERENCE_ALIASES.get(preference_raw)
        if preference_direction is None:
            raise ProjectDataError(f"属性“{label}”的偏好方向 {config.get('preference_direction')!r} 不支持。")
        if preference_direction == "unspecified":
            warnings.append(f"属性“{label}”缺少偏好方向，UTA 训练前必须确认。")
        if inferred_fields:
            warnings.append(f"属性“{label}”的 {', '.join(inferred_fields)} 由样本推断，当前仅为暂定值。")

        specs.append(
            AttributeSpec(
                key=key,
                label=label,
                unit=display_text(config.get("unit"), "-"),
                data_type=data_type,
                design_stage=design_stage,
                precision=parse_precision(config.get("precision")),
                generation_min=generation_min,
                generation_max=generation_max,
                feasible_min=feasible_min,
                feasible_max=feasible_max,
                preference_direction=preference_direction,
                marginal_trend=display_text(config.get("marginal_trend"), "待确认"),
                participates_utility=parse_bool(config.get("participates_utility"), True),
                participates_generation=parse_bool(config.get("participates_generation"), True),
                description=display_text(config.get("description")),
                inferred_fields=inferred_fields,
            )
        )
    return specs, warnings


def write_generation_ranges(path: Path | str, ranges: Dict[str, Tuple[float, float]]) -> Path:
    """Persist generation ranges to the attribute sheet and return a backup path."""
    workbook_path = Path(path).expanduser().resolve()
    if not ranges:
        raise ProjectDataError("没有收到需要保存的生成范围。")
    keep_vba = workbook_path.suffix.lower() == ".xlsm"
    try:
        workbook = load_workbook(workbook_path, keep_vba=keep_vba)
    except Exception as exc:
        raise ProjectDataError(f"无法打开 Excel 进行范围修改：{exc}") from exc
    temporary = workbook_path.with_name(f".{workbook_path.stem}.range_update{workbook_path.suffix}")
    try:
        if ATTRIBUTE_SHEET not in workbook.sheetnames:
            raise ProjectDataError(f"Excel 缺少“{ATTRIBUTE_SHEET}”工作表，无法回写生成范围。")
        sheet = workbook[ATTRIBUTE_SHEET]
        headers: Dict[str, int] = {}
        for column, value in enumerate(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)), start=1):
            canonical = ATTRIBUTE_HEADER_ALIASES.get(normalize_header(value))
            if canonical:
                headers[canonical] = column
        required = {"label", "key", "generation_min", "generation_max"}
        if not required.issubset(headers):
            missing = "、".join(sorted(required - set(headers)))
            raise ProjectDataError(f"“{ATTRIBUTE_SHEET}”缺少回写所需列：{missing}。")

        updated: set[str] = set()
        for row in range(2, sheet.max_row + 1):
            key = display_text(sheet.cell(row, headers["key"]).value)
            if key not in ranges:
                continue
            lo, hi = ranges[key]
            sheet.cell(row, headers["generation_min"], float(lo))
            sheet.cell(row, headers["generation_max"], float(hi))
            updated.add(key)
        missing_keys = sorted(set(ranges) - updated)
        if missing_keys:
            raise ProjectDataError(f"以下属性ID未在“{ATTRIBUTE_SHEET}”中找到：{'、'.join(missing_keys)}。")

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup = workbook_path.with_name(f"{workbook_path.stem}.before_range_update_{timestamp}{workbook_path.suffix}")
        shutil.copy2(workbook_path, backup)
        workbook.save(temporary)
        os.replace(temporary, workbook_path)
        return backup
    except ProjectDataError:
        raise
    except Exception as exc:
        raise ProjectDataError(f"生成范围写入 Excel 失败，请确认文件未被 Excel 占用：{exc}") from exc
    finally:
        workbook.close()
        if temporary.exists():
            try:
                temporary.unlink()
            except OSError:
                pass


def build_scheme_records(raw_rows: List[Dict[str, Any]], specs: List[AttributeSpec]) -> Tuple[List[SchemeRecord], List[str]]:
    label_to_spec = {item.label: item for item in specs}
    schemes: List[SchemeRecord] = []
    warnings: List[str] = []
    used_ids: set[str] = set()
    for index, row in enumerate(raw_rows, start=1):
        metadata = row["metadata"]
        scheme_id = display_text(metadata.get("id"), f"S-{index:03d}")
        if scheme_id in used_ids:
            raise ProjectDataError(f"方案编号“{scheme_id}”重复。")
        used_ids.add(scheme_id)
        params: Dict[str, Any] = {}
        for label, raw_value in row["params"].items():
            spec = label_to_spec[label]
            if spec.data_type == "continuous":
                value: Any = optional_float(raw_value, f"第 {row['excel_row']} 行属性“{label}”")
            elif spec.data_type == "integer":
                numeric = optional_float(raw_value, f"第 {row['excel_row']} 行属性“{label}”")
                assert numeric is not None
                value = int(round(numeric))
            else:
                value = display_text(raw_value)
            params[spec.key] = value
            if spec.is_numeric and value is not None:
                assert spec.generation_min is not None and spec.generation_max is not None
                if float(value) < spec.generation_min or float(value) > spec.generation_max:
                    warnings.append(
                        f"方案 {scheme_id} 的“{label}”={value} 超出生成范围 [{spec.generation_min}, {spec.generation_max}]。"
                    )
        schemes.append(
            SchemeRecord(
                id=scheme_id,
                source=display_text(metadata.get("source"), "已有样本"),
                known_feasibility=display_text(metadata.get("known_feasibility"), "未标注"),
                notes=display_text(metadata.get("notes")),
                params=params,
                excel_row=int(row["excel_row"]),
            )
        )
    return schemes, warnings


def ensure_acyclic(attributes: List[AttributeSpec], couplings: List[CouplingSpec]) -> None:
    nodes = {item.key for item in attributes}
    outgoing: Dict[str, List[str]] = {node: [] for node in nodes}
    indegree = {node: 0 for node in nodes}
    for edge in couplings:
        outgoing[edge.source_key].append(edge.target_key)
        indegree[edge.target_key] += 1
    queue = [node for node, degree in indegree.items() if degree == 0]
    visited = 0
    while queue:
        node = queue.pop()
        visited += 1
        for target in outgoing[node]:
            indegree[target] -= 1
            if indegree[target] == 0:
                queue.append(target)
    if visited != len(nodes):
        cyclic_labels = [item.label for item in attributes if indegree[item.key] > 0]
        raise ProjectDataError(f"耦合关系存在有向环，当前序列生成器无法确定先后顺序：{'、'.join(cyclic_labels)}。")


def read_couplings(workbook: Any, specs: List[AttributeSpec]) -> Tuple[List[CouplingSpec], List[str]]:
    warnings: List[str] = []
    if COUPLING_SHEET not in workbook.sheetnames:
        warnings.append(f"缺少“{COUPLING_SHEET}”工作表，系统将按属性独立处理。")
        return [], warnings
    by_label = {normalize_header(item.label): item for item in specs}
    couplings: List[CouplingSpec] = []
    seen: set[Tuple[str, str]] = set()
    for excel_row, record in row_dicts(workbook[COUPLING_SHEET], COUPLING_HEADER_ALIASES):
        source_label = display_text(record.get("source"))
        target_label = display_text(record.get("target"))
        if not source_label or not target_label:
            raise ProjectDataError(f"“{COUPLING_SHEET}”第 {excel_row} 行必须同时填写源属性和目标属性。")
        source = by_label.get(normalize_header(source_label))
        target = by_label.get(normalize_header(target_label))
        if not source:
            raise ProjectDataError(f"“{COUPLING_SHEET}”第 {excel_row} 行引用了未知源属性“{source_label}”。")
        if not target:
            raise ProjectDataError(f"“{COUPLING_SHEET}”第 {excel_row} 行引用了未知目标属性“{target_label}”。")
        if source.key == target.key:
            raise ProjectDataError(f"属性“{source.label}”不能耦合到自身。")
        edge_key = (source.key, target.key)
        if edge_key in seen:
            raise ProjectDataError(f"耦合边“{source.label} -> {target.label}”重复。")
        seen.add(edge_key)
        direction_raw = normalize_header(record.get("direction"))
        direction = DIRECTION_ALIASES.get(direction_raw)
        if not direction:
            raise ProjectDataError(f"“{COUPLING_SHEET}”第 {excel_row} 行方向 {record.get('direction')!r} 不支持。")
        coefficient = optional_float(record.get("coefficient_prior"), f"第 {excel_row} 行先验系数")
        if coefficient is not None:
            if direction == "positive" and coefficient < 0:
                raise ProjectDataError(f"“{source.label} -> {target.label}”是正向关系，但先验系数为负。")
            if direction == "negative" and coefficient > 0:
                raise ProjectDataError(f"“{source.label} -> {target.label}”是负向关系，但先验系数为正。")
        couplings.append(
            CouplingSpec(
                source_key=source.key,
                source_label=source.label,
                target_key=target.key,
                target_label=target.label,
                direction=direction,
                relation_type=display_text(record.get("relation_type"), "monotonic"),
                coefficient_prior=coefficient,
                status=display_text(record.get("status"), "待确认"),
                description=display_text(record.get("description")),
            )
        )
    ensure_acyclic(specs, couplings)
    by_key = {item.key: item for item in specs}
    for edge in couplings:
        source = by_key[edge.source_key]
        target = by_key[edge.target_key]
        if source.design_stage > target.design_stage:
            raise ProjectDataError(
                f"耦合边“{source.label} -> {target.label}”与设计顺序冲突："
                f"第 {source.design_stage} 批属性不能反向决定第 {target.design_stage} 批属性。"
            )
    return couplings, warnings


def read_requirement_profiles(
    workbook: Any,
    specs: List[AttributeSpec],
) -> Tuple[List[RequirementProfile], List[str]]:
    """Read protocol reference vectors, with legacy sheet compatibility."""
    if PROTOCOL_SHEET in workbook.sheetnames:
        return read_protocol_profiles(workbook, specs)
    return read_legacy_requirement_profiles(workbook, specs)


def read_protocol_profiles(
    workbook: Any,
    specs: List[AttributeSpec],
) -> Tuple[List[RequirementProfile], List[str]]:
    """Read one protocol per row; attribute directions stay in 属性配置."""
    sheet = workbook[PROTOCOL_SHEET]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], [f"“{PROTOCOL_SHEET}”为空，将继续使用通用效能模式。"]
    headers = [display_text(value) for value in rows[0]]
    normalized = [normalize_header(value) for value in headers]
    if any(not value for value in normalized):
        raise ProjectDataError(f"“{PROTOCOL_SHEET}”第一行不能有空白表头。")
    if len(set(normalized)) != len(normalized):
        raise ProjectDataError(f"“{PROTOCOL_SHEET}”第一行存在重复表头。")

    by_label = {normalize_header(item.label): item for item in specs}
    metadata_columns: Dict[str, int] = {}
    attribute_columns: List[Tuple[int, AttributeSpec]] = []
    for index, (header, normalized_header) in enumerate(zip(headers, normalized)):
        metadata_key = PROTOCOL_METADATA_HEADERS.get(normalized_header)
        if metadata_key:
            metadata_columns[metadata_key] = index
            continue
        spec = by_label.get(normalized_header)
        if spec is None:
            raise ProjectDataError(
                f"“{PROTOCOL_SHEET}”表头“{header}”既不是协议说明列，也不是“{DATA_SHEET}”中的属性。"
            )
        if not spec.is_numeric:
            raise ProjectDataError(f"“{PROTOCOL_SHEET}”暂不支持类别属性“{spec.label}”。")
        attribute_columns.append((index, spec))

    expected = {item.key for item in specs if item.is_numeric and item.participates_utility}
    provided = {item.key for _, item in attribute_columns}
    missing = [item.label for item in specs if item.key in expected - provided]
    if missing:
        raise ProjectDataError(
            f"“{PROTOCOL_SHEET}”缺少参与效能属性：{'、'.join(missing)}。"
        )

    profiles: List[RequirementProfile] = []
    seen_ids: set = set()
    for excel_row, values in enumerate(rows[1:], start=2):
        if all(value is None or str(value).strip() == "" for value in values):
            continue
        profile_id = display_text(
            values[metadata_columns["profile_id"]]
            if "profile_id" in metadata_columns and metadata_columns["profile_id"] < len(values)
            else None,
            f"PROTOCOL-{len(profiles) + 1:03d}",
        )
        if not re.fullmatch(r"[A-Za-z0-9_-]+", profile_id):
            raise ProjectDataError(
                f"“{PROTOCOL_SHEET}”第 {excel_row} 行协议编号 {profile_id!r} 无效，"
                "只能使用英文字母、数字、下划线和连字符。"
            )
        if profile_id in seen_ids:
            raise ProjectDataError(f"“{PROTOCOL_SHEET}”协议编号“{profile_id}”重复。")
        seen_ids.add(profile_id)
        profile_name = display_text(
            values[metadata_columns["profile_name"]]
            if "profile_name" in metadata_columns and metadata_columns["profile_name"] < len(values)
            else None,
            profile_id,
        )
        description = display_text(
            values[metadata_columns["description"]]
            if "description" in metadata_columns and metadata_columns["description"] < len(values)
            else None
        )
        requirements: List[RequirementSpec] = []
        for column, spec in attribute_columns:
            raw_value = values[column] if column < len(values) else None
            target_value = optional_float(raw_value, f"“{PROTOCOL_SHEET}”第 {excel_row} 行“{spec.label}”")
            if target_value is None:
                if spec.participates_utility:
                    raise ProjectDataError(
                        f"“{PROTOCOL_SHEET}”第 {excel_row} 行缺少属性“{spec.label}”的协议要求值。"
                    )
                continue
            if spec.preference_direction not in {"higher_better", "lower_better", "interval"}:
                if spec.participates_utility:
                    raise ProjectDataError(
                        f"属性“{spec.label}”参与效能，但“属性配置”的偏好方向未明确。"
                    )
                continue
            requirements.append(
                RequirementSpec(
                    attribute_key=spec.key,
                    attribute_label=spec.label,
                    requirement_type=(
                        "higher_better"
                        if spec.preference_direction == "higher_better"
                        else "lower_better"
                        if spec.preference_direction == "lower_better"
                        else "target"
                    ),
                    target_value=target_value,
                    minimum=target_value if spec.preference_direction == "higher_better" else None,
                    maximum=target_value if spec.preference_direction == "lower_better" else None,
                    description=description,
                )
            )
        profiles.append(
            RequirementProfile(
                id=profile_id,
                name=profile_name,
                requirements=requirements,
                direct_reuse_threshold=1.0,
                improvement_threshold=0.80,
                redesign_threshold=0.60,
            )
        )
    warnings: List[str] = []
    if not profiles:
        warnings.append(f"“{PROTOCOL_SHEET}”没有有效协议行，将继续使用通用效能模式。")
    return profiles, warnings


def read_legacy_requirement_profiles(
    workbook: Any,
    specs: List[AttributeSpec],
) -> Tuple[List[RequirementProfile], List[str]]:
    """Read the version-9 vertical sheet without using it to train BT/UTA."""
    if LEGACY_REQUIREMENT_SHEET not in workbook.sheetnames:
        return [], []
    warnings: List[str] = [
        f"检测到旧版“{LEGACY_REQUIREMENT_SHEET}”表；仍可读取，但建议改为一行一份协议的“{PROTOCOL_SHEET}”表。"
    ]
    by_label = {normalize_header(item.label): item for item in specs}
    grouped: Dict[str, Dict[str, Any]] = {}
    profile_order: List[str] = []
    for excel_row, record in row_dicts(workbook[LEGACY_REQUIREMENT_SHEET], REQUIREMENT_HEADER_ALIASES):
        profile_id = display_text(record.get("profile_id"), "default")
        if not re.fullmatch(r"[A-Za-z0-9_-]+", profile_id):
            raise ProjectDataError(
                f"“{LEGACY_REQUIREMENT_SHEET}”第 {excel_row} 行基准编号 {profile_id!r} 无效，"
                "只能使用英文字母、数字、下划线和连字符。"
            )
        profile_name = display_text(record.get("profile_name"), profile_id)
        label = display_text(record.get("label"))
        if not label:
            raise ProjectDataError(f"“{LEGACY_REQUIREMENT_SHEET}”第 {excel_row} 行缺少属性名。")
        spec = by_label.get(normalize_header(label))
        if not spec:
            raise ProjectDataError(
                f"“{LEGACY_REQUIREMENT_SHEET}”第 {excel_row} 行引用了未知属性“{label}”。"
            )
        if not spec.is_numeric:
            raise ProjectDataError(f"“{LEGACY_REQUIREMENT_SHEET}”第 {excel_row} 行暂不支持类别属性“{label}”。")
        requirement_raw = normalize_header(record.get("requirement_type"))
        requirement_type = REQUIREMENT_TYPE_ALIASES.get(requirement_raw)
        if not requirement_type:
            raise ProjectDataError(
                f"“{LEGACY_REQUIREMENT_SHEET}”第 {excel_row} 行要求类型 "
                f"{record.get('requirement_type')!r} 不支持。"
            )
        target_value = optional_float(record.get("target_value"), f"第 {excel_row} 行要求值")
        minimum = optional_float(record.get("minimum"), f"第 {excel_row} 行要求下限")
        maximum = optional_float(record.get("maximum"), f"第 {excel_row} 行要求上限")
        tolerance = optional_float(record.get("tolerance"), f"第 {excel_row} 行容差") or 0.0
        if tolerance < 0.0:
            raise ProjectDataError(f"“{LEGACY_REQUIREMENT_SHEET}”第 {excel_row} 行容差不能为负数。")
        if requirement_type == "at_least":
            minimum = minimum if minimum is not None else target_value
            if minimum is None:
                raise ProjectDataError(f"属性“{label}”的“达到下限即可”必须填写要求下限或要求值。")
        elif requirement_type == "at_most":
            maximum = maximum if maximum is not None else target_value
            if maximum is None:
                raise ProjectDataError(f"属性“{label}”的“不超过上限即可”必须填写要求上限或要求值。")
        elif requirement_type == "within_range":
            if minimum is None or maximum is None:
                raise ProjectDataError(f"属性“{label}”的“落入区间即可”必须同时填写要求下限和上限。")
            if minimum > maximum:
                raise ProjectDataError(f"属性“{label}”的要求下限不能大于要求上限。")
        elif requirement_type == "target" and target_value is None:
            raise ProjectDataError(f"属性“{label}”的“接近目标值最好”必须填写要求值。")

        if profile_id not in grouped:
            profile_order.append(profile_id)
            grouped[profile_id] = {
                "name": profile_name,
                "requirements": [],
                "seen": set(),
                "direct": parse_fraction(
                    record.get("direct_reuse_threshold"), f"基准“{profile_name}”直接复用阈值", 0.95
                ),
                "improve": parse_fraction(
                    record.get("improvement_threshold"), f"基准“{profile_name}”改进复用阈值", 0.80
                ),
                "redesign": parse_fraction(
                    record.get("redesign_threshold"), f"基准“{profile_name}”重新研制阈值", 0.60
                ),
            }
        profile = grouped[profile_id]
        if spec.key in profile["seen"]:
            raise ProjectDataError(f"“{LEGACY_REQUIREMENT_SHEET}”中基准“{profile_name}”的属性“{label}”重复。")
        profile["seen"].add(spec.key)
        profile["requirements"].append(
            RequirementSpec(
                attribute_key=spec.key,
                attribute_label=spec.label,
                requirement_type=requirement_type,
                target_value=target_value,
                minimum=minimum,
                maximum=maximum,
                tolerance=tolerance,
                hard_requirement=parse_bool(record.get("hard_requirement"), False),
                overachievement_bonus=parse_bool(record.get("overachievement_bonus"), False),
                description=display_text(record.get("description")),
            )
        )

    profiles: List[RequirementProfile] = []
    for profile_id in profile_order:
        data = grouped[profile_id]
        if not data["requirements"]:
            continue
        if not data["redesign"] <= data["improve"] <= data["direct"]:
            raise ProjectDataError(
                f"评价基准“{data['name']}”的阈值必须满足：重新研制阈值 <= 改进复用阈值 <= 直接复用阈值。"
            )
        profiles.append(
            RequirementProfile(
                id=profile_id,
                name=data["name"],
                requirements=data["requirements"],
                direct_reuse_threshold=data["direct"],
                improvement_threshold=data["improve"],
                redesign_threshold=data["redesign"],
            )
        )
    if LEGACY_REQUIREMENT_SHEET in workbook.sheetnames and not profiles:
        warnings.append(f"“{LEGACY_REQUIREMENT_SHEET}”没有有效要求行，将继续使用旧版通用效能模式。")
    return profiles, warnings


def load_project_workbook(path: Path | str) -> ProjectDataset:
    workbook_path = Path(path).expanduser().resolve()
    if not workbook_path.exists():
        raise ProjectDataError(f"Excel 文件不存在：{workbook_path}")
    if workbook_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ProjectDataError("当前仅支持 .xlsx 或 .xlsm 工作簿。")
    try:
        workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    except Exception as exc:
        raise ProjectDataError(f"无法读取 Excel：{exc}") from exc
    try:
        if DATA_SHEET not in workbook.sheetnames:
            raise ProjectDataError(f"缺少必需工作表“{DATA_SHEET}”。")

        product_code, project_name, identity_warnings = read_project_identity(
            workbook, workbook_path.stem
        )
        _, attribute_columns, metadata_columns = read_data_headers(workbook[DATA_SHEET])
        configs, config_warnings = read_attribute_config(workbook)
        raw_rows, values_by_label = read_raw_data(workbook[DATA_SHEET], attribute_columns, metadata_columns)
        specs, spec_warnings = build_attribute_specs(attribute_columns, configs, values_by_label)
        schemes, scheme_warnings = build_scheme_records(raw_rows, specs)
        couplings, coupling_warnings = read_couplings(workbook, specs)
        requirement_profiles, requirement_warnings = read_requirement_profiles(workbook, specs)
    finally:
        workbook.close()
    return ProjectDataset(
        product_code=product_code,
        project_name=project_name,
        workbook_path=str(workbook_path),
        workbook_fingerprint=workbook_fingerprint(workbook_path),
        learning_fingerprint=learning_fingerprint(specs, schemes, couplings),
        attributes=specs,
        schemes=schemes,
        couplings=couplings,
        requirement_profiles=requirement_profiles,
        warnings=(
            identity_warnings
            + config_warnings
            + spec_warnings
            + scheme_warnings
            + coupling_warnings
            + requirement_warnings
        ),
    )


def validation_text(project: ProjectDataset) -> str:
    lines = [
        f"项目：{project.project_name}",
        f"属性：{len(project.attributes)} 个",
        f"方案：{len(project.schemes)} 个",
        f"耦合：{len(project.couplings)} 条",
        f"新技术协议：{len(project.requirement_profiles)} 个",
        f"指纹：{project.workbook_fingerprint}",
        "属性顺序：" + "、".join(item.label for item in project.attributes),
    ]
    if project.warnings:
        lines.append("警告：")
        lines.extend(f"- {item}" for item in project.warnings)
    else:
        lines.append("校验通过，无警告。")
    return "\n".join(lines)


def main() -> None:
    parser = argparse.ArgumentParser(description="Validate and inspect a parameter-project Excel workbook.")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--json", action="store_true", help="Print a JSON summary.")
    parser.add_argument("--dump", type=Path, help="Write the normalized complete project data to JSON.")
    args = parser.parse_args()
    try:
        project = load_project_workbook(args.workbook)
    except ProjectDataError as exc:
        raise SystemExit(f"Excel 校验失败：{exc}") from exc
    print(json.dumps(project.summary(), ensure_ascii=False, indent=2) if args.json else validation_text(project))
    if args.dump:
        args.dump.parent.mkdir(parents=True, exist_ok=True)
        args.dump.write_text(json.dumps(project.to_dict(), ensure_ascii=False, indent=2), encoding="utf-8")


if __name__ == "__main__":
    main()
